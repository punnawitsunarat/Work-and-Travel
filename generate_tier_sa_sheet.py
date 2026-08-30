import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load existing workbook
excel_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
excel_clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
except Exception:
    wb = openpyxl.Workbook()

sheet_name = 'Tier S-A Jobs & 2nd Options'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(title=sheet_name)
ws.views.sheetView[0].showGridLines = True

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid') # Amber tint
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid') # Sky blue tint
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

headers = [
    'Tier & รัฐ (State)',
    'ชื่องาน / สถานที่ทำงานหลัก (Employer)',
    'Agency ที่ถือสัญญา',
    'ตำแหน่งงานหลัก (Role)',
    '⏱️ เวลาทำงานหลัก (ปกติ / พีค)',
    'ค่าแรงฐาน / OT ($/ชม.)',
    'ทิป ($/สัปดาห์)',
    'ชั่วโมงงานหลัก (ชม./wk)',
    'ค่าที่พัก ($/สัปดาห์)',
    'สวัสดิการ (อาหาร ฟรี หรือ อื่นๆ)',
    '🔥 งานที่สองที่ทำต่อได้จริง (ตัวเลือกที่ 1)',
    '🔥 งานที่สองที่ทำต่อได้จริง (ตัวเลือกที่ 2)',
    '🔥 งานที่สองที่ทำต่อได้จริง (ตัวเลือกที่ 3)'
]

rows = [
    # ==========================================
    # 🏔️ 1. ALASKA (Tier S - 0% State Tax)
    # ==========================================
    [
        'Tier S - Alaska (AK)',
        'Denali Princess Wilderness Lodge (Denali NP)',
        'Acadex, OEG, IEE, Higher',
        'Housekeeping / Room Attendant',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $16.00 / OT $24.00',
        '$15 - $30',
        '50 – 54 ชม./wk',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน, ทิปหัวเตียง, รถรับส่งพนักงาน, ส่วนลดทัวร์ 50%',
        '🚌 Fleet Detailer (ในโรงแรม): ล้างรถทัวร์กะดึก 21:00-01:30 น. ได้เรต OT $24.00/ชม. ทันที (ไม่ต้องขอสปอนเซอร์)',
        '🍕 Prospectors Pizzeria: เดิน 10 นาที ล้างจาน/Barback 18:30-00:30 น. ($15/ชม. + ทิปสด $40-$70/คืน)',
        '🍽️ Fannie Q’s Saloon (ในโรงแรม): ช่วยจัดเลี้ยง/Runner 17:00-22:00 น. ได้เรต OT $24/ชม. + ทิป'
    ],
    [
        'Tier S - Alaska (AK)',
        'Denali Princess Wilderness Lodge (Denali NP)',
        'Acadex, OEG, IEE, Higher',
        'Kitchen Stewarding / Dishwasher',
        'ปกติ: 16:00 – 00:00 (8 ชม.)\nพีค: 15:00 – 01:30 (10.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 55 ชม./wk',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อใน EDR + เชฟทำสเต๊ก/เบอร์เกอร์เลี้ยงในครัวฟรีทุกกะ',
        '🧹 เช้าช่วยแม่บ้าน (Housekeeping): 08:30-14:30 น. ขอกะเช้าเพิ่มในโรงแรม ได้เรต OT $24/ชม.',
        '🧳 Luggage Handler: ช่วยยกกระเป๋ารอบรถไฟเช้า 08:00-12:00 น. ได้ทิปสดจากทัวร์',
        '🛍️ Denali Boardwalk Gift Shops: แคชเชียร์/จัดของกะสาย 09:00-14:00 น. ($15/ชม.)'
    ],
    [
        'Tier S - Alaska (AK)',
        'Grande Denali Lodge & Denali Bluffs Hotel',
        'Acadex Thailand',
        'Housekeeping / Public Area',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 46 ชม./wk',
        '$140',
        'โบนัส +$1.00/ชม. ท้ายทริป, โรงแรมหรู 4 ดาวบนยอดเขา, รถชัตเติลบัส, ส่วนลดอาหาร',
        '🍷 Alpenglow Restaurant (ในโรงแรม): Busser 17:00-23:00 น. ($14/ชม. + ทิปสดเศรษฐี $50-$80/คืน)',
        '🥪 Subway Denali / The Black Bear: นั่งชัตเติลลงเขาไปทำกะเย็น 17:00-22:00 น. ($15/ชม.)',
        '🍽️ Denali Princess Kitchens: เดินลงเขา 15 นาที ล้างจานกะค่ำ 18:00-00:00 น. ($16/ชม.)'
    ],
    [
        'Tier S - Alaska (AK)',
        'Mt. McKinley Princess Lodge (Trapper Creek)',
        'Higher, OEG, Acadex',
        'Housekeeping / Kitchen',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$10 - $20',
        '44 – 48 ชม./wk',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, รถรับส่ง, ชมวิวยอดเขาเดนาลี บรรยากาศสงบ',
        '🥩 20,320 Alaskan Grill (ในโรงแรม): Busser/Steward กะค่ำ 17:30-22:30 น. (เรต OT $24/ชม. + ทิป)',
        '🍕 Northland Wood Fired Pizza (ในโรงแรม): ผู้ช่วยทำพิซซ่า/ครัว 16:30-21:30 น. (เรต OT $24/ชม.)',
        '🥐 Talkeetna Roadhouse: นั่งรถตู้พนักงานเข้าเมือง Talkeetna ช่วยล้างจาน/ทำขนมวันหยุด'
    ],
    [
        'Tier S - Alaska (AK)',
        'Fairbanks Princess Riverside Lodge',
        'IEE, Acadex, Higher',
        'Food Runner / Laundry / Steward',
        'ปกติ: 06:30 – 14:30 หรือ 15:00 – 23:00\nพีค: กะสลับตามรอบรถไฟ 8 ชม.',
        'ฐาน $15.50 / OT $23.25',
        '$20 - $40',
        '36 – 40 ชม./wk',
        '$105',
        'หอพักในเมืองแฟร์แบงก์ส + รถชัตเติลบัส, ใกล้ Walmart และร้านอาหารไทย หางาน 2 ง่าย',
        '🛒 Walmart Supercenter Fairbanks: จัดสต็อกสินค้า/แคชเชียร์ 16:30-22:30 น. ($16-$17/ชม. มีรถเมล์ผ่าน)',
        '🍜 Pad Thai Restaurant / Thai House: ผู้ช่วยครัว/เสิร์ฟ 17:00-22:30 น. (ทิปสด + กินอาหารไทยฟรี)',
        '🍺 The Pump House Restaurant: Busser/Dishwasher ร้านอาหารริมน้ำดัง 17:30-23:00 น. ($15/ชม. + ทิป)'
    ],
    [
        'Tier S - Alaska (AK)',
        'Seward Windsong Lodge & Kenai Fjords Tours (Pursuit)',
        'OEG, Acadex, ALC',
        'Housekeeping / Tour Operations / F&B',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$20 - $50',
        '44 – 50 ชม./wk',
        '$120',
        'เมืองตากอากาศชายทะเล Seward ท่าเรือล่องเรือชมวาฬและธารน้ำแข็ง, หอพักพนักงาน, ล่องเรือฟรี',
        '🦀 Ray’s Waterfront (ริมท่าเรือ Seward): Busser/Steward ร้านซีฟู้ดดัง 17:30-23:00 น. ($14/ชม. + ทิปสด $60-$100)',
        '🍺 Seward Brewing Company: Barback/Dishwasher กะค่ำ 18:00-00:00 น. ($15/ชม. + ทิป)',
        '🛥️ Kenai Fjords Boat Cleaning: ล้างทำความสะอาดเรือทัวร์รอบเย็น 18:00-21:30 น. ($16.50/ชม.)'
    ],
    [
        'Tier S - Alaska (AK)',
        'Talkeetna Alaskan Lodge (Pursuit Collection)',
        'OEG, Acadex, Higher',
        'Housekeeping / Kitchen / Steward',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$15 - $30',
        '44 – 50 ชม./wk',
        '$115',
        'โรงแรมหรูชมวิวยอดเขา Denali ในเมืองคาวบอย Talkeetna, หอพักพนักงาน + รถรับส่ง',
        '🥩 Foraker Dining Room (ในโรงแรม): Busser ร้านอาหาร Fine Dining 17:30-23:00 น. (ทิปสด $50-$90/คืน)',
        '🍕 Mountain High Pizza Pie: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น. ($15/ชม. + ทิป)',
        '🍺 Denali Brewing Company (Talkeetna Pub): Barback กะค่ำ 18:30-00:30 น. ($14/ชม. + ทิปสด)'
    ],
    [
        'Tier S - Alaska (AK)',
        'Alyeska Resort & Hotel Alyeska (Girdwood)',
        'OEG, Higher',
        'Housekeeping / Culinary / Lift Ops',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '-',
        '42 – 48 ชม./wk',
        '$135',
        'รีสอร์ตหรูระดับโลก นั่งกระเช้าชมวิวภูเขาและธารน้ำแข็ง, หอพักพนักงาน Glacier Valley',
        '🥩 Seven Glaciers Restaurant (บนยอดเขา): Busser 17:30-23:00 น. (ร้าน 4 ดาว AAA ทิปมหาศาล $70-$120)',
        '🍕 Coast Pizza / Girdwood Brewing: ผู้ช่วยครัว/Barback 18:00-23:00 น. ($15/ชม. + ทิป)',
        '🥖 The Bake Shop Girdwood: ช่วยเตรียมเบเกอรี่/ล้างจานช่วงเช้าตรู่หรือเย็น'
    ],
    [
        'Tier S - Alaska (AK)',
        'Copper River Princess Wilderness Lodge',
        'Higher, IEE, Acadex',
        'Housekeeping / Kitchen / Laundry',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$10 - $20',
        '42 – 48 ชม./wk',
        '$105',
        'รีสอร์ตริมแม่น้ำ Copper River อุทยาน Wrangell-St. Elias, อาหาร EDR ฟรี 3 มื้อ',
        '🐟 Two Rivers Salmon Grill (ในโรงแรม): Busser/Steward 17:30-22:30 น. (เรต OT $24/ชม. + ทิป)',
        '🛶 Copper Oar Rafting Base: ช่วยล้างทำความสะอาดเรือยาง/ชูชีพ 17:00-21:00 น. ($16/ชม.)',
        '🌲 Whistle Stop Gift Shop: แคชเชียร์ร้านของฝากกะค่ำ 18:00-22:00 น.'
    ],

    # ==========================================
    # 🌲 2. WYOMING (Tier S - 0% State Tax)
    # ==========================================
    [
        'Tier S - Wyoming (WY)',
        'Grand Teton Lodge Company - Jackson Lake Lodge',
        'Acadex, Higher, OEG, IEE',
        'Housekeeping / Crew Member',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '54 – 60 ชม./wk',
        '$0 (ฟรี!)',
        'หอพักฟรี 100%! + อาหารบุฟเฟต์ 3 มื้อ EDR ($105/wk), ส่วนลด 40%, บัตรผ่าน 2 อุทยานฟรี',
        '🍽️ Mural Room & Pioneer Grill (ในโรงแรม): Busser 17:30-22:30 น. ได้เรต OT $26.62/ชม. + ทิป',
        '🍸 Blue Heron Lounge (ในโรงแรม): Barback ช่วยยกน้ำแข็ง/ล้างแก้ว 18:00-23:30 น. (เรต OT $26.62 + Tip-Out)',
        '🛶 Colter Bay Chuckwagon & Marina: นั่งชัตเติลฟรีไปช่วยมินิมาร์ท/เรือแคนู 17:00-21:30 น.'
    ],
    [
        'Tier S - Wyoming (WY)',
        'Grand Teton Lodge Company - Colter Bay Village',
        'Acadex, Higher, OEG, IEE',
        'Cabin Attendant / Grocery Clerk / Marina',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '52 – 58 ชม./wk',
        '$0 (ฟรี!)',
        'หอพักฟรี 100%! + อาหารบุฟเฟต์ 3 มื้อ EDR ($105/wk), หมู่บ้านริมทะเลสาบ Jackson Lake',
        '🥩 Chuckwagon Restaurant (ในหมู่บ้าน): Busser/Steward บุฟเฟต์มื้อเย็น 17:00-22:00 น. (เรต OT $26.62)',
        '🛒 Colter Bay General Store: แคชเชียร์/เติมสต็อกมินิมาร์ท 17:30-22:00 น. (เรต OT $26.62)',
        '🛶 Colter Bay Marina: ล้างเรือยนต์/เรือแคนู 16:30-20:30 น. (เรต OT $26.62)'
    ],
    [
        'Tier S - Wyoming (WY)',
        'Signal Mountain Lodge (Grand Teton NP - Forever Resorts)',
        'Higher, Acadex',
        'Housekeeping / Front Desk / Marina',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '$15 - $30',
        '45 – 50 ชม./wk',
        '$110 (รวมกิน)',
        'รีสอร์ตริมทะเลสาบ Signal Mountain, หอพักพนักงาน + อาหาร 3 มื้อ, ส่วนลดเช่าเรือ',
        '🥩 The Peaks Restaurant (ในรีสอร์ต): Busser/Runner 17:30-22:30 น. ($14/ชม. + ทิปสด $50-$90/คืน)',
        '🌮 Trapper Grill (ในรีสอร์ต): ผู้ช่วยครัว/ทำนาโชส์ 17:00-22:00 น. (ขอ OT $24.75/ชม.)',
        '⛽ Signal Mountain Gas Station & Store: แคชเชียร์รอบค่ำ 17:00-21:30 น.'
    ],
    [
        'Tier S - Wyoming (WY)',
        'Xanterra Yellowstone Lodges (Old Faithful / Canyon)',
        'OEG, IEE Thailand',
        'Hospitality Crew / Kitchen Steward',
        'ปกติ: 07:00 – 15:30 หรือ 15:00 – 23:30\nพีค: 06:30 – 16:30 (10 ชม.)',
        'ฐาน $15.70 / OT $23.55',
        '-',
        '46 – 52 ชม./wk',
        '$120 (รวมกิน)',
        'รวมหอพัก + อาหารบุฟเฟต์ 3 มื้อ EDR, หอพักใหม่ Arnica, เที่ยวบ่อน้ำพุร้อน Yellowstone ฟรี',
        '🌋 Old Faithful Snow Lodge Dining: Busser/Steward กะค่ำ 17:00-22:30 น. (เรต OT $23.55/ชม. + ทิป)',
        '🛍️ Delaware North General Stores: แคชเชียร์ร้านของฝากข้างๆ เดิน 3 นาที 17:00-21:30 น. ($15.50/ชม.)',
        '🍻 Employee Pub & Rec Hall: ผู้ช่วยผับพนักงาน/ทำความสะอาด 19:30-23:30 น.'
    ],
    [
        'Tier S - Wyoming (WY)',
        'Four Seasons Resort & Snake River Lodge (Teton Village)',
        'OEG, Acadex',
        'Housekeeping / Stewarding',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.50 / OT $26.25',
        '-',
        '45 – 50 ชม./wk',
        '$140',
        'รีสอร์ตหรู 5 ดาวใน Teton Village, หอพักพนักงาน, นั่งรถบัส START Bus ฟรี',
        '🍺 The Mangy Moose Saloon (Teton Village): Barback/Busser ผับดัง 18:00-00:30 น. (ทิปสดเศรษฐี $60-$120/คืน)',
        '🤠 Million Dollar Cowboy Bar (เมือง Jackson): นั่ง START Bus เข้าเมือง ทำ Barback 19:00-01:30 น.',
        '🍕 Alpenhof Bistro: ล้างจาน/ผู้ช่วยครัว 17:30-22:30 น. ($16/ชม. + อาหารฟรี)'
    ],
    [
        'Tier S - Wyoming (WY)',
        'The Wort Hotel & Silver Dollar Bar (Jackson Town)',
        'Acadex, OEG',
        'Housekeeping / Banquet Server / Steward',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $17.00 / OT $25.50',
        '$30 - $60',
        '44 – 50 ชม./wk',
        '$150',
        'โรงแรมประวัติศาสตร์ใจกลางเมือง Jackson, เดินทางสะดวกที่สุด มีร้านค้าล้อมรอบ',
        '💵 Silver Dollar Bar (ในโรงแรม): Barback/Busser บาร์ดนตรีสดคนแน่น 18:30-01:00 น. (ทิปสด $70-$130/คืน)',
        '🥩 Gun Barrel Steak & Game House: Busser ร้านสเต๊กเนื้อสัตว์ป่า 17:30-22:30 น. ($14/ชม. + ทิป)',
        '🛒 Albertsons Jackson: จัดสต็อกสินค้ากะดึก 20:00-00:00 น. ($18/ชม.)'
    ],
    [
        'Tier S - Wyoming (WY)',
        'Under Canvas Grand Teton / Yellowstone',
        'Higher Education',
        'Housekeeping / Guest Service',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '$50 - $100',
        '42 – 46 ชม./wk',
        '$100',
        'แคมป์กระโจมซาฟารีหรู, อาหารพนักงาน, ทิปเงินสดจากแขกไฮเอนด์ บรรยากาศเป็นกันเอง',
        '🏕️ Embers Restaurant (ในแกลมปิ้ง): Food Runner กะค่ำ 17:30-22:00 น. (ได้ทิปสดคืนละ $40-$60)',
        '☕ Moose-Wilson Road Cafes: ผู้ช่วยบาริสต้า/เบเกอรี่ช่วงเช้าตรู่ 06:00-08:00 น. หรือเย็น',
        '🛒 Jackson Town Local Diners: นั่งรถเข้าเมืองแจ็กสัน ทำงานร้านอาหารวันหยุด'
    ],

    # ==========================================
    # 🎡 3. WISCONSIN (Tier S - เมืองหลวง 2 งาน)
    # ==========================================
    [
        'Tier S - Wisconsin (WI)',
        'Kalahari Resorts & Conventions (Wisconsin Dells)',
        'OEG, Acadex, Higher',
        'Lifeguard / Housekeeping',
        'ปกติ: 09:30 – 18:00 หรือ 12:00 – 20:30\nพีค: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '38 – 44 ชม./wk',
        '$105',
        'หอพัก Kalahari Village, บัตรเล่นสวนน้ำฟรี, เมืองปั่นจักรยาน 100%, หางาน 2 ง่ายที่สุดในอเมริกา',
        '🍕 Moosejaw Pizza & Dells Brewing: ปั่นจักรยาน 10 นาที Busser/Barback 18:30-23:30 น. ($13/ชม. + ทิปสด $60-$90/คืน)',
        '🍔 Buffalo Phil’s Grille: Food Runner/Busser เสิร์ฟอาหารด้วยรถไฟจิ๋ว 18:00-22:30 น. ($13/ชม. + ทิปสด $50-$80)',
        '🧀 MACS Mac & Cheese / Dairy Queen: แคชเชียร์/ครัวกะดึก 18:30-23:00 น. ($14.50-$15.50/ชม.)'
    ],
    [
        'Tier S - Wisconsin (WI)',
        'Wilderness Resort & Glacier Canyon Lodge (WI Dells)',
        'OEG, IEE, Acadex',
        'Waterpark Attendant / Housekeeping',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีค: 08:30 – 19:00 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '40 – 46 ชม./wk',
        '$105',
        'อาณาจักรรีสอร์ตสวนน้ำที่ใหญ่ที่สุดในอเมริกา (600 เอเคอร์), หอพักพนักงานในรีสอร์ต',
        '🥩 Field’s at the Wilderness (สเต๊กหรูในรีสอร์ต): Busser 17:30-23:00 น. ($12/ชม. + ทิปสด $70-$120/คืน)',
        '🍕 Sarento’s Restaurant (ในรีสอร์ต): Food Runner/ล้างจาน 17:00-22:30 น. (ขอ OT ในรีสอร์ต $22.50)',
        '🍔 B-LUX Grill & Bar: Barback/Busser ร้านเบอร์เกอร์คราฟต์เบียร์ 18:00-00:00 น. (ทิปสดดีมาก)'
    ],
    [
        'Tier S - Wisconsin (WI)',
        'Mt. Olympus Water & Theme Park (WI Dells)',
        'New Step, Acadex, Higher',
        'Ride Operator / Housekeeping / Lifeguard',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 21:00 (11.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '42 – 48 ชม./wk',
        '$95',
        'สวนสนุกธีมกรีกโบราณ โรลเลอร์โคสเตอร์ไม้, หอพักราคาถูก, ทำเลติดถนนใหญ่ Parkway',
        '🥞 Paul Bunyan’s Cook Shanty: ล้างจาน/ทำความสะอาดรอบค่ำ 17:30-22:00 น. ($14.50/ชม. + กินฟรี)',
        '🌭 Hot Dog Avenue: แคชเชียร์/ครัวปิด 18:00-22:30 น. ($14.50/ชม.)',
        '🍺 Showboat Saloon (Downtown): Barback ผับดนตรีสด 19:00-01:30 น. (ทิปสดแน่น)'
    ],
    [
        'Tier S - Wisconsin (WI)',
        'Noah’s Ark Waterpark (Wisconsin Dells)',
        'OEG, IEE, New Step',
        'Ride Operator / Park Services',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีค: 08:30 – 18:30 (9.5 ชม.)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '40 – 46 ชม./wk',
        '$100',
        'สวนน้ำกลางแจ้งใหญ่ที่สุดในอเมริกา, เล่นสวนน้ำฟรี, หอพักพนักงานใกล้ที่ทำงาน',
        '🍻 Monk’s Bar & Grill (Downtown Dells): ปั่นจักรยานไปทำ Barback/Busser 18:30-00:00 น. (ทิปสดแน่นมาก)',
        '🍕 Pizza Pub Dells: ช่วยส่งอาหาร/ทำพิซซ่า/ล้างจานรอบดึก 18:00-00:00 น. ($14/ชม. + ทิป)',
        '🍦 Dairy Queen / Cold Stone Dells: ตักไอศกรีมกะค่ำ 18:00-22:30 น. ($14.50/ชม.)'
    ],
    [
        'Tier S - Wisconsin (WI)',
        'Chula Vista Resort & Waterpark (WI Dells)',
        'Higher, Acadex, IEE',
        'Housekeeping / Lifeguard',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '40 – 45 ชม./wk',
        '$105',
        'รีสอร์ตริมแม่น้ำ Wisconsin River, สวนน้ำในร่ม/กลางแจ้ง, หอพักในรีสอร์ต',
        '🥩 Kaminski’s Chop House (ในรีสอร์ต): Busser ร้านสเต๊กหรู 17:30-23:00 น. ($12/ชม. + ทิปสดเศรษฐี $70-$120/คืน)',
        '🌮 Mexicali Rose: Busser/Runner ร้านอาหารเม็กซิกันริมน้ำ 18:00-23:00 น. ($13/ชม. + ทิปสด)',
        '🍕 Kilbourn City Grill (ในรีสอร์ต): ล้างจาน/ผู้ช่วยครัว 17:00-22:00 น. (ขอ OT ในรีสอร์ตได้)'
    ],
    [
        'Tier S - Wisconsin (WI)',
        'Grand Geneva Resort & Spa (Lake Geneva, WI)',
        'Acadex, OEG',
        'Housekeeping / Banquet / Culinary',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$20 - $50',
        '42 – 48 ชม./wk',
        '$110',
        'รีสอร์ตหรูระดับ 4 เพชร AAA เมืองตากอากาศทะเลสาบคนรวยชิคาโก, หอพักพนักงาน',
        '🥩 Geneva ChopHouse (ในรีสอร์ต): Busser 17:30-23:00 น. ($13/ชม. + ทิปสด $60-$100/คืน)',
        '🍕 Popeye’s on Lake Geneva: Busser/Dishwasher ร้านอาหารริมทะเลสาบดัง 18:00-23:30 น. (ทิปดี)',
        '🍺 The Bottle Shop / Local Pubs: Barback ร้านคราฟต์เบียร์ 18:30-00:00 น.'
    ],

    # ==========================================
    # 🤠 4. TENNESSEE (Tier S - 0% State Tax)
    # ==========================================
    [
        'Tier S - Tennessee (TN)',
        'Dollywood Theme Park & Splash Country (Pigeon Forge)',
        'New Step, Acadex',
        'Ride Operator / Food Service / Retail',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 21:30 (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '44 – 50 ชม./wk',
        '$110',
        'ปลอดภาษีเงินได้รัฐ 0%, นั่งรถราง Pigeon Forge Trolley $1/วัน, บัตรเข้าสวนสนุก Dollywood ฟรี',
        '🍗 Paula Deen’s Family Kitchen (The Island): นั่งรถรางไปทำ Busser 18:30-23:00 น. ($13/ชม. + ทิปสด $60-$90)',
        '🍕 Mellow Mushroom Pizza: Barback/Busser ร้านพิซซ่าคราฟต์เบียร์ 18:30-00:00 น. ($12/ชม. + ทิปสด $50-$80)',
        '🥞 Puckett’s Grocery & Restaurant: Food Runner 18:00-23:00 น. ($13/ชม. + ทิป)'
    ],
    [
        'Tier S - Tennessee (TN)',
        'The Island in Pigeon Forge (Arcades & Attractions)',
        'New Step, Acadex',
        'Attractions Host / Retail / F&B',
        'ปกติ: 10:00 – 18:00 (8 ชม.)\nพีค: 10:00 – 23:00 (กะดึก)',
        'ฐาน $15.00 / OT $22.50',
        '$20 - $40',
        '42 – 48 ชม./wk',
        '$110',
        'ศูนย์รวมความบันเทิง ชิงช้าสวรรค์ยักษ์และร้านอาหารกลางเมือง, รถรางผ่านหน้าร้าน',
        '🦜 Margaritaville Restaurant (ใน The Island): Busser/Barback 18:30-00:00 น. ($13/ชม. + ทิปสด $60-$100)',
        '🥞 Timberwood Grill (ใน The Island): ล้างจาน/ผู้ช่วยครัว 18:00-23:30 น. ($14.50/ชม.)',
        '🍭 The Island Candy Kitchen: แคชเชียร์/ทำขนม 18:00-22:30 น. ($15/ชม.)'
    ],
    [
        'Tier S - Tennessee (TN)',
        'Wilderness at the Smokies (Sevierville)',
        'OEG, Acadex, Higher',
        'Lifeguard / Housekeeping',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีค: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '42 – 48 ชม./wk',
        '$105',
        'สวนน้ำในร่มใหญ่ที่สุดในเทนเนสซี, ปลอดภาษีรัฐ 0%, หอพักพนักงาน Sevierville',
        '🍏 Applewood Farmhouse Restaurant: Busser ร้านอาหารดังคนแน่น 18:00-22:30 น. (ทิปสด $60-$100/คืน)',
        '🥩 Hidden Trail Restaurant (ในรีสอร์ต): ล้างจาน/Busser กะค่ำ 18:00-22:30 น. (ขอ OT ในรีสอร์ต $22.50/ชม.)',
        '🛍️ Tanger Outlets Sevierville: จัดสต็อกสินค้า/ปิดร้าน 18:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier S - Tennessee (TN)',
        'Anakeesta Mountaintop Theme Park (Gatlinburg)',
        'Higher, IEE',
        'Guest Host / F&B / Retail',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีค: 08:30 – 20:00 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '$15 - $30',
        '42 – 48 ชม./wk',
        '$110',
        'สวนสนุกธรรมชาติบนยอดเขา Gatlinburg นั่งกระเช้า Chondola, ปลอดภาษีรัฐ 0%',
        '🍺 Smoky Mountain Brewery (Gatlinburg): Barback/Busser 18:30-00:30 น. ($12/ชม. + ทิปสด $70-$120/คืน)',
        '🥩 Alamo Steakhouse: Busser/Dishwasher ร้านสเต๊กคาวบอย 17:30-23:00 น. ($13/ชม. + ทิป)',
        '🥞 Pancake Pantry (ร้านแพนเค้กอันดับ 1): ช่วยเตรียมวัตถุดิบ/ล้างจานรอบบ่าย-เย็น'
    ],
    [
        'Tier S - Tennessee (TN)',
        'Ober Mountain / Gatlinburg Area Lodges',
        'IEE, New Step',
        'Attractions Operator / Retail',
        'ปกติ: 09:00 – 17:00 (8 ชม.)\nพีค: 08:30 – 19:00 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '40 – 46 ชม./wk',
        '$110',
        'กระเช้าลอยฟ้าและสวนสนุกบนยอดเขา Gatlinburg, ปลอดภาษีรัฐ 0%, เมืองเดินเท้า 100%',
        '🥃 Ole Smoky Moonshine Distillery: พนักงานเช็ดแก้ว/เติมสต็อก/แคชเชียร์ 17:30-22:30 น. ($15/ชม.)',
        '🥩 Calhoun’s in Gatlinburg: Busser/Runner ร้านบาร์บีคิวซี่โครงหมูชื่อดัง 17:30-23:00 น. (ทิปดี)',
        '🍕 Best Italian Cafe & Pizzeria: ผู้ช่วยครัว/ล้างจาน 18:00-23:30 น. ($14/ชม. + ทิป)'
    ],

    # ==========================================
    # 🤠 5. TEXAS (Tier S - 0% State Tax)
    # ==========================================
    [
        'Tier S - Texas (TX)',
        'Schlitterbahn Waterpark (New Braunfels / Galveston)',
        'OEG, Acadex, IEE',
        'Lifeguard / Park Operations',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '44 – 50 ชม./wk',
        '$110',
        'สวนน้ำระดับตำนานของเท็กซัส, ปลอดภาษีเงินได้รัฐ 0%, เล่นสวนน้ำฟรี',
        '🥩 Gristmill River Restaurant (Gruene): Busser ร้านอาหารริมแม่น้ำสุดคึกคัก 18:30-23:30 น. ($12/ชม. + ทิปสด $60-$100)',
        '🦫 Buc-ee’s New Braunfels (ใหญ่ที่สุดในโลก): แคชเชียร์/Food Prep กะดึก 19:00-00:00 น. ($17-$19/ชม. จ่ายหนักมาก)',
        '🍺 Krause’s Cafe & Biergarten: Barback/Runner ร้านเบียร์เยอรมัน 18:00-23:00 น. ($13/ชม. + ทิป)'
    ],
    [
        'Tier S - Texas (TX)',
        'Kalahari Resorts & Conventions Round Rock (Austin, TX)',
        'OEG, Acadex',
        'Lifeguard / Housekeeping / F&B',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 48 ชม./wk',
        '$120',
        'สวนน้ำในร่มใหญ่ที่สุดในอเมริกา สาขาเท็กซัส, ปลอดภาษีรัฐ 0%, ใกล้ออสติน',
        '🥩 Double Cut Steak House (ในรีสอร์ต): Busser ร้านสเต๊กหรู 17:30-23:00 น. ($13/ชม. + ทิปสด $70-$120)',
        '🍕 Tom Foolerys Adventure Park (ในรีสอร์ต): คุมเครื่องเล่น/เกมรอบค่ำ 18:00-23:00 น. (ขอ OT $23.25)',
        '🍔 Round Rock Local Eateries: Busser/ล้างจานร้านอาหารรอบนอก 18:30-23:30 น.'
    ],
    [
        'Tier S - Texas (TX)',
        'Six Flags Fiesta Texas & SeaWorld San Antonio',
        'New Step, Higher, IEE',
        'Ride Operator / Lifeguard / Culinary',
        'ปกติ: 10:00 – 18:30 (8 ชม.)\nพีค: 09:30 – 21:30 (11.5 ชม.)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '42 – 48 ชม./wk',
        '$115',
        'สวนสนุกและสวนน้ำระดับโลก ซานอันโตนิโอ, ปลอดภาษีรัฐ 0%, หอพักพนักงาน',
        '🥩 The County Line BBQ (River Walk): Busser ร้านบาร์บีคิวริมแม่น้ำ 19:00-00:00 น. (ทิปสด $60-$100/คืน)',
        '🎸 Hard Rock Cafe San Antonio: Food Runner/Busser 19:00-00:30 น. ($13/ชม. + ทิป)',
        '🌮 Boudro’s on the Riverwalk: ล้างจาน/ผู้ช่วยครัว 19:00-01:00 น. ($15/ชม.)'
    ],
    [
        'Tier S - Texas (TX)',
        'Kemah Boardwalk & Galveston Pleasure Pier',
        'OEG, New Step',
        'Ride Operator / Retail / F&B',
        'ปกติ: 10:30 – 18:30 (8 ชม.)\nพีค: 10:00 – 22:30 (12 ชม.)',
        'ฐาน $14.50 / OT $21.75',
        '$15 - $30',
        '42 – 48 ชม./wk',
        '$110',
        'สวนสนุกริมอ่าวเม็กซิโก เครือ Landry’s, ปลอดภาษีรัฐ 0%, บรรยากาศชายทะเล',
        '🦞 Aquarium Restaurant Kemah (ในเครือ): Busser ร้านอาหารใต้น้ำ 18:30-23:30 น. (ทิปสด $60-$100/คืน)',
        '🍤 Bubba Gump Shrimp Co. (Galveston Pier): Food Runner/Busser 18:30-00:00 น. (ทิปดี)',
        '🏖️ The Spot (Galveston Seawall): Barback/Busser ร้านอาหารริมหาดคนแน่น 19:00-01:00 น.'
    ],

    # ==========================================
    # 🐻 6. MONTANA (Tier A - 0% Sales Tax)
    # ==========================================
    [
        'Tier A - Montana (MT)',
        'Glacier National Park Lodges (Pursuit / Xanterra)',
        'OEG, Acadex, Higher',
        'Hospitality Crew / Housekeeping',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '46 – 52 ชม./wk',
        '$115 (รวมกิน)',
        'อุทยานธรรมชาติธารน้ำแข็ง Glacier NP, ปลอดภาษีมูลค่าเพิ่ม 0% Sales Tax, หอพัก+อาหาร EDR',
        '🥩 Belton Chalet Dining Room: Busser/Steward ร้านอาหารประวัติศาสตร์ 17:30-22:30 น. ($15/ชม. + ทิป)',
        '🛶 Glacier Raft Company Base: พนักงานล้างทำความสะอาดเรือยาง/อุปกรณ์ล่องแก่ง 17:00-21:30 น. ($16/ชม.)',
        '🥐 West Glacier Bakery & Cafe: ผู้ช่วยเตรียมวัตถุดิบ/ล้างจาน 17:00-22:00 น.'
    ],
    [
        'Tier A - Montana (MT)',
        'Many Glacier Hotel & Glacier Park Lodge (East Glacier)',
        'OEG, Acadex',
        'Housekeeping / Dining Room / Steward',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$15 - $30',
        '46 – 52 ชม./wk',
        '$115 (รวมกิน)',
        'โรงแรมสไตล์สวิสชาเลต์ริมทะเลสาบ Swiftcurrent Lake สวยที่สุดใน Glacier NP, อาหาร EDR 3 มื้อ',
        '🥩 Ptarmigan Dining Room (ในโรงแรม): Busser 17:00-22:30 น. (เรต OT $24/ชม. + ทิปสด)',
        '☕ Heidi’s Snack Shop (ในโรงแรม): แคชเชียร์/ชงกาแฟ 16:30-21:30 น. (ขอ OT $24/ชม.)',
        '🛶 Many Glacier Boat Tours: ช่วยเทียบเรือ/ล้างเรือทัวร์ 17:00-20:30 น.'
    ],
    [
        'Tier A - Montana (MT)',
        'Big Sky Resort / Montage Big Sky',
        'Acadex, Higher, IEE',
        'Mountain Operations / Housekeeping',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.50 / OT $26.25',
        '-',
        '46 – 52 ชม./wk',
        '$140',
        'รีสอร์ตภูเขาหรูหราไฮเอนด์, ค่าแรงฐานสูงมาก, มีรถบัส Skyline Bus ฟรี',
        '🍺 Lone Peak Brewery (Town Center): Barback/Busser 17:30-23:30 น. ($15/ชม. + ทิปสดเศรษฐี $60-$100/คืน)',
        '🍖 The Riverhouse BBQ: Busser/Runner ร้านบาร์บีคิวริมแม่น้ำ Gallatin 17:00-22:30 น. ($14/ชม. + ทิป)',
        '🛒 Roxy’s Market Big Sky: พนักงานจัดสต็อกซูเปอร์มาร์เก็ต 18:00-22:00 น. ($17/ชม.)'
    ],

    # ==========================================
    # 🏊 7. MARYLAND & VIRGINIA (Tier A)
    # ==========================================
    [
        'Tier A - Maryland (MD)',
        'Premier Aquatics / High Sierra Pools (Bethesda/Rockville)',
        'Acadex, New Step, ALC',
        'Pool Lifeguard',
        'ปกติ: 10:30 – 19:30 (9 ชม.)\nพีค: 10:00 – 20:30 (10.5 ชม. สระเปิดยาว)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 54 ชม./wk (การันตี OT)',
        '$120',
        'การันตีชั่วโมง OT แน่นอนสัปดาห์ละ 10-14 ชม., อพาร์ตเมนต์แชร์กับเพื่อน, รถไฟใต้ดิน Metro เข้า DC',
        '🛒 Giant Food / Safeway: พนักงานจัดสต็อกสินค้ากะดึก 20:30-01:00 น. ($16-$17.50/ชม. มีสาขาใกล้ที่พัก)',
        '🍜 Thai Tanic / Ruan Thai: ผู้ช่วยครัว/แพ็กอาหาร 20:00-23:30 น. (มีอาหารไทยฟรี)',
        '🎯 Target / CVS Pharmacy: แคชเชียร์/ปิดร้าน 20:00-23:30 น. ($16.00/ชม.)'
    ],
    [
        'Tier A - Maryland (MD)',
        'Ocean City Boardwalk Hotels & Resorts',
        'IEE, New Step, OEG',
        'Housekeeping / Ride Operator',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '44 – 50 ชม./wk',
        '$125',
        'เมืองตากอากาศชายหาด Atlantic, รถบัส Coastal Highway วิ่ง 24 ชม., หางานสองง่ายมาก',
        '🌴 Seacrets Jamaica USA (ผับริมหาดยักษ์ใหญ่): Barback/Busser 18:00-01:30 น. ($12/ชม. + ทิปสดมหาศาล $80-$150/คืน)',
        '🍟 Thrasher’s French Fries (ริมหาด): แคชเชียร์/ทอดเฟรนช์ฟรายส์ 18:00-23:00 น. ($15/ชม.)',
        '🍺 Shenanigans Irish Pub: ล้างจาน/ผู้ช่วยบาร์ 18:30-00:30 น. ($14/ชม. + ทิป)'
    ],
    [
        'Tier A - Maryland (MD)',
        'Jolly Roger Amusement Park & Splash Mountain (Ocean City)',
        'New Step, IEE',
        'Ride Operator / Lifeguard / Park Services',
        'ปกติ: 11:00 – 19:00 (8 ชม.)\nพีค: 10:00 – 23:00 (12 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '45 – 52 ชม./wk',
        '$120',
        'สวนสนุกและสวนน้ำที่ใหญ่ที่สุดใน Ocean City, เล่นสวนสนุกฟรี, มีรถเมล์ผ่านหน้าสวนสนุก',
        '🦀 Higgins Crab House: Busser ร้านปูยักษ์ 19:30-00:30 น. ($13/ชม. + ทิปสด $60-$90/คืน)',
        '🍕 Dough Roller Pizza (Boardwalk): ล้างจาน/ทำพิซซ่ากะดึก 19:30-01:00 น. ($15/ชม.)',
        '🍦 Dumser’s Dairyland: ตักไอศกรีม/แคชเชียร์ 19:00-23:30 น. ($15/ชม.)'
    ],
    [
        'Tier A - Virginia (VA)',
        'Busch Gardens Williamsburg & Water Country USA',
        'New Step, Acadex',
        'Ride Operator / Culinary Operations',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 21:30 (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '44 – 50 ชม./wk',
        '$115',
        'สวนสนุกธีมยุโรปชื่อดัง, บัตรเข้าสวนสนุกและสวนน้ำฟรี, หอพักพนักงาน Williamsburg',
        '🏰 Kingsmill Resort (ริมแม่น้ำ James River): Banquet Server/Busser งานจัดเลี้ยง 18:30-23:00 น. (ทิปสด $50-$90/คืน)',
        '🦀 Captain George’s Seafood Buffet: Busser/Dishwasher บุฟเฟต์ซีฟู้ดยักษ์ใหญ่ 18:30-23:30 น. (คนแน่นมาก ทิปดี)',
        '🥞 Cracker Barrel / Pancake House: ผู้ช่วยครัว/ล้างจานรอบค่ำ 18:00-22:00 น. ($14.50/ชม.)'
    ],
    [
        'Tier A - Virginia (VA)',
        'Virginia Beach Oceanfront Resorts & Boardwalk',
        'IEE, New Step, Acadex',
        'Housekeeping / Beach Attendant',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '$15 - $30',
        '42 – 48 ชม./wk',
        '$130',
        'เมืองตากอากาศชายทะเลเวอร์จิเนีย มีทางเดินเลียบหาดยาว 3 ไมล์, รถราง Wave Trolley สะดวก',
        '🍺 Waterman’s Surfside Grille: Barback/Busser ร้านริมหาดคนแน่น 18:00-00:30 น. (ทิปสด $70-$120/คืน)',
        '🦀 Catch 31 Fish Bar: Food Runner/Busser 18:00-23:30 น. (ทิปดี)',
        '🍕 Dough Boy’s California Pizza: ผู้ช่วยครัว/ล้างจาน 18:30-00:00 น. ($14.50/ชม.)'
    ],

    # ==========================================
    # 🗿 8. SOUTH DAKOTA (Tier A - 0% State Tax)
    # ==========================================
    [
        'Tier A - South Dakota (SD)',
        'Xanterra Mount Rushmore / Keystone Lodges',
        'OEG, IEE, Higher',
        'Retail / Food Service / Housekeeping',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '44 – 50 ชม./wk',
        '$105',
        'ทำงานหน้าอนุสรณ์สถานแห่งชาติ Mount Rushmore, ปลอดภาษีรัฐ 0%, หอพัก+อาหาร EDR',
        '🍔 Carver’s Cafe at Mt Rushmore: ล้างจาน/ปิดครัว 17:00-21:30 น. (ขอ OT ในอุทยาน $23.25/ชม.)',
        '🥩 Ruby House Restaurant (Keystone): Busser ร้านอาหารธีมคาวบอยย้อนยุค 17:30-22:30 น. ($13/ชม. + ทิป)',
        '🚂 1880 Train / Keystone Gift Shops: พนักงานร้านของฝาก/สถานีรถไฟโบราณ 17:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier A - South Dakota (SD)',
        'Custer State Park Resorts (State Game Lodge / Sylvan)',
        'Higher, Acadex',
        'Guest Service / Housekeeping',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 16:30 (9 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 48 ชม./wk',
        '$105 (รวมกิน)',
        'อุทยานธรรมชาติฝูงควายไบซัน Custer SP, ปลอดภาษีรัฐ 0%, หอพัก+อาหาร EDR',
        '🥩 State Game Lodge Dining Room: Busser ร้านสเต๊กควายไบซันหรู 17:00-22:00 น. ($13/ชม. + ทิปสด $50-$80)',
        '🍔 Black Hills Burger & Bun (เมือง Custer): ล้างจาน/ครัว 17:00-21:30 น. ($15/ชม. + กินฟรี)',
        '🏕️ Sylvan Lake General Store: พนักงานมินิมาร์ท/เช่าเรือแคนู 16:30-21:00 น.'
    ],
    [
        'Tier A - South Dakota (SD)',
        'Wall Drug Store (Wall, SD)',
        'IEE, Acadex',
        'Retail / Restaurant Staff / Fast Food',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '45 – 50 ชม./wk',
        '$85 (ถูกมาก)',
        'ห้างคาวบอยระดับตำนานของอเมริกา จุดแวะพักยอดฮิตก่อนเข้า Badlands NP, หอพักราคาถูก',
        '☕ Wall Drug Western Art Dining Room: Busser/Steward 17:00-22:00 น. (ขอ OT ในตัว $23.25)',
        '🍩 Wall Drug Famous Donut Kitchen: ช่วยทำโดนัท/เบเกอรี่รอบเย็น 16:30-21:00 น.',
        '🌵 Badlands Saloon & Grille: Barback/ล้างจาน 18:00-23:00 น. ($14.50/ชม. + ทิป)'
    ],
    [
        'Tier A - South Dakota (SD)',
        'Deadwood Historic Hotels & Casinos (The Lodge at Deadwood)',
        'Higher, New Step',
        'Housekeeping / Food & Beverage',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '$20 - $50',
        '42 – 48 ชม./wk',
        '$100',
        'เมืองคาวบอยคาสิโนประวัติศาสตร์ ปลอดภาษีรัฐ 0%, มีรถราง Deadwood Trolley $1',
        '🥩 Deadwood Grille / O’Shea’s: Busser/Runner 17:30-23:00 น. ($13/ชม. + ทิปสด $50-$90/คืน)',
        '🎰 Mineral Palace / Silverado Casino: Barback บาร์ในคาสิโน 18:30-01:00 น. (ทิปดี)',
        '🍕 Mustang Sally’s Sports Bar: ผู้ช่วยครัว/ล้างจาน 18:00-00:00 น. ($14.50/ชม.)'
    ],

    # ==========================================
    # 🏜️ 9. UTAH (Tier A)
    # ==========================================
    [
        'Tier A - Utah (UT)',
        'Zion National Park Lodge (Xanterra Springdale)',
        'OEG, IEE, Acadex',
        'Hospitality Crew / Housekeeping',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 16:30 (9 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '44 – 50 ชม./wk',
        '$110',
        'อุทยานแห่งชาติหุบเขาผาแดง Zion NP, มีรถชัตเติลบัสฟรีวิ่งตลอดถนน Springdale',
        '🌮 Bit & Spur Restaurant & Saloon: Busser/Barback 17:30-23:00 น. ($13/ชม. + ทิปสดแขกต่างชาติ $50-$90/คืน)',
        '🍔 Oscar’s Cafe (Springdale): Food Runner/Busser 17:00-22:00 น. ($13/ชม. + ทิป)',
        '🍺 Zion Canyon Brew Pub: ล้างจาน/ผู้ช่วยครัวหน้าปากทางเข้าอุทยาน 18:00-23:30 น. ($15/ชม.)'
    ],
    [
        'Tier A - Utah (UT)',
        'Ruby’s Inn / Bryce Canyon Grand Hotel',
        'Higher, IEE, Acadex',
        'Housekeeping / Fast Food / Retail',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '42 – 48 ชม./wk',
        '$100',
        'อาณาจักรโรงแรมและรีสอร์ต Ruby’s Inn หน้าปากทาง Bryce Canyon NP, หอพักพนักงานใกล้ที่ทำงาน',
        '🥩 Cowboy’s Buffet & Steak Room (ในเครือ): Busser/Steward 17:00-22:30 น. (ขอ OT ในเครือ $22.50/ชม.)',
        '🤠 Ebenezer’s Barn & Grill: Food Runner งานดินเนอร์โชว์คาวบอย 18:00-22:00 น. (ได้ส่วนแบ่งทิป)',
        '🛒 Ruby’s General Store & Diner: แคชเชียร์/ผู้ช่วยครัว 17:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier A - Utah (UT)',
        'Park City Mountain Resort (Vail Resorts - Park City)',
        'OEG, Acadex',
        'Mountain Host / Housekeeping / F&B',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $17.50 / OT $26.25',
        '-',
        '42 – 48 ชม./wk',
        '$150',
        'สกีรีสอร์ตและเมืองภูเขาที่ใหญ่ที่สุดในสหรัฐฯ, ระบบรถเมล์ฟรีทั้งเมือง Park City Transit',
        '🥩 High West Distillery & Saloon (Main Street): Barback/Busser โรงกลั่นวิสกี้ดัง 17:30-23:30 น. (ทิปสด $80-$140/คืน)',
        '🍕 Red Banjo Pizza (Historic Main St): ล้างจาน/ผู้ช่วยครัว 18:00-23:00 น. ($16/ชม. + ทิป)',
        '🛒 Freshies Lobster Co. / Local Cafes: Food Runner 17:00-21:30 น. ($15/ชม. + ทิป)'
    ],

    # ==========================================
    # 🦞 10. MAINE (Tier A)
    # ==========================================
    [
        'Tier A - Maine (ME)',
        'Bar Harbor Grand Hotel / Harborside Hotel',
        'Acadex, OEG, Higher',
        'Housekeeping / Laundry',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '-',
        '44 – 50 ชม./wk',
        '$130',
        'เมืองตากอากาศชายทะเล Bar Harbor & อุทยาน Acadia NP, รถบัส Island Explorer ฟรี',
        '🦞 Stewman’s Lobster Pound (ริมทะเล): Busser/Barback ร้านกุ้งล็อบสเตอร์ 17:30-23:00 น. ($13/ชม. + ทิปสด $60-$120/คืน)',
        '🍺 Geddy’s Down Under / The Chart Room: Busser/Runner 18:00-23:30 น. (ทิปสดแน่นมาก)',
        '🍦 Ben & Bill’s Chocolate Emporium: พนักงานตักไอศกรีม/แคชเชียร์ 17:00-22:00 น. ($15.50/ชม.)'
    ],
    [
        'Tier A - Maine (ME)',
        'Cliff House Maine (Cape Neddick / Ogunquit)',
        'Higher, Acadex',
        'Housekeeping / Culinary / Stewarding',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.00 / OT $25.50',
        '-',
        '42 – 48 ชม./wk',
        '$135',
        'รีสอร์ตหรูบนหน้าผาริมมหาสมุทรแอตแลนติก แขกไฮเอนด์ระดับมหาเศรษฐีบอสตันและนิวยอร์ก',
        '🦞 The Tiller Restaurant (ในรีสอร์ต): Busser ร้าน Fine Dining วิวมหาสมุทร 17:30-23:00 น. (ทิปสด $70-$130/คืน)',
        '🍺 Nubb’s Lobster Shack (ในรีสอร์ต): Food Runner/Barback 17:00-22:30 น. (ขอ OT ในรีสอร์ต $25.50)',
        '🍦 Ogunquit Beach Cafes: ผู้ช่วยครัว/บริการรอบเย็น'
    ],
    [
        'Tier A - Maine (ME)',
        'The Nonantum Resort (Kennebunkport, ME)',
        'Higher, Acadex',
        'Housekeeping / Banquet / F&B',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '$20 - $50',
        '42 – 48 ชม./wk',
        '$125',
        'รีสอร์ตตากอากาศริมน้ำประวัติศาสตร์ เมืองของประธานาธิบดีบุช, หอพักพนักงาน Kennebunkport',
        '🦞 The Clam Shack (สะพาน Kennebunkport): ล้างจาน/ทอดซีฟู้ดร้านดัง 17:30-22:30 น. ($16/ชม. + ทิป)',
        '🥩 Ocean Restaurant: Busser ร้านอาหารหรูริมน้ำ 18:00-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍦 Rococo Artisan Ice Cream: ตักไอศกรีม 17:00-22:00 น. ($15/ชม.)'
    ],

    # ==========================================
    # 🎢 11. OHIO (Tier A - หอพักถูกที่สุด $80/wk)
    # ==========================================
    [
        'Tier A - Ohio (OH)',
        'Cedar Point Amusement Park & Resorts (Sandusky)',
        'OEG, IEE Thailand',
        'Ride Operator / Food Service',
        'ปกติ: 10:00 – 18:30 (8 ชม.)\nพีค: 09:30 – 22:30 (12 ชม. สวนสนุกปิดดึก)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '46 – 52 ชม./wk',
        '$80 (ถูกสุด)',
        'สวนสนุกเครื่องเล่นระดับโลก, หอพักราคาถูกที่สุดในอเมริกา ($80/wk), รถบัสรับส่งฟรี, เล่นสวนสนุกฟรี',
        '🏨 Hotel Breakers / Castaway Bay (ในสวนสนุก): ล้างจาน/แม่บ้านกะค่ำ 19:00-00:00 น. (ขอ OT ในสวนสนุก $21.75/ชม.)',
        '🍖 Famous Dave’s BBQ (Cedar Point Marina): Busser/Dishwasher 18:30-23:30 น. ($13/ชม. + ทิปสด)',
        '🐴 Thirsty Pony / Kalahari Sandusky: Barback/Busser ร้านอาหารใกล้เคียง 19:00-00:30 น.'
    ],
    [
        'Tier A - Ohio (OH)',
        'Kings Island Amusement Park (Mason / Cincinnati)',
        'OEG, IEE Thailand',
        'Ride Operator / Park Services / Culinary',
        'ปกติ: 10:00 – 18:30 (8 ชม.)\nพีค: 09:30 – 22:30 (12 ชม.)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '44 – 50 ชม./wk',
        '$90',
        'สวนสนุกชั้นนำเครือ Cedar Fair ใกล้เมือง Cincinnati, หอพักพนักงาน + รถรับส่งฟรี',
        '🍕 Mason Local Pizzerias & Diners: ล้างจาน/ผู้ช่วยครัวรอบดึก 19:00-00:00 น. ($14.50/ชม.)',
        '🍦 Greaters Ice Cream: พนักงานตักไอศกรีมชื่อดัง 19:00-23:00 น. ($14.50/ชม.)',
        '🏨 Great Wolf Lodge Mason (ข้างๆ สวนสนุก): แม่บ้าน/ผู้ช่วยครัวกะค่ำ (ขอทำเพิ่มได้)'
    ],

    # ==========================================
    # ⛰️ 12. COLORADO (Tier A)
    # ==========================================
    [
        'Tier A - Colorado (CO)',
        'YMCA of the Rockies (Estes Park Center)',
        'Higher, IEE, Acadex',
        'Housekeeping / Food Service',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 16:30 (9 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 48 ชม./wk',
        '$110 (รวมกิน)',
        'แคมป์รีสอร์ตกลางเทือกเขา Rocky Mountain NP, หอพัก+อาหารบุฟเฟต์ 3 มื้อ, รถชัตเติลบัสเข้าเมืองฟรี',
        '🏨 The Stanley Hotel (โรงแรมประวัติศาสตร์ The Shining): Busser/Steward ร้านอาหารหรู 17:30-23:00 น. (ทิปสด $50-$90)',
        '🍔 Penelope’s Old Time Burgers: ผู้ช่วยครัว/ทอดเบอร์เกอร์ 17:00-21:30 น. ($15/ชม. + กินฟรี)',
        '🍺 Estes Park Brewery / Rock Cut Brewing: ล้างแก้ว/ผู้ช่วยบาร์ 18:00-22:30 น. ($15/ชม.)'
    ]
]

ws.append(headers)
for r in rows:
    ws.append(r)

# Header style
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Row styling
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(headers)), start=2):
    tier_text = str(row[0].value or '')
    if 'Tier S' in tier_text:
        row_fill = tier_s_fill
    else:
        row_fill = tier_a_fill
    
    # Alternate with subtle zebra
    if row_idx % 2 == 0:
        fill_to_use = row_fill
    else:
        fill_to_use = PatternFill(fill_type=None)

    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
        
        if col_idx in [1, 6, 7, 8, 9]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx in [5, 10, 11, 12, 13]:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Column width auto-fit
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        lines = val_str.split('\n')
        for l in lines:
            length = len(l.encode('utf-8')) // 2 if any(ord(c) > 127 for c in l) else len(l)
            if length > max_len:
                max_len = length
    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 55)

# Freeze top row and 2 leftmost columns
ws.freeze_panes = 'C2'

for p in [excel_path, excel_clean_path]:
    try:
        wb.save(p)
        print('Successfully saved comprehensive database to:', p)
    except PermissionError:
        print('Locked path:', p)
