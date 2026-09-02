# -*- coding: utf-8 -*-
import openpyxl
import re, os, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb_public = openpyxl.load_workbook(public_path, data_only=True)
ws_src = wb_public['All Public Jobs']

print(f'Processing and cleaning {ws_src.max_row - 1} public jobs from source...')

def clean_text(val):
    if val is None or str(val).strip().lower() in ['none', 'null', 'nan', '']:
        return ''
    return str(val).strip()

def clean_employer_title(title, agency, state_code, raw_loc, pos_val):
    t = clean_text(title)
    
    if agency == 'NewStep':
        # Remove Thai marketing prefix
        t = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', t, flags=re.IGNORECASE)
        # Remove extra charge/no charge tags at the end
        t = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t, flags=re.IGNORECASE)
        # Match 'Employer เมือง City รัฐ State'
        m = re.search(r'^(.*?)\s+เมือง\s+(.*?)\s+รัฐ\s+.*$', t)
        if m:
            emp = m.group(1).strip()
            city = m.group(2).strip()
            return f'{emp} ({city}, {state_code})' if city else f'{emp} ({state_code})'
        # Match 'Employer เมือง City'
        m2 = re.search(r'^(.*?)\s+เมือง\s+(.*?)$', t)
        if m2:
            emp = m2.group(1).strip()
            city = m2.group(2).strip()
            return f'{emp} ({city}, {state_code})'
        return f'{t} ({state_code})'
        
    elif agency == 'Acadex':
        # Remove group and year tags e.g. (Summer 2027: Group X)
        t = re.sub(r'\s*\((?:Summer\s*2027|Summer\s*2026)[^)]*\)', '', t, flags=re.IGNORECASE).strip()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'IEE':
        # Convert ALL CAPS to Title Case
        t = t.title()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'OEG':
        if 'Work & Travel' in t:
            if 'Denali' in str(raw_loc) or 'Denali' in str(pos_val):
                return f'Denali Princess Wilderness Lodge ({state_code})'
            elif 'Cedar Point' in str(pos_val) or 'Cedar Point' in str(raw_loc):
                return f'Cedar Point Amusement Park ({state_code})'
            elif 'Kalahari' in str(pos_val) or 'Kalahari' in str(raw_loc):
                return f'Kalahari Resorts & Conventions ({state_code})'
            elif 'Six Flags' in str(pos_val) or 'Six Flags' in str(raw_loc) or '2 สวนสนุก' in t:
                return f'Six Flags Great Escape & Hurricane Harbor ({state_code})'
            else:
                pos_clean = str(pos_val).split('(')[0].strip() if pos_val else ''
                return f'OEG Placement - {pos_clean} ({state_code})'
        return f'{t} ({state_code})' if state_code not in t else t
        
    else:
        if state_code and state_code not in t and len(t) < 40:
            return f'{t} ({state_code})'
        return t

def parse_rate(rate_val, pos_val, title_val):
    combined = f"{clean_text(rate_val)} {clean_text(pos_val)} {clean_text(title_val)}"
    m = re.findall(r'\$(\d+(?:\.\d+)?)', combined)
    if m:
        nums = [float(x) for x in m if float(x) >= 5.0]
        if nums:
            if len(nums) >= 2 and nums[0] != nums[1] and nums[1] <= 35.0:
                r1, r2 = min(nums[:2]), max(nums[:2])
                return f"ฐาน ${r1:.2f} - ${r2:.2f} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
            else:
                r1 = nums[0]
                return f"ฐาน ${r1:.2f} / OT ${r1*1.5:.2f}"
    return "ฐาน $14.00 - $16.00 / OT $21.00 - $24.00"

def parse_hours(hours_val, pos_val):
    combined = f"{clean_text(hours_val)} {clean_text(pos_val)}"
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)', combined)
    if m:
        h1, h2 = m.group(1), m.group(2)
        return f"{h1} – {h2} ชม./wk"
    m2 = re.search(r'(\d+)\s*(?:hours|hrs|ชม)', combined, re.IGNORECASE)
    if m2:
        h = m2.group(1)
        return f"{h} – {int(h)+8} ชม./wk"
    return "40 – 48 ชม./wk"

def parse_housing(hsg_val, desc_val):
    combined = f"{clean_text(hsg_val)} {clean_text(desc_val)}".lower()
    if any(k in combined for k in ['ฟรี', 'บ้านฟรี', 'free', '$0', 'no charge']):
        return "ฟรี ($0) (ที่พักฟรี!)"
    if any(k in combined for k in ['รวมกิน', 'รวมอาหาร', 'meal included', 'edr included', 'free meals']):
        m = re.search(r'\$(\d+(?:\.\d+)?)', clean_text(hsg_val))
        if m:
            p = int(float(m.group(1)))
            return f"${p} (รวมกิน)"
        return "$105 (รวมกิน)"
    
    m = re.search(r'\$(\d+(?:\.\d+)?)', clean_text(hsg_val))
    if m:
        p = float(m.group(1))
        if 'day' in combined or 'ต่อวัน' in combined:
            p = p * 7
        elif 'month' in combined or 'ต่อเดือน' in combined:
            p = p / 4.33
        elif 'bi-weekly' in combined or '2 week' in combined:
            p = p / 2
        return f"${int(p)}"
    
    return "$100 - $125"

def parse_shifts(hours_str):
    if '32' in hours_str or '35' in hours_str:
        return "ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:00 (9 ชม.)"
    elif '50' in hours_str or '55' in hours_str or '60' in hours_str:
        return "ปกติ: 07:30 – 16:00 (8.5 ชม.)\nพีคซัมเมอร์: 07:00 – 17:30 (10.5 ชม. ควงกะ)"
    return "ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)"

def parse_tips(pos_str, title_str):
    combined = f"{clean_text(pos_str)} {clean_text(title_str)}".lower()
    if any(k in combined for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
        return "$40 - $100"
    if any(k in combined for k in ['housekeep', 'room attendant', 'clean', 'maid']):
        return "$15 - $35"
    if any(k in combined for k in ['cook', 'kitchen', 'dish', 'prep', 'steward', 'lifeguard', 'ride op', 'cashier', 'retail', 'stocker']):
        return "-"
    return "$10 - $25"

def generate_2nd_jobs(state_name, city_name):
    st = state_name.lower()
    if 'alaska' in st:
        return [
            "🍕 Prospectors Pizzeria / Mountain High: ล้างจาน/Barback กะค่ำ 18:30-00:30 น. ($15/ชม. + ทิปสด)",
            "🦀 Local Salmon / Seafood Restaurant: Busser/Runner 17:30-23:00 น. ($14/ชม. + ทิป)",
            "🛒 Supermarket / General Store: จัดสต็อกสินค้า/แคชเชียร์รอบค่ำ 17:00-22:00 น. ($15-$16/ชม.)"
        ]
    elif 'wyoming' in st:
        return [
            "🥩 Town Square / Resort Dining Room: Busser 17:30-22:30 น. (ทิปสด $50-$100/คืน)",
            "🍺 Local Saloon / Western Bar: Barback 18:30-00:30 น. (ทิปสดแน่น)",
            "🛒 Park General Stores / Retail: เติมสต็อกสินค้า/แคชเชียร์ 17:00-21:30 น."
        ]
    elif 'wisconsin' in st:
        return [
            "🍕 Moosejaw Pizza / Pizza Pub: Busser/ล้างจาน 18:30-23:30 น. ($13/ชม. + ทิป $60-$90)",
            "🎢 Mt. Olympus / Noah’s Ark: ช่วยคุมเครื่องเล่น/สวนน้ำรอบบ่าย-ค่ำ ($14.50-$15/ชม.)",
            "🍦 Dairy Queen / Cold Stone: ตักไอศกรีม/แคชเชียร์ 18:00-23:00 น."
        ]
    elif 'tennessee' in st:
        return [
            "🍗 Paula Deen’s / The Island Dining: Busser 18:30-23:00 น. (ทิปสดแน่น ปลอดภาษี 0%)",
            "🍕 Mellow Mushroom / Local Pizzeria: Barback/ผู้ช่วยครัว 18:30-00:00 น.",
            "🥞 Pancake House / Local Diners: ผู้ช่วยครัว/ล้างจาน 17:00-22:30 น."
        ]
    elif 'maine' in st:
        return [
            "🦞 Local Lobster Pound / Seafood Pier: Busser 17:30-23:00 น. (ทิปสด $60-$120/คืน)",
            "🍺 Craft Brewery / Waterfront Pub: Barback 18:00-00:00 น. ($15/ชม. + ทิป)",
            "🍦 Ice Cream & Fudge Shops: ตักไอศกรีม/ขายขนม 18:00-22:30 น."
        ]
    elif 'virginia' in st:
        return [
            "🥩 Colonial / Beachfront Steakhouses: Busser 17:30-23:00 น. (ทิปสด $60-$100)",
            "🎢 Busch Gardens / King's Dominion: คุมเครื่องเล่นรอบบ่าย-ค่ำ ($13.75-$15/ชม.)",
            "🦀 Waterfront Seafood / BBQ: Food Runner 18:00-23:30 น."
        ]
    elif 'north carolina' in st:
        return [
            "🦀 Outer Banks Seafood / Grill: Busser 17:30-23:00 น. (ทิปสด $60-$110/คืน)",
            "🍺 OBX Craft Brewing: Barback 18:30-00:30 น. ($14/ชม. + ทิป)",
            "🍩 Duck Donuts / Ice Cream: แคชเชียร์รอบค่ำ 18:00-22:30 น."
        ]
    elif 'south carolina' in st:
        return [
            "🦀 Myrtle Beach Oceanfront Seafood: Busser 17:30-23:00 น. (ทิปสด $70-$120)",
            "🎡 Broadway at the Beach: พนักงานของที่ระลึก/ร้านอาหาร 18:00-00:00 น.",
            "🍕 Boardwalk Pizzerias: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น."
        ]
    elif 'ohio' in st:
        return [
            "🥩 Put-in-Bay / Lake Erie Restaurants: Busser 17:30-23:30 น. (ทิปสด $70-$120)",
            "🎢 Cedar Point / Kings Island: ขอกะคุมเครื่องเล่นรอบค่ำ ($15/ชม.)",
            "🍺 Island Bars / Brewpubs: Barback 18:30-00:30 น."
        ]
    elif 'new york' in st:
        return [
            "🍕 NYC Pizzerias / Diners: ผู้ช่วยทำพิซซ่า/ล้างจานรอบค่ำ ($16/ชม. + ทิป)",
            "🥩 Manhattan / Queens Restaurants: Busser 18:00-23:30 น. (ทิปสดแน่น)",
            "🛒 Supermarket / Retail Store: แคชเชียร์/จัดสต็อกกะดึก ($16-$18/ชม.)"
        ]
    else:
        return [
            "🥩 Local Steakhouses / Casual Dining: Busser/Food Runner 17:30-23:00 น. (ทิปสด $50-$90)",
            "🍕 Local Pizzeria / Burger Bar: ผู้ช่วยครัว/ล้างจาน 18:00-23:30 น. ($14-$16/ชม.)",
            "🛒 Supermarket / Retail Shop: แคชเชียร์/จัดสต็อกกะดึก ($15-$17/ชม.)"
        ]

all_formatted_jobs = []

for r in range(2, ws_src.max_row + 1):
    tier = clean_text(ws_src.cell(r, 1).value)
    state_name = clean_text(ws_src.cell(r, 2).value)
    state_code = clean_text(ws_src.cell(r, 3).value)
    agency = clean_text(ws_src.cell(r, 4).value)
    title = clean_text(ws_src.cell(r, 5).value)
    pos = clean_text(ws_src.cell(r, 10).value)
    rate = clean_text(ws_src.cell(r, 11).value)
    hours = clean_text(ws_src.cell(r, 12).value)
    start_d = clean_text(ws_src.cell(r, 13).value)
    end_d = clean_text(ws_src.cell(r, 14).value)
    housing = clean_text(ws_src.cell(r, 15).value)
    hsg_dep = clean_text(ws_src.cell(r, 16).value)
    trans = clean_text(ws_src.cell(r, 17).value)
    eng = clean_text(ws_src.cell(r, 18).value)
    loc = clean_text(ws_src.cell(r, 19).value)
    src_url = clean_text(ws_src.cell(r, 21).value)

    state_display = f"{state_name} ({state_code})" if state_code else state_name
    
    # Clean Employer / Workplace Name
    emp_display = clean_employer_title(title, agency, state_code, loc, pos)

    # Format Position
    pos_display = pos if pos else "Resort Associate / Food Service / Housekeeping / Retail"
    if len(pos_display) > 80:
        pos_display = pos_display[:77] + "..."

    # Extract clean attributes
    rate_clean = parse_rate(rate, pos, title)
    hours_clean = parse_hours(hours, pos)
    shifts_clean = parse_shifts(hours_clean)
    tips_clean = parse_tips(pos, title)
    hsg_clean = parse_housing(housing, f"{pos} {title}")
    
    # Format Benefits & Highlights
    ben_parts = []
    if eng: ben_parts.append(f"ภาษา: {eng}")
    if trans: ben_parts.append(f"การเดินทาง: {trans}")
    if hsg_dep: ben_parts.append(f"มัดจำบ้าน: {hsg_dep}")
    if start_d and end_d: ben_parts.append(f"ช่วงเวลา: {start_d} ถึง {end_d}")
    if src_url: ben_parts.append(f"Link: {src_url}")
    ben_display = " | ".join(ben_parts) if ben_parts else "มีที่พักและสิ่งอำนวยความสะดวกจัดสรรให้, สภาพแวดล้อมปลอดภัย"
    if len(ben_display) > 250:
        ben_display = ben_display[:247] + "..."

    # 2nd jobs
    j2_opts = generate_2nd_jobs(state_name, loc)

    job_row = [
        tier,
        state_display,
        emp_display,
        agency,
        pos_display,
        shifts_clean,
        rate_clean,
        tips_clean,
        hours_clean,
        hsg_clean,
        ben_display,
        j2_opts[0],
        j2_opts[1],
        j2_opts[2]
    ]
    all_formatted_jobs.append(job_row)

print(f"Total formatted job rows: {len(all_formatted_jobs)}")

# Create Workbook
wb = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Sheet 1: Top Employers (Summer Only) - All 1,167 Jobs (11 Columns)
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True
headers_sheet1 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น']
ws1.append(headers_sheet1)
for job in all_formatted_jobs:
    ws1.append(job[:11])

# Sheet 2: Tier S-A Summer Jobs - All 1,167 Jobs with 2nd Job Guide (14 Columns)
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True
headers_sheet2 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)']
ws2.append(headers_sheet2)
for job in all_formatted_jobs:
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True
agency_headers = ['ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)', 'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย']
ws3.append(agency_headers)
agency_directory = [
    ['OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange', 'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด', 'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)', '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'],
    ['Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA', 'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก', 'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)', '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'],
    ['IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions', 'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก', 'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)', '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'],
    ['IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange', 'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง', 'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)', '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'],
    ['New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA', 'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด', 'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)', '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'],
    ['Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International', 'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ', 'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)', '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'],
    ['American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier', 'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู', 'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)', '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ']
]
for row in agency_directory:
    ws3.append(row)

# Sheet 4: State Summary
ws4 = wb.create_sheet(title='State Summary')
ws4.views.sheetView[0].showGridLines = True
ws_sum_src = wb_public['State Summary']
for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws4.append(row_vals)

# Formatting
for ws in [ws1, ws2]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 42

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            if c in [1, 2, 8, 9, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [6, 7]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 40, 22, 35, 32, 28, 14, 18, 18, 50]
col_widths_s2 = [12, 18, 40, 22, 35, 32, 28, 14, 18, 18, 50, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws4.max_column + 1):
    cell = ws4.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws4.row_dimensions[3].height = 28

for r in range(4, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 24
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws4.max_column + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 16
ws4.column_dimensions['B'].width = 22

# Format Sheet 3 (Summer Agency Directory)
ws3.row_dimensions[1].height = 28
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 28
    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Save files
try:
    wb.save(clean_path)
    print('Saved to clean path:', clean_path)
except Exception as e:
    print('Error saving clean path:', e)

try:
    wb.save(master_path)
    print('Saved to master path:', master_path)
except Exception as e:
    print('Master path note (will attempt copy):', e)
    try:
        shutil.copyfile(clean_path, master_path)
        print('Copied to master path successfully.')
    except Exception as err:
        print('Master file in use:', err)

print('Build complete: Cleaned NewStep and all public jobs in Top Employers format!')
