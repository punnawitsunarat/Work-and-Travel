import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'WAT 2027 Jobs Masterlist'
ws.views.sheetView[0].showGridLines = True

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

headers = [
    'ชื่องาน / สถานที่ทำงาน (Employer)',
    'รัฐ (State)',
    'Agency ที่ถือสัญญา',
    'ตำแหน่ง (Position / Role)',
    'เวลาทำงานจริง (ช่วงปกติ / ช่วงพีค)',
    'ค่าแรงฐาน / OT ($/ชม.)',
    'ทิป ($/สัปดาห์)',
    'ชั่วโมงทำงานจริง (คิดเผื่อหัว-ท้ายซีซัน)',
    'ค่าที่พัก ($/สัปดาห์)',
    'สวัสดิการ (อาหาร ฟรี หรือ อื่นๆ)'
]

rows = [
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher',
        'Housekeeping / Room Attendant',
        'ปกติ: 07:30 – 15:30 น. (8 ชม.)\nพีค: 07:00 – 17:30 น. (10 ชม. ควงกะ)',
        'ฐาน $16.00 / OT $24.00',
        '$15 - $30',
        '50 – 54 ชม./สัปดาห์ (พีค 58-62 ชม. / หัวท้าย 35-42 ชม.)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน, ทิปหัวเตียง, รถรับส่งพนักงาน, ส่วนลดทัวร์/รถไฟ 50%'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher',
        'Kitchen Stewarding / Dishwasher (ล้างจาน)',
        'ปกติ: 16:00 – 00:00 น. (8 ชม.)\nพีค: 15:00 – 01:30 น. (10.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 55 ชม./สัปดาห์ (ควงกะค่ำได้ง่าย)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อใน EDR + เชฟทำสเต๊ก/เบอร์เกอร์เลี้ยงในครัวฟรีทุกกะ'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher',
        'Luggage Handler (พนักงานยกกระเป๋าทัวร์)',
        'ปกติ: 08:00 – 16:30 น. หรือ 12:00 – 20:30 น.\nพีค: 07:00 – 18:30 น. (ตามรอบรถไฟ)',
        'ฐาน $15.50 / OT $23.25',
        '$40 - $70',
        '46 – 52 ชม./สัปดาห์',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, ทิปเงินสดจากกรุ๊ปทัวร์ผู้สูงอายุ'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'ขอทำเพิ่มหน้างาน (Internal 2nd Job)',
        'Fleet Detailer (งานเสริมล้างรถทัวร์กะดึก)',
        'ปกติ: 21:00 – 01:30 น. (4.5 ชม.)\nพีค: 21:00 – 02:30 น. (5.5 ชม.)',
        'คิดเรต OT $24.00/ชม. ทันที',
        '-',
        '20 – 25 ชม./สัปดาห์ (ทำควบหลังงานหลักแม่บ้าน)',
        '-',
        'นับเป็นชั่วโมง OT $24.00 ทั้งหมด, ทำงานในอู่รถบัส Princess, อาหาร EDR 3 มื้อ'
    ],
    [
        'Denali Bluffs Hotel',
        'Alaska (AK)',
        'Acadex Thailand',
        'Housekeeping / Laundry',
        'ปกติ: 08:00 – 16:00 น. (8 ชม.)\nพีค: 08:00 – 17:00 น. (9 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '40 – 45 ชม./สัปดาห์ (หัวท้าย 32-38 ชม.)',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป (ครบสัญญา), รถชัตเติลบัสขึ้น-ลงเขา Sugarloaf, ส่วนลดอาหารร้าน The Perk'
    ],
    [
        'Grande Denali Lodge',
        'Alaska (AK)',
        'Acadex Thailand',
        'Housekeeping / Public Area Cleaner',
        'ปกติ: 08:00 – 16:00 น. (8 ชม.)\nพีค: 07:30 – 17:00 น. (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 46 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป (ครบสัญญา), โรงแรมหรู 4 ดาว, รถชัตเติลบัส, ส่วนลดอาหารห้องอาหาร Alpenglow'
    ],
    [
        'Grande Denali Lodge',
        'Alaska (AK)',
        'Acadex Thailand',
        'Busser / Food Runner (Alpenglow Restaurant)',
        'ปกติ: 16:30 – 22:30 น. (6 ชม.)\nพีค: 16:00 – 23:30 น. (7.5 ชม.)',
        'ฐาน $14.00 / OT $21.00',
        '$200 - $350',
        '40 – 45 ชม./สัปดาห์',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป, ทิปเงินสดจากแขกไฮเอนด์ร้านอาหารวิวพาโนรามา ($40-$70/คืน)'
    ],
    [
        'Mt. McKinley Princess Lodge',
        'Alaska (AK)',
        'Higher, OEG, Acadex',
        'Housekeeping / Kitchen / Laundry',
        'ปกติ: 08:00 – 16:00 น. (8 ชม.)\nพีค: 07:30 – 17:00 น. (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$10 - $20',
        '44 – 48 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, รถรับส่งพนักงาน, ชมวิวยอดเขาเดนาลี บรรยากาศสงบ'
    ],
    [
        'Fairbanks Princess Riverside Lodge',
        'Alaska (AK)',
        'IEE, Acadex, Higher',
        'Food Runner / Laundry / Steward',
        'ปกติ: 06:30 – 14:30 น. หรือ 15:00 – 23:00 น.\nพีค: กะสลับตามรอบรถไฟ/เครื่องบิน 8 ชม.',
        'ฐาน $15.50 / OT $23.25',
        '$20 - $40',
        '36 – 40 ชม./สัปดาห์ (ไม่ค่อยมี OT)',
        '$105',
        'หอพักในเมืองแฟร์แบงก์ส + รถชัตเติลบัส, ใกล้ Walmart และร้านอาหารไทย หางาน 2 ง่าย'
    ],
    [
        'Grand Teton Lodge Company (GTLC)',
        'Wyoming (WY)',
        'Acadex, Higher, OEG, IEE',
        'Housekeeping / Crew Member',
        'ปกติ: 08:00 – 16:30 น. (8 ชม.)\nพีค: 07:00 – 17:30 น. (10 ชม. ควงกะ)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '54 – 60 ชม./สัปดาห์ (พีค 60-64 ชม. / หัวท้าย 40 ชม.)',
        '$0',
        'หอพักฟรี 100%! + อาหารบุฟเฟต์ 3 มื้อ EDR ($105/wk), ส่วนลดซื้อของ 40%, เข้าอุทยาน Grand Teton & Yellowstone ฟรี'
    ],
    [
        'Grand Teton Lodge Company (GTLC)',
        'Wyoming (WY)',
        'Acadex, Higher, OEG, IEE',
        'Kitchen Crew / Dishwasher / Food Prep',
        'ปกติ: 06:00 – 14:30 น. หรือ 15:00 – 23:00 น.\nพีค: 14:00 – 00:30 น. (ช่วยจัดเลี้ยง Banquet)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '52 – 58 ชม./สัปดาห์ (ควงกะ Banquet ได้)',
        '$0',
        'หอพักฟรี 100%! + อาหาร 3 มื้อ EDR + กินฟรีในครัว, สิทธิประโยชน์เครือ Vail Resorts'
    ],
    [
        'Xanterra Yellowstone Lodges',
        'Wyoming/MT',
        'OEG, IEE Thailand',
        'Hospitality Crew / Kitchen / Steward',
        'ปกติ: 07:00 – 15:30 น. หรือ 15:00 – 23:30 น.\nพีค: 06:30 – 16:30 น. (10 ชม.)',
        'ฐาน $15.70 / OT $23.55',
        '-',
        '46 – 52 ชม./สัปดาห์ (หัวท้าย 38-42 ชม.)',
        '$120',
        'รวมหอพัก + อาหารบุฟเฟต์ 3 มื้อ EDR, หอพักใหม่ Arnica, เที่ยวบ่อน้ำพุร้อนและไกเซอร์ Yellowstone ฟรี'
    ],
    [
        'Under Canvas Grand Teton / Yellowstone',
        'Wyoming (WY)',
        'Higher Education',
        'Housekeeping / Guest Service',
        'ปกติ: 08:30 – 16:30 น. (8 ชม.)\nพีค: 08:00 – 17:30 น. (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '$50 - $100',
        '42 – 46 ชม./สัปดาห์',
        '$100',
        'แคมป์กระโจมซาฟารีหรู, อาหารพนักงาน, ทิปเงินสดจากแขกไฮเอนด์ บรรยากาศเป็นกันเอง'
    ],
    [
        'Kalahari Resorts & Conventions',
        'Wisconsin (WI)',
        'OEG, Acadex, Higher',
        'Lifeguard / Housekeeping (งานหลัก)',
        'ปกติ: 09:30 – 18:00 น. หรือ 12:00 – 20:30 น.\nพีค: 09:00 – 19:30 น. (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '38 – 44 ชม./สัปดาห์ (เกลี่ยกะ 40 ชม.)',
        '$105',
        'หอพัก Kalahari Village, บัตรเล่นสวนน้ำฟรี, เมืองปั่นจักรยาน 100%, หางาน 2 ร้านอาหารง่าย'
    ],
    [
        'Wisconsin Dells Riverfront Eateries',
        'Wisconsin (WI)',
        'Walk-in สมัครเองหน้างาน',
        'Barback / Busser / Runner (งานที่สอง)',
        'ปกติ: 17:30 – 22:30 น. (5 ชม.)\nพีค: 17:00 – 00:30 น. (7.5 ชม.)',
        'ฐาน $12.00 – $14.00',
        '$250 - $500',
        '20 – 25 ชม./สัปดาห์ (กะค่ำ 17:30-23:00)',
        '-',
        'รับทิปเงินสดทุกคืน ยัดกระเป๋ากลับห้องทันที ($50-$100/คืน), อาหารพนักงานในกะ (ทำควบกับงานหลัก)'
    ],
    [
        'Dollywood Theme Park & Resorts',
        'Tennessee (TN)',
        'New Step, Acadex',
        'Ride Operator / Food Service / Retail',
        'ปกติ: 09:30 – 18:00 น. (8 ชม.)\nพีค: 09:00 – 21:30 น. (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '44 – 50 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$110',
        'ปลอดภาษีเงินได้รัฐ 0%, นั่งรถราง Pigeon Forge Trolley $1-$2.50/วัน, บัตรเข้าสวนสนุก Dollywood & Splash Country ฟรี'
    ],
    [
        'Premier Aquatics / High Sierra Pools',
        'Maryland/VA',
        'Acadex, New Step, ALC',
        'Pool Lifeguard',
        'ปกติ: 10:30 – 19:30 น. (9 ชม.)\nพีค: 10:00 – 20:30 น. (10.5 ชม. สระเปิดยาว)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 54 ชม./สัปดาห์ (การันตี OT 10-14 ชม.)',
        '$120',
        'การันตีชั่วโมง OT สม่ำเสมอ 10-14 ชม./สัปดาห์ ไม่ผันผวนตามนักท่องเที่ยว, อพาร์ตเมนต์แชร์กับเพื่อน'
    ],
    [
        'Cedar Point Amusement Park',
        'Ohio (OH)',
        'OEG, IEE Thailand',
        'Ride Operator / Food Service',
        'ปกติ: 10:00 – 18:30 น. (8 ชม.)\nพีค: 09:30 – 22:30 น. (12 ชม. สวนสนุกปิดดึก)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '46 – 52 ชม./สัปดาห์ (หัวท้าย 38-42 ชม.)',
        '$80',
        'หอพักราคาถูกมาก ($80/wk), รถบัสรับส่งฟรี, บัตรเล่นเครื่องเล่นสวนสนุกระดับโลกฟรี'
    ]
]

ws.append(headers)
for r in rows:
    ws.append(r)

# Header style
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Row styling
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(headers)), start=2):
    fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if fill.fill_type:
            cell.fill = fill
        if col_idx in [2, 6, 7, 9]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 5:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Column width auto-fit
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        # check each line if multiline
        lines = val_str.split('\n')
        for l in lines:
            length = len(l.encode('utf-8')) // 2 if any(ord(c) > 127 for c in l) else len(l)
            if length > max_len:
                max_len = length
    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Freeze top row
ws.freeze_panes = 'A2'

paths_to_save = [
    r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx',
    r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'
]

for p in paths_to_save:
    try:
        wb.save(p)
        print('Saved clean workbook with shift hours to:', p)
    except PermissionError:
        print('Could not save to locked path:', p)
