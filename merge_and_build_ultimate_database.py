# -*- coding: utf-8 -*-
import openpyxl
import os, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb_public = openpyxl.load_workbook(public_path, data_only=True)
wb_old_master = openpyxl.load_workbook(master_path, data_only=True)

wb_new = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
hyperlink_font = Font(name='Segoe UI', size=10, color='2563EB', underline='single')
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# 1. Sheet: All Public Jobs (1,167 Rows)
ws_pub_src = wb_public['All Public Jobs']
ws_pub = wb_new.active
ws_pub.title = 'All Public Jobs (1,167 Jobs)'
ws_pub.views.sheetView[0].showGridLines = True

for r in range(1, ws_pub_src.max_row + 1):
    row_vals = [ws_pub_src.cell(r, c).value for c in range(1, ws_pub_src.max_column + 1)]
    ws_pub.append(row_vals)

# 2. Sheet: Verified Summer 2027 (531 Rows)
ws_ver_src = wb_public['Verified Summer 2027']
ws_ver = wb_new.create_sheet(title='Verified Summer 2027 (531 Jobs)')
ws_ver.views.sheetView[0].showGridLines = True

for r in range(1, ws_ver_src.max_row + 1):
    row_vals = [ws_ver_src.cell(r, c).value for c in range(1, ws_ver_src.max_column + 1)]
    ws_ver.append(row_vals)

# 3. Sheet: Top Employers & 2nd Job Guide (180 Rows)
ws_top_src = wb_old_master['Tier S-A Summer Jobs']
ws_top = wb_new.create_sheet(title='Top Employers & 2nd Job Guide')
ws_top.views.sheetView[0].showGridLines = True

for r in range(1, ws_top_src.max_row + 1):
    row_vals = [ws_top_src.cell(r, c).value for c in range(1, ws_top_src.max_column + 1)]
    ws_top.append(row_vals)

# 4. Sheet: State Summary
ws_sum_src = wb_public['State Summary']
ws_sum = wb_new.create_sheet(title='State Summary')
ws_sum.views.sheetView[0].showGridLines = True

for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws_sum.append(row_vals)

# 5. Sheet: Summer Agency Directory
ws_dir_src = wb_old_master['Summer Agency Directory']
ws_dir = wb_new.create_sheet(title='Summer Agency Directory')
ws_dir.views.sheetView[0].showGridLines = True

for r in range(1, ws_dir_src.max_row + 1):
    row_vals = [ws_dir_src.cell(r, c).value for c in range(1, ws_dir_src.max_column + 1)]
    ws_dir.append(row_vals)

# 6. Sheet: Coverage & Audit
ws_aud_src = wb_public['Coverage & Audit']
ws_aud = wb_new.create_sheet(title='Coverage & Audit')
ws_aud.views.sheetView[0].showGridLines = True

for r in range(1, ws_aud_src.max_row + 1):
    row_vals = [ws_aud_src.cell(r, c).value for c in range(1, ws_aud_src.max_column + 1)]
    ws_aud.append(row_vals)

# Format Sheets 1 & 2 (Public Datasets)
for ws in [ws_pub, ws_ver]:
    ws.row_dimensions[1].height = 28
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 36
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            
            # Tier highlight
            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            # Hyperlinks for column 21 (Source URL)
            if c == 21 and cell.value and str(cell.value).startswith('http'):
                url_str = str(cell.value)
                cell.hyperlink = url_str
                cell.font = hyperlink_font

            # Alignments
            if c in [1, 3, 4, 6, 7, 8, 9, 22]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [11, 12, 13, 14, 15, 16]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    widths = [12, 16, 8, 18, 32, 16, 14, 24, 20, 35, 28, 25, 22, 22, 35, 18, 28, 22, 28, 20, 45, 14]
    for i, w in enumerate(widths, start=1):
        if i <= ws.max_column:
            ws.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 3 (Top Employers & 2nd Job Guide)
ws_top.row_dimensions[1].height = 28
for c in range(1, ws_top.max_column + 1):
    cell = ws_top.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws_top.max_row + 1):
    tier_val = str(ws_top.cell(row=r, column=1).value or '')
    ws_top.row_dimensions[r].height = 42
    for c in range(1, ws_top.max_column + 1):
        cell = ws_top.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if 'Tier S' in tier_val and c == 1:
            cell.fill = tier_s_fill
            cell.font = bold_font
        elif 'Tier A' in tier_val and c == 1:
            cell.fill = tier_a_fill
            cell.font = bold_font

        if c in [1, 2, 8, 9, 10]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

top_widths = [12, 18, 40, 30, 32, 32, 25, 14, 18, 16, 45, 42, 42, 42]
for i, w in enumerate(top_widths, start=1):
    if i <= ws_top.max_column:
        ws_top.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws_sum.max_column + 1):
    cell = ws_sum.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws_sum.row_dimensions[3].height = 28

for r in range(4, ws_sum.max_row + 1):
    ws_sum.row_dimensions[r].height = 24
    for c in range(1, ws_sum.max_column + 1):
        cell = ws_sum.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws_sum.max_column + 1):
    ws_sum.column_dimensions[get_column_letter(c)].width = 16
ws_sum.column_dimensions['B'].width = 22

# Format Sheet 5 (Summer Agency Directory)
ws_dir.row_dimensions[1].height = 28
for c in range(1, ws_dir.max_column + 1):
    cell = ws_dir.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws_dir.max_row + 1):
    ws_dir.row_dimensions[r].height = 28
    for c in range(1, ws_dir.max_column + 1):
        cell = ws_dir.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

dir_widths = [25, 20, 28, 42, 38, 16, 20, 30, 35]
for i, w in enumerate(dir_widths, start=1):
    if i <= ws_dir.max_column:
        ws_dir.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 6 (Coverage & Audit)
for c in range(1, ws_aud.max_column + 1):
    cell = ws_aud.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws_aud.row_dimensions[3].height = 28

for r in range(4, ws_aud.max_row + 1):
    ws_aud.row_dimensions[r].height = 26
    for c in range(1, ws_aud.max_column + 1):
        cell = ws_aud.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

for c in range(1, ws_aud.max_column + 1):
    ws_aud.column_dimensions[get_column_letter(c)].width = 22

# Save files
try:
    wb_new.save(master_path)
    print('Master database saved successfully to:', master_path)
except Exception as e:
    print('Master path lock exception:', e)

try:
    wb_new.save(clean_path)
    print('Clean mirror saved successfully to:', clean_path)
except Exception as e:
    print('Clean path lock exception:', e)

print('Build process completed successfully!')
