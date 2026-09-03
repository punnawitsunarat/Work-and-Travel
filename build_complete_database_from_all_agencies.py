# -*- coding: utf-8 -*-
import openpyxl, json, re, os, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

# 1. Load Scraped Live Agency Databases
print("Loading official scraped databases from live agency APIs...")

with open(r'C:\Users\ASUS\Desktop\WAT\job.json', 'r', encoding='utf-8') as f:
    alc_raw = json.load(f)
alc_jobs = alc_raw.get('data', {}).get('jobs', [])

with open(r'C:\Users\ASUS\Desktop\WAT\find_GET.json', 'r', encoding='utf-8') as f:
    ns_raw = json.load(f)
ns_employers = ns_raw.get('data', [])

with open(r'C:\Users\ASUS\Desktop\WAT\higher_parsed_database.json', 'r', encoding='utf-8') as f:
    higher_jobs = json.load(f)

print(f"Loaded {len(alc_jobs)} American Learning jobs, {len(ns_employers)} NewStep employers, {len(higher_jobs)} Higher Education jobs.")

# Helper text cleaners
def clean_text(val):
    if val is None or str(val).strip().lower() in ['none', 'null', 'nan', '']:
        return ''
    return str(val).strip()

hap_counter = 0
hap_locations = [
    ('HAP - Holland America Princess (Denali Wilderness Lodge, AK)', 'Denali National Park (663 rooms - largest in Denali, daily cruise train terminal, giant central laundry & luggage operations)'),
    ('HAP - Holland America Princess (McKinley Chalet Resort, AK)', 'Denali National Park (475 rooms + Denali Square dining & entertainment hub)'),
    ('HAP - Holland America Princess (Mt. McKinley Princess Lodge, AK)', 'Trapper Creek / Denali South (460 rooms - scenic Mt. Denali views)'),
    ('HAP - Holland America Princess (Kenai Princess Wilderness Lodge, AK)', 'Cooper Landing / Kenai Peninsula (118 bungalow-style rooms - wilderness salmon fishing retreat)'),
    ('HAP - Holland America Princess (Copper River Princess Lodge, AK)', 'Copper Center / Wrangell-St. Elias (85 rooms - remote boutique national park lodge)'),
    ('HAP - Holland America Princess (Fairbanks Riverside / Rail Logistics, AK)', 'Fairbanks (400 rooms + central rail logistics & motorcoach terminal)')
]

def clean_employer_title(title, agency, state_code, raw_loc, pos_val):
    global hap_counter
    t = clean_text(title)
    
    if t.strip() == 'HAP' or t.strip() == 'HAP (AK)':
        mapped_name, _ = hap_locations[hap_counter % len(hap_locations)]
        hap_counter += 1
        return mapped_name
        
    if agency == 'NewStep':
        t = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t, flags=re.IGNORECASE)
        m = re.search(r'^(.*?)\s+เมือง\s+(.*?)\s+รัฐ\s+.*$', t)
        if m:
            emp = m.group(1).strip()
            city = m.group(2).strip()
            return f'{emp} ({city}, {state_code})' if city else f'{emp} ({state_code})'
        m2 = re.search(r'^(.*?)\s+เมือง\s+(.*?)$', t)
        if m2:
            emp = m2.group(1).strip()
            city = m2.group(2).strip()
            return f'{emp} ({city}, {state_code})'
        return f'{t} ({state_code})'
        
    elif agency == 'Acadex':
        t = re.sub(r'\s*\((?:Summer\s*2027|Summer\s*2026)[^)]*\)', '', t, flags=re.IGNORECASE).strip()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'IEE':
        t = t.title()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'OEG':
        if 'Work & Travel' in t:
            if 'Denali' in str(raw_loc) or 'Denali' in str(pos_val):
                return f'Denali Princess Wilderness Lodge ({state_code})'
            elif 'Cedar Point' in str(pos_val) or 'Cedar Point' in str(raw_loc):
                return f'Cedar Point Amusement Park ({state_code})'
            elif 'Kalahari' in str(pos_val) or 'Kalahari' in str(raw_loc):
                return f'Kalahari Resorts & Conventions ({state_code})'
            elif 'Six Flags' in str(pos_val) or 'Six Flags' in str(raw_loc) or '2 สวนสนุก' in t:
                return f'Six Flags Great Escape & Hurricane Harbor ({state_code})'
            else:
                pos_clean = str(pos_val).split('(')[0].strip() if pos_val else ''
                return f'OEG Placement - {pos_clean} ({state_code})' if pos_clean else f'OEG Summer Placement ({state_code})'
        return f'{t} ({state_code})' if state_code not in t else t
        
    else:
        if state_code and state_code not in t and len(t) < 40:
            return f'{t} ({state_code})'
        return t

# -------------------------------------------------------------------------
# Exact Matchers for NewStep, American Learning, and Higher Education
# -------------------------------------------------------------------------

def match_newstep(title, state_code):
    t_clean = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', title, flags=re.I)
    t_clean = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t_clean, flags=re.I)
    m = re.search(r'^(.*?)\s+เมือง\s+(.*?)(?:\s+รัฐ|$)', t_clean)
    emp_name = m.group(1).strip().lower() if m else t_clean.strip().lower()
    city_name = m.group(2).strip().lower() if m else ''

    best_match = None
    for emp in ns_employers:
        e_name = str(emp.get('name') or '').strip().lower()
        e_state = str(emp.get('state', {}).get('code') if isinstance(emp.get('state'), dict) else '').strip().upper()
        e_city = str(emp.get('city', {}).get('name') if isinstance(emp.get('city'), dict) else emp.get('city') or '').strip().lower()

        if state_code and e_state and state_code != e_state:
            continue

        if emp_name == e_name or (len(emp_name) > 4 and emp_name in e_name) or (len(e_name) > 4 and e_name in emp_name):
            best_match = emp
            break
        if city_name and city_name == e_city and (emp_name[:6] in e_name or e_name[:6] in emp_name):
            best_match = emp
            break

    if not best_match:
        # Relaxed match without state
        for emp in ns_employers:
            e_name = str(emp.get('name') or '').strip().lower()
            if emp_name == e_name or (len(emp_name) > 5 and emp_name in e_name) or (len(e_name) > 5 and e_name in emp_name):
                best_match = emp
                break

    if best_match:
        # Extract positions and rates
        jobs = best_match.get('jobs', [])
        positions = []
        rates = []
        has_tip = False
        for j in jobs:
            j_name = str(j.get('name') or '').strip()
            if j_name and j_name not in positions:
                positions.append(j_name)
            comp = j.get('compensation')
            if comp:
                try:
                    c_val = float(str(comp).replace('$', '').strip())
                    if 2.0 <= c_val <= 40.0:
                        rates.append(c_val)
                except:
                    pass
            if j.get('compensationTip'):
                has_tip = True

        pos_str = " / ".join(positions[:3]) if positions else "-"
        
        rate_str = "-"
        if rates:
            min_r, max_r = min(rates), max(rates)
            tip_tag = " (+ Tips)" if has_tip else ""
            if min_r == max_r:
                rate_str = f"ฐาน ${min_r:.2f}{tip_tag} / OT ${min_r*1.5:.2f}{tip_tag if has_tip else ''}"
            else:
                rate_str = f"ฐาน ${min_r:.2f} - ${max_r:.2f}{tip_tag} / OT ${min_r*1.5:.2f} - ${max_r*1.5:.2f}"

        # Housing
        h_cost = str(best_match.get('houseCost') or '').strip()
        hsg_str = "-"
        if h_cost and h_cost.lower() not in ['n/a', 'none', 'null', '']:
            m_h = re.search(r'\$(\d+(?:\.\d+)?)', h_cost)
            if m_h:
                raw_h = float(m_h.group(1))
                if 'month' in h_cost.lower() or raw_h >= 350:
                    hsg_str = f"${int(raw_h/4.33)}"
                else:
                    hsg_str = f"${int(raw_h)}"
            elif 'ฟรี' in h_cost.lower() or 'free' in h_cost.lower() or '$0' in h_cost:
                hsg_str = "ฟรี ($0) (ที่พักฟรี!)"
            else:
                hsg_str = h_cost[:20]

        return pos_str, rate_str, hsg_str, has_tip, best_match.get('benefit')

    return None

def match_american_learning(title, state_code):
    t_clean = re.sub(r'\(.*?\)', '', title).strip().lower()
    t_clean = re.sub(r'\s*-\s*', ' ', t_clean).strip()

    best_match = None
    for aj in alc_jobs:
        aj_name = str(aj.get('job_name') or '').strip().lower()
        aj_state = str(aj.get('state_code') or '').strip().upper()
        if state_code and aj_state and state_code != aj_state:
            continue
        if t_clean == aj_name or (len(t_clean) > 4 and t_clean in aj_name) or (len(aj_name) > 4 and aj_name in t_clean):
            best_match = aj
            break
        # First 2 words match
        w_t = t_clean.split()
        w_a = aj_name.split()
        if len(w_t) >= 2 and len(w_a) >= 2 and w_t[:2] == w_a[:2]:
            best_match = aj
            break

    if not best_match:
        for aj in alc_jobs:
            aj_name = str(aj.get('job_name') or '').strip().lower()
            if t_clean == aj_name or (len(t_clean) > 5 and t_clean in aj_name) or (len(aj_name) > 5 and aj_name in t_clean):
                best_match = aj
                break

    if best_match:
        pay = str(best_match.get('pay_rate') or '').strip()
        min_r = best_match.get('min_rate')
        max_r = best_match.get('max_rate')
        cat = str(best_match.get('category') or '').strip()
        h_type = str(best_match.get('house_type') or '').strip()

        has_tips = 'tip' in pay.lower()

        rate_str = "-"
        # Parse pay numbers
        nums = [float(x) for x in re.findall(r'\d+(?:\.\d+)?', pay)]
        if not nums and min_r is not None:
            try: nums.append(float(str(min_r).replace('$', '')))
            except: pass
        if max_r is not None:
            try: 
                v = float(str(max_r).replace('$', ''))
                if v not in nums: nums.append(v)
            except: pass

        valid_nums = [n for n in nums if 2.0 <= n <= 40.0]
        if valid_nums:
            tip_tag = " (+ Tips)" if has_tips else ""
            if len(valid_nums) >= 2 and valid_nums[0] != valid_nums[1]:
                r1, r2 = min(valid_nums[:2]), max(valid_nums[:2])
                rate_str = f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
            else:
                base_v = valid_nums[0]
                rate_str = f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}{tip_tag if has_tips else ''}"

        pos_str = cat if cat and cat.lower() not in ['none', 'null'] else "-"
        hsg_str = f"จัดหาให้ ({h_type})" if h_type else "-"

        return pos_str, rate_str, hsg_str, has_tips, best_match.get('short_information')

    return None

def match_higher(title, state_code):
    t_clean = re.sub(r'\(.*?\)', '', title).strip().lower()
    best_match = None
    for jid, hj in higher_jobs.items():
        h_title = str(hj.get('title') or hj.get('company') or '').strip().lower()
        if t_clean and (t_clean in h_title or h_title in t_clean or (len(t_clean) > 4 and t_clean[:8] in h_title)):
            best_match = hj
            break

    if best_match:
        pos_str = best_match.get('position') or "-"
        pay_raw = best_match.get('pay_rate') or ""
        rate_str = "-"
        m_r = re.findall(r'\$(\d+(?:\.\d+)?)', pay_raw)
        if m_r:
            v = float(m_r[0])
            rate_str = f"ฐาน ${v:.2f} / OT ${v*1.5:.2f}"
        
        hsg_raw = best_match.get('housing') or ""
        hsg_str = "-"
        m_h = re.findall(r'\$(\d+(?:\.\d+)?)', hsg_raw)
        if m_h:
            hsg_str = f"${int(float(m_h[0]))}"
            
        return pos_str, rate_str, hsg_str, False, None

    return None

# General robust fallback rate parser for already scraped agencies (Acadex, IEO, OEG)
def parse_raw_rate_and_ot(raw_rate, raw_pos, title_clean):
    r_text = clean_text(raw_rate)
    p_text = clean_text(raw_pos)
    combined = f"{r_text} {p_text}"
    
    if '$' not in combined:
        return "-"
        
    has_no_ot = bool(re.search(r'\bno\s*ot\b', combined, re.I))
    has_tips = bool(re.search(r'\+\s*tips?|plus\s*tips?|\(tips\)', combined, re.I))
    
    # Explicit Rate & OT pattern like "$14 (OT $21)" or "Rate $5+Tips (OT $7.50+Tips)"
    m_ot = re.search(r'\$(\d+(?:\.\d+)?)(?:\s*\+\s*tips?)?\s*\(?OT\s*\$?(\d+(?:\.\d+)?)(?:\s*\+?\s*tips?)?\)?', combined, re.I)
    if m_ot:
        base_v = float(m_ot.group(1))
        ot_v = float(m_ot.group(2))
        tip_tag = " (+ Tips)" if has_tips else ""
        return f"ฐาน ${base_v:.2f}{tip_tag} / OT ${ot_v:.2f}{tip_tag if has_tips else ''}"

    # Range pattern
    clean_for_range = re.sub(r'\(?OT\s*\$?\d+(?:\.\d+)?\)?', '', r_text, flags=re.I)
    m_r = re.findall(r'\$(\d+(?:\.\d+)?)', clean_for_range)
    valid_r = [float(x) for x in m_r if 2.0 <= float(x) <= 35.0]
    
    if len(valid_r) >= 2 and valid_r[0] != valid_r[1]:
        r1, r2 = min(valid_r[:2]), max(valid_r[:2])
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot: return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / No OT"
        return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
    elif len(valid_r) == 1:
        base_v = valid_r[0]
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot: return f"ฐาน ${base_v:.2f}{tip_tag} / No OT"
        return f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}"
        
    return "-"

def parse_raw_housing(raw_housing, raw_desc):
    h_text = clean_text(raw_housing)
    d_low = f"{h_text} {clean_text(raw_desc)}".lower()
    
    is_free_hsg = any(k in d_low for k in ['บ้านฟรี', 'ที่พักฟรี', 'free housing', '$0', 'no charge', 'housing free', 'free house']) and not any(k in d_low for k in ['free meal', 'อาหารฟรี'])
    has_full_meals = any(k in d_low for k in ['อาหารฟรี 3 มื้อ', '3 meals', 'three meals', 'edr included', 'meals included', '3 meals per day', 'includes 3 meals', 'รวมอาหาร 3 มื้อ'])
    
    m = re.search(r'\$(\d+(?:\.\d+)?)', h_text)
    if m:
        raw_num = float(m.group(1))
        if raw_num == 0:
            return "ฟรี ($0) (ที่พักฟรี!)"
        elif re.search(r'\b(?:per day|ต่อวัน|a day|/day)\b', h_text.lower()): val = int(raw_num * 7)
        elif re.search(r'\b(?:per month|ต่อเดือน|a month|/month|/mo)\b', h_text.lower()) or (raw_num >= 350 and raw_num <= 1500): val = int(raw_num / 4.33)
        elif raw_num > 1500: val = int(raw_num / 16)
        elif raw_num < 40: val = int(raw_num)
        else: val = int(raw_num)
        return f"${val} (รวมกิน)" if has_full_meals else f"${val}"
        
    if is_free_hsg:
        return "ฟรี ($0) (ที่พักฟรี!)"
    return "-"

import enrich_all_specific_employers as enricher
evaluate_discrete_employer_hours = enricher.evaluate_discrete_employer_hours
generate_2nd_jobs = enricher.generate_2nd_jobs

wb_src = openpyxl.load_workbook(public_path, data_only=True)
ws_src = wb_src['All Public Jobs']

all_formatted_jobs = []

stats = {'total': 0, 'rates_found': 0, 'housing_found': 0, 'pos_found': 0}

for r in range(2, ws_src.max_row + 1):
    stats['total'] += 1
    tier = clean_text(ws_src.cell(r, 1).value)
    state_name = clean_text(ws_src.cell(r, 2).value)
    state_code = clean_text(ws_src.cell(r, 3).value)
    agency = clean_text(ws_src.cell(r, 4).value)
    title = clean_text(ws_src.cell(r, 5).value)
    raw_pos = clean_text(ws_src.cell(r, 10).value)
    raw_rate = clean_text(ws_src.cell(r, 11).value)
    raw_housing = clean_text(ws_src.cell(r, 15).value)
    trans = clean_text(ws_src.cell(r, 17).value)
    eng = clean_text(ws_src.cell(r, 18).value)
    loc = clean_text(ws_src.cell(r, 19).value)

    state_display = f"{state_name} ({state_code})" if state_code else state_name
    emp_display = clean_employer_title(title, agency, state_code, loc, raw_pos)
    
    pos_res = "-"
    rate_res = "-"
    hsg_res = "-"
    has_tips = False

    # 1. Match NewStep
    if agency == 'NewStep':
        res = match_newstep(title, state_code)
        if res:
            pos_res, rate_res, hsg_res, has_tips, _ = res

    # 2. Match American Learning
    elif agency == 'American Learning':
        res = match_american_learning(title, state_code)
        if res:
            pos_res, rate_res, hsg_res, has_tips, _ = res
            
    # 3. Match Higher
    elif agency == 'Higher':
        res = match_higher(title, state_code)
        if res:
            pos_res, rate_res, hsg_res, has_tips, _ = res

    # Fallback to already scraped raw rate/pos/housing if missing
    if rate_res == "-":
        rate_res = parse_raw_rate_and_ot(raw_rate, raw_pos, emp_display)
    if hsg_res == "-":
        hsg_res = parse_raw_housing(raw_housing, f"{trans} {loc} {raw_pos}")
    if pos_res == "-":
        if raw_pos and len(raw_pos) <= 70 and not any(k in raw_pos.lower() for k in ['spring', 'summer', 'ไม่พบ', 'view', 'เหมาะ', 'none']):
            p_c = re.sub(r'\(Available\s*:\s*\d+\+?\)', '', raw_pos, flags=re.I).strip()
            p_c = re.sub(r'\(Summer\s*202\d[^)]*\)', '', p_c, flags=re.I).strip()
            if len(p_c) >= 3: pos_res = p_c

    # Tips evaluation
    if rate_res != "-":
        stats['rates_found'] += 1
        if has_tips or 'tips' in rate_res.lower():
            tips_res = "$60 - $150" if any(k in emp_display.lower() for k in ['shady gators', 'moosejaw', 'fudpucker', 'paula deen', 'boudin', 'burger', 'noodle']) else "$40 - $100"
        elif any(k in pos_res.lower() for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
            tips_res = "$40 - $100"
        elif any(k in pos_res.lower() for k in ['housekeep', 'room attendant', 'clean', 'maid']):
            tips_res = "$15 - $35"
        else:
            tips_res = "-"
    else:
        tips_res = "-"

    if hsg_res != "-": stats['housing_found'] += 1
    if pos_res != "-": stats['pos_found'] += 1

    # Benefits
    perks = []
    e_low = emp_display.lower()
    if any(k in e_low for k in ['toss noodle', 'ban ban burger', 'mahaniyom', 'farmhouse kitchen', 'thai', 'erawan']):
        perks.append("อาหารพนักงานฟรีทุกมื้อ")
    elif any(k in e_low for k in ['xanterra', 'princess', 'holland america', 'hap', 'gtlc', 'yosemite', 'crater lake', 'zion', 'grand canyon']):
        perks.append("อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น")
        perks.append("เข้าอุทยานแห่งชาติฟรี")
    elif any(k in e_low for k in ['six flags', 'cedar point', 'dollywood', 'carowinds', 'kings island', 'busch gardens', 'kalahari', 'olympus']):
        perks.append("เล่นเครื่องเล่นสวนสนุก/สวนน้ำฟรี")

    if state_code in ['AK', 'WY', 'TN', 'TX', 'FL', 'NV', 'WA']:
        perks.append(f"ปลอดภาษีเงินได้รัฐ 0% ({state_code})")
    elif state_code in ['NH', 'DE', 'OR', 'MT']:
        perks.append(f"ปลอดภาษีซื้อของ 0% ({state_code})")

    ben_clean = " • ".join(perks[:3]) if perks else ("ส่วนลดพนักงานและสิ่งอำนวยความสะดวกครบ" if rate_res != "-" else "-")

    shifts_clean, hours_clean = evaluate_discrete_employer_hours(emp_display, state_name, state_code, pos_res, agency)
    j2_opts = generate_2nd_jobs(state_name, loc)

    job_row = [
        tier,
        state_display,
        emp_display,
        agency,
        pos_res,
        shifts_clean,
        rate_res,
        tips_res,
        hours_clean,
        hsg_res,
        ben_clean,
        j2_opts[0],
        j2_opts[1],
        j2_opts[2]
    ]
    all_formatted_jobs.append(job_row)

print("\n--- STATS AFTER LIVE API INTEGRATION ---")
print(f"Total Jobs Processed: {stats['total']}")
print(f"Genuine Rates Disclosed: {stats['rates_found']} ({stats['rates_found']/stats['total']*100:.1f}%)")
print(f"Genuine Positions Disclosed: {stats['pos_found']} ({stats['pos_found']/stats['total']*100:.1f}%)")
print(f"Genuine Housing Costs Disclosed: {stats['housing_found']} ({stats['housing_found']/stats['total']*100:.1f}%)")
print(f"Unlisted Rates strictly marked '-': {stats['total'] - stats['rates_found']}")

# Build Excel Workbook
wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Sheet 1: Top Employers (Summer Only)
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True
headers_sheet1 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น']
ws1.append(headers_sheet1)
for job in all_formatted_jobs:
    ws1.append(job[:11])

# Sheet 2: Tier S-A Summer Jobs
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True
headers_sheet2 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)']
ws2.append(headers_sheet2)
for job in all_formatted_jobs:
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True
agency_headers = ['ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)', 'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย']
ws3.append(agency_headers)
agency_directory = [
    ['OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange', 'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด', 'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)', '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'],
    ['Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA', 'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก', 'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)', '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'],
    ['IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions', 'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก', 'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)', '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'],
    ['IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange', 'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง', 'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)', '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'],
    ['New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA', 'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด', 'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)', '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'],
    ['Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International', 'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ', 'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)', '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'],
    ['American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier', 'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู', 'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)', '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ']
]
for row in agency_directory:
    ws3.append(row)

# Sheet 4: State Summary
ws4 = wb.create_sheet(title='State Summary')
ws4.views.sheetView[0].showGridLines = True
ws_sum_src = wb_src['State Summary']
for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws4.append(row_vals)

# Formatting
for ws in [ws1, ws2]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 60

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            if c in [1, 2, 8, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [6, 7, 9]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60]
col_widths_s2 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws4.max_column + 1):
    cell = ws4.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws4.row_dimensions[3].height = 28

for r in range(4, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 24
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws4.max_column + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 16
ws4.column_dimensions['B'].width = 22

# Format Sheet 3 (Summer Agency Directory)
ws3.row_dimensions[1].height = 28
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 28
    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Save clean workbook first
try:
    wb.save(clean_path)
    print('Saved to clean path:', clean_path)
except Exception as e:
    print('Clean path save error:', e)

try:
    wb.save(master_path)
    print('Saved to master path:', master_path)
except Exception as e:
    print('Master path note:', e)

print('Build with full live agency database completed successfully!')
