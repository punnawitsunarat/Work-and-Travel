import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid') # Amber tint
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid') # Sky blue tint
zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# ==============================================================================
# 📊 SHEET 1: Top Employers Masterlist (ตารางงานเด่น 10 คอลัมน์ คลีนๆ)
# ==============================================================================
ws1 = wb.active
ws1.title = 'Top Employers Masterlist'
ws1.views.sheetView[0].showGridLines = True

headers1 = [
    'ชื่องาน / สถานที่ทำงาน (Employer)',
    'รัฐ (State)',
    'Agency ที่ถือสัญญา',
    'ตำแหน่ง (Position / Role)',
    '⏱️ เวลาทำงานจริง (ช่วงปกติ / ช่วงพีค)',
    'ค่าแรงฐาน / OT ($/ชม.)',
    'ทิป ($/สัปดาห์)',
    'ชั่วโมงทำงานจริง (คิดเผื่อหัว-ท้ายซีซัน)',
    'ค่าที่พัก ($/สัปดาห์)',
    'สวัสดิการ (อาหาร ฟรี หรือ อื่นๆ)'
]

rows1 = [
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher, IEO',
        'Housekeeping / Room Attendant',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีค: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $16.00 / OT $24.00',
        '$15 - $30',
        '50 – 54 ชม./สัปดาห์ (พีค 58-62 ชม. / หัวท้าย 35-42 ชม.)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน, ทิปหัวเตียง, รถรับส่งพนักงาน, ส่วนลดทัวร์/รถไฟ 50%'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher, IEO',
        'Kitchen Stewarding / Dishwasher (ล้างจาน)',
        'ปกติ: 16:00 – 00:00 (8 ชม.)\nพีค: 15:00 – 01:30 (10.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 55 ชม./สัปดาห์ (ควงกะค่ำได้ง่าย)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อใน EDR + เชฟทำสเต๊ก/เบอร์เกอร์เลี้ยงในครัวฟรีทุกกะ'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'Acadex, OEG, IEE, Higher, IEO',
        'Luggage Handler (พนักงานยกกระเป๋าทัวร์)',
        'ปกติ: 08:00 – 16:30 หรือ 12:00 – 20:30\nพีค: 07:00 – 18:30 (ตามรอบรถไฟ)',
        'ฐาน $15.50 / OT $23.25',
        '$40 - $70',
        '46 – 52 ชม./สัปดาห์',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, ทิปเงินสดจากกรุ๊ปทัวร์ผู้สูงอายุ'
    ],
    [
        'Denali Princess Wilderness Lodge',
        'Alaska (AK)',
        'ขอทำเพิ่มหน้างาน (Internal 2nd Job)',
        'Fleet Detailer (งานเสริมล้างรถทัวร์กะดึก)',
        'ปกติ: 21:00 – 01:30 (4.5 ชม.)\nพีค: 21:00 – 02:30 (5.5 ชม.)',
        'คิดเรต OT $24.00/ชม. ทันที',
        '-',
        '20 – 25 ชม./สัปดาห์ (ทำควบหลังงานหลักแม่บ้าน)',
        '-',
        'นับเป็นชั่วโมง OT $24.00 ทั้งหมด, ทำงานในอู่รถบัส Princess, อาหาร EDR 3 มื้อ'
    ],
    [
        'Denali Bluffs Hotel',
        'Alaska (AK)',
        'Acadex Thailand, IEO',
        'Housekeeping / Laundry',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 08:00 – 17:00 (9 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '40 – 45 ชม./สัปดาห์ (หัวท้าย 32-38 ชม.)',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป (ครบสัญญา), รถชัตเติลบัสขึ้น-ลงเขา Sugarloaf, ส่วนลดอาหารร้าน The Perk'
    ],
    [
        'Grande Denali Lodge',
        'Alaska (AK)',
        'Acadex Thailand, IEO',
        'Housekeeping / Public Area Cleaner',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25',
        '-',
        '42 – 46 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป (ครบสัญญา), โรงแรมหรู 4 ดาว, รถชัตเติลบัส, ส่วนลดอาหารห้องอาหาร Alpenglow'
    ],
    [
        'Grande Denali Lodge',
        'Alaska (AK)',
        'Acadex Thailand, IEO',
        'Busser / Food Runner (Alpenglow Restaurant)',
        'ปกติ: 16:30 – 22:30 (6 ชม.)\nพีค: 16:00 – 23:30 (7.5 ชม.)',
        'ฐาน $14.00 / OT $21.00',
        '$200 - $350',
        '40 – 45 ชม./สัปดาห์',
        '$140',
        'โบนัสพิเศษ +$1.00/ชม. ท้ายทริป, ทิปเงินสดจากแขกไฮเอนด์ร้านอาหารวิวพาโนรามา ($40-$70/คืน)'
    ],
    [
        'Mt. McKinley Princess Lodge',
        'Alaska (AK)',
        'Higher, OEG, Acadex, IEO',
        'Housekeeping / Kitchen / Laundry',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีค: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00',
        '$10 - $20',
        '44 – 48 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$105',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, รถรับส่งพนักงาน, ชมวิวยอดเขาเดนาลี บรรยากาศสงบ'
    ],
    [
        'Fairbanks Princess Riverside Lodge',
        'Alaska (AK)',
        'IEE, Acadex, Higher, IEO',
        'Food Runner / Laundry / Steward',
        'ปกติ: 06:30 – 14:30 หรือ 15:00 – 23:00\nพีค: กะสลับตามรอบรถไฟ 8 ชม.',
        'ฐาน $15.50 / OT $23.25',
        '$20 - $40',
        '36 – 40 ชม./สัปดาห์ (ไม่ค่อยมี OT)',
        '$105',
        'หอพักในเมืองแฟร์แบงก์ส + รถชัตเติลบัส, ใกล้ Walmart และร้านอาหารไทย หางาน 2 ง่าย'
    ],
    [
        'Grand Teton Lodge Company (GTLC)',
        'Wyoming (WY)',
        'Acadex, Higher, OEG, IEE, IEO',
        'Housekeeping / Crew Member',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีค: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '54 – 60 ชม./สัปดาห์ (พีค 60-64 ชม. / หัวท้าย 40 ชม.)',
        '$0',
        'หอพักฟรี 100%! + อาหารบุฟเฟต์ 3 มื้อ EDR ($105/wk), ส่วนลดซื้อของ 40%, เข้าอุทยาน Grand Teton & Yellowstone ฟรี'
    ],
    [
        'Grand Teton Lodge Company (GTLC)',
        'Wyoming (WY)',
        'Acadex, Higher, OEG, IEE, IEO',
        'Kitchen Crew / Dishwasher / Food Prep',
        'ปกติ: 06:00 – 14:30 หรือ 15:00 – 23:00\nพีค: 14:00 – 00:30 (ช่วย Banquet)',
        'ฐาน $17.75 / OT $26.62',
        '-',
        '52 – 58 ชม./สัปดาห์ (ควงกะ Banquet ได้)',
        '$0',
        'หอพักฟรี 100%! + อาหาร 3 มื้อ EDR + กินฟรีในครัว, สิทธิประโยชน์เครือ Vail Resorts'
    ],
    [
        'Xanterra Yellowstone Lodges',
        'Wyoming/MT',
        'OEG, IEE Thailand, IEO',
        'Hospitality Crew / Kitchen / Steward',
        'ปกติ: 07:00 – 15:30 หรือ 15:00 – 23:30\nพีค: 06:30 – 16:30 (10 ชม.)',
        'ฐาน $15.70 / OT $23.55',
        '-',
        '46 – 52 ชม./สัปดาห์ (หัวท้าย 38-42 ชม.)',
        '$120',
        'รวมหอพัก + อาหารบุฟเฟต์ 3 มื้อ EDR, หอพักใหม่ Arnica, เที่ยวบ่อน้ำพุร้อนและไกเซอร์ Yellowstone ฟรี'
    ],
    [
        'Under Canvas Grand Teton / Yellowstone',
        'Wyoming (WY)',
        'Higher Education, IEO',
        'Housekeeping / Guest Service',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีค: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75',
        '$50 - $100',
        '42 – 46 ชม./สัปดาห์',
        '$100',
        'แคมป์กระโจมซาฟารีหรู, อาหารพนักงาน, ทิปเงินสดจากแขกไฮเอนด์ บรรยากาศเป็นกันเอง'
    ],
    [
        'Kalahari Resorts & Conventions',
        'Wisconsin (WI)',
        'OEG, Acadex, Higher, IEO, New Step',
        'Lifeguard / Housekeeping (งานหลัก)',
        'ปกติ: 09:30 – 18:00 หรือ 12:00 – 20:30\nพีค: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '38 – 44 ชม./สัปดาห์ (เกลี่ยกะ 40 ชม.)',
        '$105',
        'หอพัก Kalahari Village, บัตรเล่นสวนน้ำฟรี, เมืองปั่นจักรยาน 100%, หางาน 2 ร้านอาหารง่าย'
    ],
    [
        'Wisconsin Dells Riverfront Eateries',
        'Wisconsin (WI)',
        'Walk-in สมัครเองหน้างาน',
        'Barback / Busser / Runner (งานที่สอง)',
        'ปกติ: 17:30 – 22:30 (5 ชม.)\nพีค: 17:00 – 00:30 (7.5 ชม.)',
        'ฐาน $12.00 – $14.00',
        '$250 - $500',
        '20 – 25 ชม./สัปดาห์ (กะค่ำ 17:30-23:00)',
        '-',
        'รับทิปเงินสดทุกคืน ยัดกระเป๋ากลับห้องทันที ($50-$100/คืน), อาหารพนักงานในกะ (ทำควบกับงานหลัก)'
    ],
    [
        'Dollywood Theme Park & Resorts',
        'Tennessee (TN)',
        'New Step, Acadex, IEO',
        'Ride Operator / Food Service / Retail',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีค: 09:00 – 21:30 (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.00 / OT $22.50',
        '-',
        '44 – 50 ชม./สัปดาห์ (หัวท้าย 35-40 ชม.)',
        '$110',
        'ปลอดภาษีเงินได้รัฐ 0%, นั่งรถราง Pigeon Forge Trolley $1-$2.50/วัน, บัตรเข้าสวนสนุก Dollywood & Splash Country ฟรี'
    ],
    [
        'Premier Aquatics / High Sierra Pools',
        'Maryland/VA',
        'Acadex, New Step, ALC, IEO',
        'Pool Lifeguard',
        'ปกติ: 10:30 – 19:30 (9 ชม.)\nพีค: 10:00 – 20:30 (10.5 ชม. สระเปิดยาว)',
        'ฐาน $16.00 / OT $24.00',
        '-',
        '50 – 54 ชม./สัปดาห์ (การันตี OT 10-14 ชม.)',
        '$120',
        'การันตีชั่วโมง OT สม่ำเสมอ 10-14 ชม./สัปดาห์ ไม่ผันผวนตามนักท่องเที่ยว, อพาร์ตเมนต์แชร์กับเพื่อน'
    ],
    [
        'Cedar Point Amusement Park',
        'Ohio (OH)',
        'OEG, IEE Thailand, IEO',
        'Ride Operator / Food Service',
        'ปกติ: 10:00 – 18:30 (8 ชม.)\nพีค: 09:30 – 22:30 (12 ชม. สวนสนุกปิดดึก)',
        'ฐาน $14.50 / OT $21.75',
        '-',
        '46 – 52 ชม./สัปดาห์ (หัวท้าย 38-42 ชม.)',
        '$80',
        'หอพักราคาถูกมาก ($80/wk), รถบัสรับส่งฟรี, บัตรเล่นเครื่องเล่นสวนสนุกระดับโลกฟรี'
    ]
]

ws1.append(headers1)
for r in rows1:
    ws1.append(r)

for col_idx in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row in enumerate(ws1.iter_rows(min_row=2, max_row=len(rows1)+1, min_col=1, max_col=len(headers1)), start=2):
    fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if fill.fill_type:
            cell.fill = fill
        if col_idx in [2, 6, 7, 9]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx == 5:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

for col in ws1.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        lines = val_str.split('\n')
        for l in lines:
            length = len(l.encode('utf-8')) // 2 if any(ord(c) > 127 for c in l) else len(l)
            if length > max_len:
                max_len = length
    ws1.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 45)

ws1.freeze_panes = 'C2'


# ==============================================================================
# 📊 SHEET 2: Tier S-A Jobs & 2nd Options (38 งานหลัก + งาน 2 ครบ 13 คอลัมน์)
# ==============================================================================
# Import rows from generate_tier_sa_sheet
import generate_tier_sa_sheet
ws2 = wb.create_sheet(title='Tier S-A Jobs & 2nd Options')
ws2.views.sheetView[0].showGridLines = True

headers2 = generate_tier_sa_sheet.headers
rows2 = generate_tier_sa_sheet.rows

ws2.append(headers2)
for r in rows2:
    ws2.append(r)

for col_idx in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row in enumerate(ws2.iter_rows(min_row=2, max_row=len(rows2)+1, min_col=1, max_col=len(headers2)), start=2):
    tier_text = str(row[0].value or '')
    if 'Tier S' in tier_text:
        row_fill = tier_s_fill
    else:
        row_fill = tier_a_fill
    
    if row_idx % 2 == 0:
        fill_to_use = row_fill
    else:
        fill_to_use = PatternFill(fill_type=None)

    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
        if col_idx in [1, 6, 7, 8, 9]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx in [5, 10, 11, 12, 13]:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

for col in ws2.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        lines = val_str.split('\n')
        for l in lines:
            length = len(l.encode('utf-8')) // 2 if any(ord(c) > 127 for c in l) else len(l)
            if length > max_len:
                max_len = length
    ws2.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 55)

ws2.freeze_panes = 'C2'


# ==============================================================================
# 📊 SHEET 3: 12-Week Cashflow Simulator
# ==============================================================================
ws3 = wb.create_sheet(title='12-Week Cashflow Simulator')
ws3.views.sheetView[0].showGridLines = True

headers3 = [
    'สัปดาห์ (Week)', 'สถานการณ์ & รายละเอียดกะงาน', 'ชม. งานหลัก (ชม.)', 'ชม. งานสอง (ชม.)',
    'รวม ชม./สัปดาห์', 'รายรับงานหลัก ($)', 'รายรับงานสอง ($)', 'เงินทิปสด ($)',
    'รายรับรวมรายสัปดาห์ ($)', 'รายรับรวม (บาท)', 'หักภาษี Fed 10% ($)', 'หักหอพัก/EDR ($)',
    'หักของใช้ส่วนตัว ($)', 'เงินสดสุทธิสัปดาห์นั้น ($)', 'เงินสดสะสมคงเหลือ ($)', 'เงินสดสะสมคงเหลือ (บาท)'
]

sim_data = [
    ['Week 1', 'ปฐมนิเทศ, อบรม, ทำ SSN ที่แฟร์แบงก์ส', 32, 0, 32, 512, 0, 0, 512, 16819, 51.2, 105, 25, 330.8, 330.8, 10867],
    ['Week 2', 'เริ่มแม่บ้านจริง, เรือสำราญรอบแรก, สมัครงาน 2', 38, 0, 38, 608, 0, 0, 608, 19973, 60.8, 105, 25, 417.2, 748.0, 24572],
    ['Week 3', 'รถไฟทัวร์แน่น, งานสองเริ่มทำ 2-3 วัน', 48, 10, 58, 832, 150, 60, 1042, 34230, 104.2, 105, 25, 807.8, 1555.8, 51108],
    ['Week 4', 'เข้าที่เต็มตัว, งานสอง Barback 4 วัน', 54, 14, 68, 976, 210, 120, 1306, 42902, 130.6, 105, 25, 1045.4, 2601.2, 85449],
    ['Week 5', 'PEAK จัด วันชาติ 4th of July แขกล้น', 58, 15, 73, 1072, 225, 160, 1457, 47862, 145.7, 105, 25, 1181.3, 3782.5, 124255],
    ['Week 6', 'PEAK ต่อเนื่อง เรือสำราญเข้าทุกวัน', 56, 14, 70, 1024, 210, 140, 1374, 45136, 137.4, 105, 25, 1106.6, 4889.1, 160607],
    ['Week 7', 'ฝนตกหนัก/รถไฟดีเลย์ ร่างกายเริ่มล้า', 44, 12, 56, 736, 180, 80, 996, 32719, 99.6, 105, 25, 766.4, 5655.5, 185783],
    ['Week 8', 'ฟื้นตัวกลับมาพีค ทัวร์ครอบครัวแน่น', 55, 14, 69, 1000, 210, 140, 1350, 44348, 135.0, 105, 25, 1085.0, 6740.5, 221425],
    ['Week 9', 'ช่วงท้ายพีคซีซัน ชั่วโมงยังแน่น', 52, 12, 64, 928, 180, 120, 1228, 40340, 122.8, 105, 25, 975.2, 7715.7, 253461],
    ['Week 10', 'อากาศเย็นลง นักศึกษาฝรั่งเริ่มกลับ', 48, 10, 58, 832, 150, 100, 1082, 35544, 108.2, 105, 25, 843.8, 8559.5, 281180],
    ['Week 11', 'ทัวร์เริ่มซา ปิดห้องพักบางโซน', 44, 8, 52, 736, 120, 60, 916, 30091, 91.6, 105, 25, 694.4, 9253.9, 303991],
    ['Week 12', 'สัปดาห์สุดท้าย เคลียร์ห้องพัก คืนมัดจำ', 36, 0, 36, 576, 0, 0, 576, 18922, 57.6, 105, 25, 388.4, 9642.3, 316750]
]

ws3.append(headers3)
for row in sim_data:
    ws3.append(row)

for col_num in range(1, len(headers3) + 1):
    cell = ws3.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row in enumerate(ws3.iter_rows(min_row=2, max_row=len(sim_data)+1, min_col=1, max_col=len(headers3)), start=2):
    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if col_idx in [1, 3, 4, 5]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_idx >= 6:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if col_idx in [6, 7, 8, 9, 11, 12, 13, 14, 15]:
                cell.number_format = '$#,##0.00'
            elif col_idx in [10, 16]:
                cell.number_format = '#,##0฿'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

for col in ws3.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        length = len(val_str.encode('utf-8')) // 2 if any(ord(c) > 127 for c in val_str) else len(val_str)
        if length > max_len:
            max_len = length
    ws3.column_dimensions[col_letter].width = max(max_len + 4, 15)

ws3.freeze_panes = 'B2'


# ==============================================================================
# 📊 SHEET 4: Agency Directory & Contacts (6 เอเจนซี่)
# ==============================================================================
ws4 = wb.create_sheet(title='Agency Directory & Contacts')
ws4.views.sheetView[0].showGridLines = True

headers4 = [
    'ชื่อเอเจนซี่ (Agency)', 'ระดับ (Tier)', 'ค่าโครงการโดยประมาณ (บาท)', 'สปอนเซอร์สหรัฐฯ',
    'งานเด่นใน Tier S & Tier A', 'เบอร์โทรศัพท์', 'LINE Official',
    'ที่ตั้งสำนักงาน (Office Address)', 'จุดเด่น & คำแนะนำ'
]

data4 = [
    ['OEG (Overseas Ed Group)', 'Tier S (Mass)', '106,900 - 111,900 ฿', 'CIEE (อันดับ 1 สหรัฐฯ), Spirit', 'Grand Teton (WY), Yellowstone (WY), Denali (AK), Kalahari (WI), Schlitterbahn (TX)', '02-263-3666', '@oegworkandtravel', 'อาคารสินธร ทาวเวอร์ 1 ชั้น 7 ถ.วิทยุ กทม. & เชียงใหม่', 'สปอนเซอร์ CIEE มั่นคงที่สุด ระบบติวสัมภาษณ์และดูแลมาตรฐานสูงสุด'],
    ['Acadex Thailand', 'Tier S (Boutique)', '69,000 - 82,000 ฿', 'Intrax, CCUSA, CHI, IENA', 'Grand Teton (WY), Denali Princess & Bluffs (AK), Kalahari (WI), Dollywood (TN), Lifeguards (MD)', '086-390-0333', '@AcadexThailand', 'อาคาร ซี.ซี.ที. (CCT Bldg) ชั้น 12A ถ.สุรวงศ์ บางรัก กทม.', 'พอร์ตงานอลาสกาและไวโอมิงครบที่สุด มีทั้ง Princess และ Bluffs ค่าโครงการคุ้มค่า'],
    ['IEE Thailand', 'Tier S (Boutique)', '68,000 - 78,000 ฿', 'CCUSA, InterExchange, GeoVisions', 'Xanterra Yellowstone (WY), Mount Rushmore (SD), Chula Vista (WI), Gatlinburg (TN), Cedar Point (OH)', '02-612-9511', '@IEEThailand', 'อาคารพญาไทพลาซ่า ชั้น 12 ถ.พญาไท ราชเทวี กทม.', 'ค่าโครงการย่อมเยา เครือข่ายสปอนเซอร์ CCUSA แน่นแฟ้นกับงานอุทยานและรีสอร์ต'],
    ['IEO Abroad (I.E.O.)', 'Tier S (Boutique)', '68,000 - 79,000 ฿', 'CIEE, Intrax, InterExchange', 'Denali (AK), Yellowstone (WY), Kalahari (WI), Dollywood (TN), Morey’s Piers (NJ), Universal (FL)', '061-426-2299', '@ieoworkandtravel', 'อาคาร K.A.N Place ซ.นราธิวาสราชนครินทร์ 8 สาทร กทม.', 'พอร์ตงานหลากหลาย มีงานสวนสนุกและรีสอร์ตทั้งฝั่งตะวันตกและตะวันออก'],
    ['Higher Education (Higher)', 'Tier S (Boutique)', '66,000 - 76,000 ฿', 'IENA, Spirit, Janus', 'Grand Teton (WY), Under Canvas (WY/MT), Mt. McKinley (AK), Custer State Park (SD), Estes Park (CO)', '02-054-9544', '@HigherEducation', 'อาคารจัตุรัสจามจุรี (Chamchuri Square) ชั้น 24 ปทุมวัน กทม.', 'ดูแลนักศึกษาใกล้ชิดเป็นกันเอง ตอบแชทเร็ว ค่าโครงการสบายกระเป๋า'],
    ['New Step Thailand', 'Tier S (Boutique)', '67,000 - 79,000 ฿', 'Intrax, CHI, AWA', 'Dollywood (TN), The Island (TN), Busch Gardens (VA/FL), Ocean City (MD), Six Flags (TX)', '063-535-9463', '@newstepworktravel', 'อาคาร SiamScape ชั้น 19 ห้อง 1910 สยามสแควร์ กทม.', 'เชี่ยวชาญงานสายสวนสนุกและงานชายฝั่งทะเล ออฟฟิศสยามเดินทางสะดวก']
]

ws4.append(headers4)
for row in data4:
    ws4.append(row)

for col_num in range(1, len(headers4) + 1):
    cell = ws4.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row_idx, row in enumerate(ws4.iter_rows(min_row=2, max_row=len(data4)+1, min_col=1, max_col=len(headers4)), start=2):
    for col_idx, cell in enumerate(row, start=1):
        cell.font = regular_font
        cell.border = thin_border
        if col_idx in [2, 6, 7]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

for col in ws4.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        length = len(val_str.encode('utf-8')) // 2 if any(ord(c) > 127 for c in val_str) else len(val_str)
        if length > max_len:
            max_len = length
    ws4.column_dimensions[col_letter].width = max(max_len + 4, 15)

ws4.freeze_panes = 'B2'

# Save to both paths
wb.save(excel_path)
wb.save(excel_clean_path)
print('Successfully rebuilt all 4 sheets in Master Excel workbook!')
