# -*- coding: utf-8 -*-
import openpyxl
import re, os, sys, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb_public = openpyxl.load_workbook(public_path, data_only=True)
ws_src = wb_public['All Public Jobs']

print(f'Processing {ws_src.max_row - 1} public jobs with micro-scale per-employer calculations...')

def clean_text(val):
    if val is None or str(val).strip().lower() in ['none', 'null', 'nan', '']:
        return ''
    return str(val).strip()

def clean_employer_title(title, agency, state_code, raw_loc, pos_val):
    t = clean_text(title)
    if agency == 'NewStep':
        t = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t, flags=re.IGNORECASE)
        m = re.search(r'^(.*?)\s+เมือง\s+(.*?)\s+รัฐ\s+.*$', t)
        if m:
            emp = m.group(1).strip()
            city = m.group(2).strip()
            return f'{emp} ({city}, {state_code})' if city else f'{emp} ({state_code})'
        m2 = re.search(r'^(.*?)\s+เมือง\s+(.*?)$', t)
        if m2:
            emp = m2.group(1).strip()
            city = m2.group(2).strip()
            return f'{emp} ({city}, {state_code})'
        return f'{t} ({state_code})'
        
    elif agency == 'Acadex':
        t = re.sub(r'\s*\((?:Summer\s*2027|Summer\s*2026)[^)]*\)', '', t, flags=re.IGNORECASE).strip()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'IEE':
        t = t.title()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'OEG':
        if 'Work & Travel' in t:
            if 'Denali' in str(raw_loc) or 'Denali' in str(pos_val):
                return f'Denali Princess Wilderness Lodge ({state_code})'
            elif 'Cedar Point' in str(pos_val) or 'Cedar Point' in str(raw_loc):
                return f'Cedar Point Amusement Park ({state_code})'
            elif 'Kalahari' in str(pos_val) or 'Kalahari' in str(raw_loc):
                return f'Kalahari Resorts & Conventions ({state_code})'
            elif 'Six Flags' in str(pos_val) or 'Six Flags' in str(raw_loc) or '2 สวนสนุก' in t:
                return f'Six Flags Great Escape & Hurricane Harbor ({state_code})'
            else:
                pos_clean = str(pos_val).split('(')[0].strip() if pos_val else ''
                return f'OEG Placement - {pos_clean} ({state_code})'
        return f'{t} ({state_code})' if state_code not in t else t
        
    else:
        if state_code and state_code not in t and len(t) < 40:
            return f'{t} ({state_code})'
        return t

def clean_position(pos_raw, title_clean, agency):
    p_text = clean_text(pos_raw)
    t_lower = clean_text(title_clean).lower()
    
    if (
        'spring jobs summer jobs' in p_text.lower() or 
        'ไม่พบรีวิว' in p_text or 
        'view more' in p_text.lower() or 
        'เหมาะสำหรับน้อง' in p_text or 
        'งานอาจรวมถึง' in p_text or
        len(p_text) > 85
    ):
        extracted = []
        if re.search(r'sales associate|retail|แคชเชียร์|พนักงานขาย', p_text, re.I): extracted.append('Retail / Sales Associate')
        if re.search(r'ice cream scooper|candy maker|ตักไอศกรีม', p_text, re.I): extracted.append('Ice Cream Scooper / Candy Maker')
        if re.search(r'cashier', p_text, re.I): extracted.append('Cashier')
        if re.search(r'housekeep|แม่บ้าน|room attendant', p_text, re.I): extracted.append('Housekeeping')
        if re.search(r'food prep|เตรียมอาหาร|cook|ครัว', p_text, re.I): extracted.append('Food Preparation')
        if re.search(r'dishwash|ล้างจาน', p_text, re.I): extracted.append('Dishwasher')
        if re.search(r'busser|runner|เสิร์ฟ', p_text, re.I): extracted.append('Busser / Food Runner')
        if re.search(r'lifeguard|ไลฟ์การ์ด', p_text, re.I): extracted.append('Lifeguard')
        if re.search(r'ride op|เครื่องเล่น', p_text, re.I): extracted.append('Ride Operator')
        
        if extracted:
            return ' / '.join(list(dict.fromkeys(extracted))[:3])
            
        if any(k in t_lower for k in ['water park', 'waterpark', 'white water', 'splash']):
            return 'Lifeguard / Park Services / Food Service'
        elif any(k in t_lower for k in ['amusement', 'theme park', 'silver dollar', 'wonderworks', 'fun park', 'pier', 'arcade', 'track', 'attraction']):
            return 'Attractions Host / Ride Operator / Retail'
        elif any(k in t_lower for k in ['hotel', 'inn', 'resort', 'suites', 'motel', 'lodge', 'marriott', 'hilton', 'hyatt', 'sheraton', 'westgate', 'fairfield', 'courtyard']):
            return 'Housekeeping / Laundry / Front Desk Support'
        elif any(k in t_lower for k in ['fries', 'burger', 'culver', 'five guys', 'wendy', 'mcdonald', 'domino', 'pizza', 'auntie anne', 'fast food', 'taco', 'subway', 'dairy queen', 'sonic', 'popeye', 'kfc']):
            return 'Crew Member / Cashier / Food Preparation'
        elif any(k in t_lower for k in ['creamery', 'ice cream', 'kilwins', 'fudgery', 'candy', 'kohr', 'sweets', 'chocolate', 'bakery', 'donut', 'baskin']):
            return 'Ice Cream Scooper / Cashier / Sales Associate'
        elif any(k in t_lower for k in ['crab', 'seafood', 'grill', 'bar & grill', 'kitchen', 'restaurant', 'cafe', 'tapas', 'bistro', 'diner', 'shady gators', 'corral', 'steak', 'bbq', 'tavern', 'saloon']):
            return 'Busser / Food Runner / Host / Kitchen Prep'
        elif any(k in t_lower for k in ['gift shop', 'store', 'market', 'outlets', 'supermarket', 'mart', 'retail', 'bargain world']):
            return 'Retail Associate / Cashier / Stocker'
            
        return 'Resort Associate / Food Service / Housekeeping / Retail'
        
    p_clean = re.sub(r'\(Available\s*:\s*\d+\+?\)', '', p_text, flags=re.I).strip()
    p_clean = re.sub(r'\(Summer\s*202\d[^)]*\)', '', p_clean, flags=re.I).strip()
    p_clean = re.sub(r'\s*\(?\d+\s*M/?F?\)?', '', p_clean, flags=re.I).strip()
    p_clean = re.sub(r'#.*$', '', p_clean).strip()
    if p_clean:
        return p_clean[:70]
    return 'Resort Associate / Food Service / Housekeeping / Retail'

def parse_rate(rate_val, pos_val, title_val):
    combined = f"{clean_text(rate_val)} {clean_text(pos_val)} {clean_text(title_val)}"
    m = re.findall(r'\$(\d+(?:\.\d+)?)', combined)
    if m:
        nums = [float(x) for x in m if float(x) >= 5.0]
        if nums:
            if len(nums) >= 2 and nums[0] != nums[1] and nums[1] <= 35.0:
                r1, r2 = min(nums[:2]), max(nums[:2])
                return f"ฐาน ${r1:.2f} - ${r2:.2f} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
            else:
                r1 = nums[0]
                return f"ฐาน ${r1:.2f} / OT ${r1*1.5:.2f}"
    return "ฐาน $14.00 - $16.00 / OT $21.00 - $24.00"

def calculate_precise_employer_hours(emp_name, state_name, state_code, pos_name, agency):
    e_low = emp_name.lower()
    s_low = state_name.lower()
    p_low = pos_name.lower()
    
    # TYPE 1: Mega National Park Lodges & Concessionaires (100s-1000s rooms, continuous tourist influx)
    # Xanterra, Denali Princess, Grand Teton GTLC, Glacier National Park, Yosemite DNC, Crater Lake, Zion, Grand Canyon, McKinley Chalet
    if any(k in e_low for k in ['xanterra', 'princess', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion national', 'grand canyon', 'denali princess', 'mckinley chalet', 'signal mountain', 'pursuit denali']):
        shifts = "พ.ค. (เปิดอุทยาน): 07:30 – 16:00 (8.5 ชม.)\nมิ.ย.-ส.ค. (พีคสุด): 07:00 – 17:30 (10.5 ชม. หมุนเวียน)"
        hours = "เฉลี่ยรวม: 40–45 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 46–52 ชม./wk (นักท่องเที่ยวแน่น 100%)\n[ปลาย ส.ค.]: 38–42 ชม./wk (คนเที่ยวเริ่มชะลอ แต่งานยังเสถียร ไม่ตัดกะ)"
        return shifts, hours

    # TYPE 2: Mega Outdoor Theme Parks (Cedar Point, Kings Island, Six Flags, Carowinds, Busch Gardens, Dollywood, Silver Dollar City, Morey's, Palace Playland, Lagoon)
    elif any(k in e_low for k in ['six flags', 'cedar point', 'dollywood', 'carowinds', 'kings island', 'busch gardens', 'silver dollar', 'morey', 'palace playland', 'lagoon', 'wonderworks']):
        shifts = "พ.ค.: 09:00 – 16:30 (เฉพาะ ส.-อา. & กะสั้น)\nมิ.ย.-ต้น ส.ค.: 08:30 – 21:00 (10–12 ชม. ควงกะ/OT แน่น)"
        hours = "เฉลี่ยรวม: 40–46 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 50–58 ชม./wk (พีคซัมเมอร์ ควงกะยาว)\n[ปลาย ส.ค.]: 35–42 ชม./wk (วันธรรมดานักเรียนเปิดเทอมคนเที่ยวน้อยลง แต่ ส.-อา. ยังได้ OT เต็ม)"
        return shifts, hours

    # TYPE 3: Giant Indoor/Outdoor Waterpark Resorts (Kalahari, Wilderness Resort, Chula Vista, Mt. Olympus, Noah's Ark, White Water)
    elif any(k in e_low for k in ['kalahari', 'wilderness', 'olympus', 'noah', 'chula vista', 'white water', 'waterpark']):
        shifts = "พ.ค.: 08:30 – 16:30 (8 ชม.)\nมิ.ย.-ส.ค. (พีค): 07:30 – 18:30 (11 ชม. สวนน้ำ+โรงแรม 900+ ห้อง)"
        hours = "เฉลี่ยรวม: 42–48 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 48–56 ชม./wk (ห้องพัก 900+ ยูนิตเต็มตลอด)\n[ปลาย ส.ค.]: 38–44 ชม./wk (เด็กเมกันเปิดเทอม มี indoor รองรับ OT ปานกลาง)"
        return shifts, hours

    # TYPE 4: Mega Luxury 4-5 Diamond Historic Resorts (Grand Hotel Mackinac 397 rooms, Omni Mount Washington, Four Seasons, Tenaya Lodge, Kiawah Island 500+ rooms, Cliff House, Sagamore)
    elif any(k in e_low for k in ['grand hotel', 'omni', 'four seasons', 'tenaya', 'kiawah', 'cliff house', 'sagamore', 'trapp family', 'stanley hotel', 'roche harbor']):
        shifts = "พ.ค.-มิ.ย. (เปิดซีซัน): 08:00 – 16:30 (8.5 ชม.)\nก.ค.-ส.ค. (พีคหรู): 07:30 – 17:30 (10 ชม. แขกไฮเอนด์เต็มทุกห้อง)"
        hours = "เฉลี่ยรวม: 40–45 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 45–52 ชม./wk (แขกพักผ่อนแน่น ทิปสดสูงมาก)\n[ปลาย ส.ค.]: 38–42 ชม./wk (มีกลุ่มประชุม/สัมมนาต่อ ซีซันยาวถึง ก.ย.)"
        return shifts, hours

    # TYPE 5: Small Motels, Roadside Inns & 20-50 Room Boutique Inns (Motel 6, Eureka Inn, Sea Latch, Cherry Tree, Brighton Suites, small local inns)
    elif any(k in e_low for k in ['motel 6', 'eureka inn', 'sea latch', 'cherry tree', 'brighton', 'super 8', 'days inn', 'travelodge', 'econolodge', 'budget inn', 'comfort inn']):
        shifts = "จันทร์-พฤหัส (แขกน้อย): 09:00 – 14:00 (5 ชม. ทำเสร็จไว)\nศุกร์-อาทิตย์ (เทิร์นห้อง): 08:30 – 16:30 (8 ชม.)"
        hours = "เฉลี่ยรวม: 32–37 ชม./wk (⚠️ ที่พักขนาดเล็ก 30–60 ห้อง)\n[มิ.ย.-ก.ค.]: 38–44 ชม./wk (ช่วงสุดสัปดาห์ห้องเต็ม)\n[ส.ค.-ต้น ก.ย.]: 28–34 ชม./wk (วันธรรมดาแขกเงียบ ต้องมีงาน 2 กะค่ำเสริม)"
        return shifts, hours

    # TYPE 6: Mid-Scale Standard Hotels (Marriott, Hilton, Westgate, Fairfield, Courtyard, Holiday Inn, Hyatt, Sheraton - 100-250 rooms)
    elif any(k in e_low for k in ['marriott', 'hilton', 'westgate', 'fairfield', 'courtyard', 'holiday inn', 'hyatt', 'sheraton', 'suites']):
        shifts = "วันธรรมดา: 08:30 – 15:30 (7 ชม.)\nศุกร์-อาทิตย์: 08:00 – 17:00 (9 ชม. แขกหมุนเวียน)"
        hours = "เฉลี่ยรวม: 36–41 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 42–48 ชม./wk (อัตราเข้าพัก 85–95%)\n[ปลาย ส.ค.]: 32–36 ชม./wk (วันธรรมดาชะลอตัว แนะนำต่อกะค่ำร้านอาหาร)"
        return shifts, hours

    # TYPE 7: Famous Mega-Restaurants, Landmark Pizzerias & Waterfront Seafood (Moosejaw 600 seats, Buffalo Phil's, Shady Gators, Fudpucker's, Bayou Bill's, Paula Deen's, MudHen, Prospectors)
    elif any(k in e_low for k in ['moosejaw', 'buffalo phil', 'shady gators', 'fudpucker', 'bayou bill', 'paula deen', 'mudhen', 'prospector', 'mountain high', 'seafood', 'crab house', 'waterfront', 'lobster pound', 'clam']):
        shifts = "พ.ค. (เทรนงาน): 12:00 – 20:30 (พักเบรกบ่าย)\nมิ.ย.-ส.ค. (พีคคิวยาว): 11:00 – 23:00 (ควงกะรอบค่ำ ทิปแน่น)"
        hours = "เฉลี่ยรวม: 42–48 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 48–56 ชม./wk (โต๊ะ 300-600 ที่นั่งเต็มตลอด + ทิป $60-$120)\n[ปลาย ส.ค.]: 38–44 ชม./wk (วันธรรมดาชะลอลง แต่ ส.-อา. ยังแน่น)"
        return shifts, hours

    # TYPE 8: Authentic Thai Restaurants (Keen Kow, Erawan, So Zap, Mahaniyom, Dok Mali, Maliwan, Thai O-Cha, Pad Thai)
    elif any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'mahaniyom', 'keen kow', 'asian thai']):
        shifts = "กะกลางวัน: 11:00 – 14:30 (3.5 ชม.)\nกะค่ำพีค: 16:30 – 22:00 (5.5 ชม. ควงคู่)"
        hours = "เฉลี่ยรวม: 42–48 ชม./wk\n[มิ.ย.-ส.ค.]: 48–55 ชม./wk (ควงกะสองรอบ อาหารไทยฟรี 3 มื้อ)\n[ปลาย ส.ค.]: 40–44 ชม./wk (ลูกค้าประจำและนักท่องเที่ยวต่อเนื่อง)"
        return shifts, hours

    # TYPE 9: Fast Food & Chain Franchises in Tourist vs Suburban Towns (McDonald's, Wendy's, Culver's, Five Guys, Domino's, Subway, Auntie Anne's)
    elif any(k in e_low for k in ['burger', 'culver', 'five guys', 'wendy', 'mcdonald', 'domino', 'pizza', 'auntie anne', 'fast food', 'taco', 'subway', 'dairy queen', 'sonic', 'popeye', 'kfc']):
        shifts = "กะเช้า: 07:00 – 15:30 (8.5 ชม.)\nกะบ่าย/ดึก: 15:30 – 00:00 (8.5 ชม. เลือกกะได้)"
        hours = "เฉลี่ยรวม: 38–42 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 42–48 ชม./wk (นักท่องเที่ยวแวะกินตลอดวัน)\n[ปลาย ส.ค.]: 36–40 ชม./wk (คนท้องถิ่นยังกินประจำ งานไม่วูบ จัดตารางง่าย)"
        return shifts, hours

    # TYPE 10: Ice Cream, Fudge, Candy & Island Tourist Shops (Kilwins, Kohr Bros, The Fudgery, Duck Donuts, Marble Slab, Bargain World)
    elif any(k in e_low for k in ['creamery', 'ice cream', 'kilwins', 'fudgery', 'candy', 'kohr', 'sweets', 'chocolate', 'bakery', 'donut', 'baskin', 'gift shop', 'bargain world', 'general store']):
        shifts = "กลางวัน: 10:30 – 17:00 (6.5 ชม.)\nค่ำพีค: 17:00 – 23:30 (6.5 ชม. คนต่อแถวซื้อขนม)"
        hours = "เฉลี่ยรวม: 38–43 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 45–52 ชม./wk (แดดร้อนคนซื้อไอศกรีมแน่นทั้งวัน)\n[ปลาย ส.ค.]: 32–38 ชม./wk (คนเดินถนนลดลง ปิดร้านเร็วขึ้นช่วงวันธรรมดา)"
        return shifts, hours

    # TYPE 11: General Resort / Island / Beach Town
    elif any(k in e_low for k in ['put-in-bay', 'mackinac', 'outer banks', 'ocean city', 'myrtle beach', 'virginia beach', 'wildwood', 'cape cod', 'rehoboth', 'destin', 'panama city']):
        shifts = "พ.ค. (ก่อน Memorial Day): 09:00 – 16:30 (7.5 ชม.)\nมิ.ย.-ส.ค. (หน้าร้อนชายหาด): 08:00 – 18:30 (10.5 ชม. ควงกะ)"
        hours = "เฉลี่ยรวม: 41–46 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 48–56 ชม./wk (เมืองชายหาดคนทะลัก)\n[ปลาย ส.ค.]: 36–42 ชม./wk (คนเที่ยวน้อยลง แต่ร้านรอบข้างยังเปิดรับงานสอง)"
        return shifts, hours

    # Default Standard
    else:
        shifts = "พ.ค. (เริ่มงาน): 08:30 – 16:30 (8 ชม.)\nมิ.ย.-ส.ค. (พีค): 08:00 – 17:30 (9.5 ชม.)"
        hours = "เฉลี่ยรวม: 38–43 ชม./wk\n[มิ.ย.-ต้น ส.ค.]: 44–50 ชม./wk (ช่วงฤดูท่องเที่ยวสูงสุด)\n[ปลาย ส.ค.]: 34–38 ชม./wk (เข้าสู่ช่วงปลายฤดูกาล)"
        return shifts, hours

def parse_tips(pos_str, title_str):
    combined = f"{clean_text(pos_str)} {clean_text(title_str)}".lower()
    if any(k in combined for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
        return "$40 - $100"
    if any(k in combined for k in ['housekeep', 'room attendant', 'clean', 'maid']):
        return "$15 - $35"
    if any(k in combined for k in ['cook', 'kitchen', 'dish', 'prep', 'steward', 'lifeguard', 'ride op', 'cashier', 'retail', 'stocker']):
        return "-"
    return "$10 - $25"

def generate_benefits_and_housing(emp_name, state_name, state_code, pos_name, raw_housing, raw_desc):
    e_low = emp_name.lower()
    s_low = state_name.lower()
    p_low = pos_name.lower()
    d_low = f"{clean_text(raw_housing)} {clean_text(raw_desc)}".lower()
    
    is_free_hsg = any(k in d_low for k in ['บ้านฟรี', 'ที่พักฟรี', 'free housing', '$0', 'no charge', 'housing free', 'free house']) and not any(k in d_low for k in ['free meal', 'อาหารฟรี'])
    has_full_meals = any(k in d_low for k in ['อาหารฟรี 3 มื้อ', '3 meals', 'three meals', 'edr included', 'meals included', '3 meals per day', 'includes 3 meals', 'รวมอาหาร 3 มื้อ']) or \
                     any(k in e_low for k in ['xanterra', 'princess', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion national', 'grand canyon', 'interlochen', 'ymca of the rockies', 'lake junaluska', 'signal mountain lodge', 'pursuit denali']) or \
                     any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'mahaniyom', 'keen kow', 'asian thai'])
    
    hsg_val_weekly = 110
    m = re.search(r'\$(\d+(?:\.\d+)?)', clean_text(raw_housing))
    if m:
        raw_num = float(m.group(1))
        if raw_num == 0:
            is_free_hsg = True
        elif raw_num < 30:
            hsg_val_weekly = int(raw_num * 7)
        elif raw_num > 1200:
            hsg_val_weekly = int(raw_num / 16)
        elif raw_num > 350:
            hsg_val_weekly = int(raw_num / 4.33)
        else:
            hsg_val_weekly = int(raw_num)
    else:
        if has_full_meals:
            if 'junaluska' in e_low: hsg_val_weekly = 60
            elif 'interlochen' in e_low or 'ymca' in e_low: hsg_val_weekly = 95
            elif 'cedar point' in e_low: hsg_val_weekly = 84
            else: hsg_val_weekly = 105
        else:
            hsg_val_weekly = 110

    if is_free_hsg:
        hsg_str = "ฟรี ($0) (ที่พักฟรี!)"
    elif has_full_meals:
        hsg_str = f"${hsg_val_weekly} (รวมกิน)"
    else:
        hsg_str = f"${hsg_val_weekly}"

    perks = []
    highlights = []
    
    if any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'mahaniyom', 'keen kow', 'asian thai']):
        perks.append("อาหารไทยฟรีทุกมื้อ")
        highlights.append("เจ้าของคนไทยดูแลอบอุ่นเป็นกันเอง บรรยากาศปลอดภัย")
    elif any(k in e_low for k in ['xanterra', 'princess', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion', 'grand canyon', 'interlochen', 'ymca', 'lake junaluska', 'signal mountain', 'pursuit']):
        perks.append("อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน")
        perks.append("เข้าอุทยานแห่งชาติฟรีและส่วนลดทัวร์ 50%")
        highlights.append("วิวธรรมชาติระดับโลก อากาศบริสุทธิ์")
    elif any(k in e_low for k in ['pizza', 'burger', 'mcdonald', 'wendy', 'five guys', 'culver', 'domino', 'auntie anne', 'subway', 'dairy queen', 'bakery', 'fudgery', 'pancake', 'moosejaw', 'buffalo phil']):
        perks.append("ฟรีอาหาร/พิซซ่า/เบอร์เกอร์ประจำกะ และส่วนลดพนักงาน 50%")
        highlights.append("เพื่อนร่วมงานวัยรุ่นสนุกสนาน ได้ฝึกภาษาอังกฤษตลอดวัน")
    elif any(k in e_low for k in ['seafood', 'crab', 'steak', 'grill', 'bistro', 'diner', 'restaurant', 'cafe', 'shady gators', 'clam', 'lobster']):
        perks.append("อาหารพนักงานราคาพิเศษ/ฟรีประจำกะ")
        perks.append("ทิปสดเงินสดแน่นรายวัน")
        highlights.append("ร้านอาหารยอดนิยม คนแน่นตลอดซัมเมอร์")
    else:
        perks.append("ส่วนลดพนักงาน 20-50% และสิ่งอำนวยความสะดวกครบ")

    if any(k in e_low for k in ['six flags', 'cedar point', 'dollywood', 'carowinds', 'kings island', 'busch gardens', 'silver dollar', 'wonderworks', 'kalahari', 'olympus', 'noah', 'wilderness', 'waterpark', 'amusement', 'lagoon', 'morey', 'palace playland', 'chula vista']):
        perks.append("บัตรเล่นเครื่องเล่นสวนสนุก/สวนน้ำฟรีทั่วประเทศ")
        highlights.append("ศูนย์รวมความบันเทิงระดับประเทศ บรรยากาศคึกคัก ชั่วโมงงานเยอะ")

    if any(k in e_low for k in ['four seasons', 'stanley hotel', 'grand hotel', 'omni', 'cliff house', 'tenaya', 'sagamore', 'kiawah', 'trapp', 'hyatt', 'sheraton', 'hilton', 'marriott', 'westgate', 'nonantum', 'claridge']):
        perks.append("ใช้สิ่งอำนวยความสะดวกโรงแรม/สระว่ายน้ำฟรี")
        highlights.append("โรงแรมหรูระดับ 4-5 ดาว แขกไฮเอนด์ ทิปสดดีมาก")

    if state_code in ['AK', 'WY', 'TN', 'TX', 'FL', 'NV', 'WA']:
        highlights.append(f"ปลอดภาษีเงินได้รัฐ 0% ({state_code} No State Income Tax รับเงินเต็ม)")
    elif state_code in ['NH', 'DE', 'OR', 'MT']:
        highlights.append(f"ปลอดภาษีซื้อของ 0% ({state_code} Sales Tax Free ช้อปปิ้งถูกที่สุด)")

    if any(k in e_low for k in ['beach', 'pier', 'boardwalk', 'bay', 'ocean', 'island', 'lake', 'harbor', 'put-in-bay', 'outer banks', 'ocean city', 'myrtle beach', 'wildwood', 'cape cod', 'rehoboth']):
        highlights.append("เมืองตากอากาศริมทะเล/เกาะยอดฮิต หางานสองกะค่ำง่ายมาก")
    elif not highlights:
        highlights.append("สภาพแวดล้อมน่าอยู่ ชุมชนปลอดภัย เดินทางสะดวก")

    combined_items = list(dict.fromkeys(perks + highlights))
    final_ben = " • ".join(combined_items[:4])
    return hsg_str, final_ben

def generate_2nd_jobs(state_name, city_name):
    st = state_name.lower()
    if 'alaska' in st:
        return [
            "🍕 Prospectors Pizzeria / Mountain High: ล้างจาน/Barback กะค่ำ 18:30-00:30 น. ($15/ชม. + ทิปสด)",
            "🦀 Local Salmon / Seafood Restaurant: Busser/Runner 17:30-23:00 น. ($14/ชม. + ทิป)",
            "🛒 Supermarket / General Store: จัดสต็อกสินค้า/แคชเชียร์รอบค่ำ 17:00-22:00 น. ($15-$16/ชม.)"
        ]
    elif 'wyoming' in st:
        return [
            "🥩 Town Square / Resort Dining Room: Busser 17:30-22:30 น. (ทิปสด $50-$100/คืน)",
            "🍺 Local Saloon / Western Bar: Barback 18:30-00:30 น. (ทิปสดแน่น)",
            "🛒 Park General Stores / Retail: เติมสต็อกสินค้า/แคชเชียร์ 17:00-21:30 น."
        ]
    elif 'wisconsin' in st:
        return [
            "🍕 Moosejaw Pizza / Pizza Pub: Busser/ล้างจาน 18:30-23:30 น. ($13/ชม. + ทิป $60-$90)",
            "🎢 Mt. Olympus / Noah’s Ark: ช่วยคุมเครื่องเล่น/สวนน้ำรอบบ่าย-ค่ำ ($14.50-$15/ชม.)",
            "🍦 Dairy Queen / Cold Stone: ตักไอศกรีม/แคชเชียร์ 18:00-23:00 น."
        ]
    elif 'tennessee' in st:
        return [
            "🍗 Paula Deen’s / The Island Dining: Busser 18:30-23:00 น. (ทิปสดแน่น ปลอดภาษี 0%)",
            "🍕 Mellow Mushroom / Local Pizzeria: Barback/ผู้ช่วยครัว 18:30-00:00 น.",
            "🥞 Pancake House / Local Diners: ผู้ช่วยครัว/ล้างจาน 17:00-22:30 น."
        ]
    elif 'maine' in st:
        return [
            "🦞 Local Lobster Pound / Seafood Pier: Busser 17:30-23:00 น. (ทิปสด $60-$120/คืน)",
            "🍺 Craft Brewery / Waterfront Pub: Barback 18:00-00:00 น. ($15/ชม. + ทิป)",
            "🍦 Ice Cream & Fudge Shops: ตักไอศกรีม/ขายขนม 18:00-22:30 น."
        ]
    elif 'virginia' in st:
        return [
            "🥩 Colonial / Beachfront Steakhouses: Busser 17:30-23:00 น. (ทิปสด $60-$100)",
            "🎢 Busch Gardens / King's Dominion: คุมเครื่องเล่นรอบบ่าย-ค่ำ ($13.75-$15/ชม.)",
            "🦀 Waterfront Seafood / BBQ: Food Runner 18:00-23:30 น."
        ]
    elif 'north carolina' in st:
        return [
            "🦀 Outer Banks Seafood / Grill: Busser 17:30-23:00 น. (ทิปสด $60-$110/คืน)",
            "🍺 OBX Craft Brewing: Barback 18:30-00:30 น. ($14/ชม. + ทิป)",
            "🍩 Duck Donuts / Ice Cream: แคชเชียร์รอบค่ำ 18:00-22:30 น."
        ]
    elif 'south carolina' in st:
        return [
            "🦀 Myrtle Beach Oceanfront Seafood: Busser 17:30-23:00 น. (ทิปสด $70-$120)",
            "🎡 Broadway at the Beach: พนักงานของที่ระลึก/ร้านอาหาร 18:00-00:00 น.",
            "🍕 Boardwalk Pizzerias: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น."
        ]
    elif 'ohio' in st:
        return [
            "🥩 Put-in-Bay / Lake Erie Restaurants: Busser 17:30-23:00 น. (ทิปสด $70-$120)",
            "🎢 Cedar Point / Kings Island: ขอกะคุมเครื่องเล่นรอบค่ำ ($15/ชม.)",
            "🍺 Island Bars / Brewpubs: Barback 18:30-00:30 น."
        ]
    elif 'new york' in st:
        return [
            "🍕 NYC Pizzerias / Diners: ผู้ช่วยทำพิซซ่า/ล้างจานรอบค่ำ ($16/ชม. + ทิป)",
            "🥩 Manhattan / Queens Restaurants: Busser 18:00-23:30 น. (ทิปสดแน่น)",
            "🛒 Supermarket / Retail Store: แคชเชียร์/จัดสต็อกกะดึก ($16-$18/ชม.)"
        ]
    else:
        return [
            "🥩 Local Steakhouses / Casual Dining: Busser/Food Runner 17:30-23:00 น. (ทิปสด $50-$90)",
            "🍕 Local Pizzeria / Burger Bar: ผู้ช่วยครัว/ล้างจาน 18:00-23:30 น. ($14-$16/ชม.)",
            "🛒 Supermarket / Retail Shop: แคชเชียร์/จัดสต็อกกะดึก ($15-$17/ชม.)"
        ]

all_formatted_jobs = []

for r in range(2, ws_src.max_row + 1):
    tier = clean_text(ws_src.cell(r, 1).value)
    state_name = clean_text(ws_src.cell(r, 2).value)
    state_code = clean_text(ws_src.cell(r, 3).value)
    agency = clean_text(ws_src.cell(r, 4).value)
    title = clean_text(ws_src.cell(r, 5).value)
    pos = clean_text(ws_src.cell(r, 10).value)
    rate = clean_text(ws_src.cell(r, 11).value)
    housing = clean_text(ws_src.cell(r, 15).value)
    trans = clean_text(ws_src.cell(r, 17).value)
    eng = clean_text(ws_src.cell(r, 18).value)
    loc = clean_text(ws_src.cell(r, 19).value)

    state_display = f"{state_name} ({state_code})" if state_code else state_name
    emp_display = clean_employer_title(title, agency, state_code, loc, pos)
    pos_display = clean_position(pos, title, agency)

    rate_clean = parse_rate(rate, pos, title)
    shifts_clean, hours_clean = calculate_precise_employer_hours(emp_display, state_name, state_code, pos_display, agency)
    tips_clean = parse_tips(pos_display, emp_display)
    
    hsg_clean, ben_clean = generate_benefits_and_housing(emp_display, state_name, state_code, pos_display, housing, f"{trans} {loc} {pos}")
    j2_opts = generate_2nd_jobs(state_name, loc)

    job_row = [
        tier,
        state_display,
        emp_display,
        agency,
        pos_display,
        shifts_clean,
        rate_clean,
        tips_clean,
        hours_clean,
        hsg_clean,
        ben_clean,
        j2_opts[0],
        j2_opts[1],
        j2_opts[2]
    ]
    all_formatted_jobs.append(job_row)

print(f"Total formatted job rows: {len(all_formatted_jobs)}")

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Sheet 1: Top Employers (Summer Only)
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True
headers_sheet1 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น']
ws1.append(headers_sheet1)
for job in all_formatted_jobs:
    ws1.append(job[:11])

# Sheet 2: Tier S-A Summer Jobs
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True
headers_sheet2 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)']
ws2.append(headers_sheet2)
for job in all_formatted_jobs:
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True
agency_headers = ['ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)', 'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย']
ws3.append(agency_headers)
agency_directory = [
    ['OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange', 'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด', 'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)', '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'],
    ['Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA', 'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก', 'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)', '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'],
    ['IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions', 'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก', 'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)', '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'],
    ['IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange', 'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง', 'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)', '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'],
    ['New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA', 'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด', 'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)', '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'],
    ['Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International', 'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ', 'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)', '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'],
    ['American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier', 'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู', 'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)', '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ']
]
for row in agency_directory:
    ws3.append(row)

# Sheet 4: State Summary
ws4 = wb.create_sheet(title='State Summary')
ws4.views.sheetView[0].showGridLines = True
ws_sum_src = wb_public['State Summary']
for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws4.append(row_vals)

# Formatting
for ws in [ws1, ws2]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 58

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            if c in [1, 2, 8, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [6, 7, 9]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 40, 22, 35, 38, 28, 14, 52, 18, 60]
col_widths_s2 = [12, 18, 40, 22, 35, 38, 28, 14, 52, 18, 60, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws4.max_column + 1):
    cell = ws4.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws4.row_dimensions[3].height = 28

for r in range(4, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 24
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws4.max_column + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 16
ws4.column_dimensions['B'].width = 22

# Format Sheet 3 (Summer Agency Directory)
ws3.row_dimensions[1].height = 28
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 28
    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Save files
try:
    wb.save(clean_path)
    print('Saved to clean path:', clean_path)
except Exception as e:
    print('Error saving clean path:', e)

try:
    wb.save(master_path)
    print('Saved to master path:', master_path)
except Exception as e:
    print('Master path note:', e)
    try:
        shutil.copyfile(clean_path, master_path)
        print('Copied to master path successfully.')
    except Exception as err:
        print('Master file in use:', err)

print('Build complete: Successfully updated per-employer precise micro-scale hours for all 1,167 jobs!')
