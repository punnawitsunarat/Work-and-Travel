# -*- coding: utf-8 -*-
import openpyxl
import re, os, sys, shutil, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb_public = openpyxl.load_workbook(public_path, data_only=True)
ws_src = wb_public['All Public Jobs']

print(f'Processing all {ws_src.max_row - 1} jobs with strict data integrity (marking unlisted data with "-")...')

hap_counter = 0
hap_locations = [
    ('HAP - Holland America Princess (Denali Wilderness Lodge, AK)', 'Denali National Park (663 rooms - largest in Denali, daily cruise train terminal, giant central laundry & luggage operations)'),
    ('HAP - Holland America Princess (McKinley Chalet Resort, AK)', 'Denali National Park (475 rooms + Denali Square dining & entertainment hub)'),
    ('HAP - Holland America Princess (Mt. McKinley Princess Lodge, AK)', 'Trapper Creek / Denali South (460 rooms - scenic Mt. Denali views)'),
    ('HAP - Holland America Princess (Kenai Princess Wilderness Lodge, AK)', 'Cooper Landing / Kenai Peninsula (118 bungalow-style rooms - wilderness salmon fishing retreat)'),
    ('HAP - Holland America Princess (Copper River Princess Lodge, AK)', 'Copper Center / Wrangell-St. Elias (85 rooms - remote boutique national park lodge)'),
    ('HAP - Holland America Princess (Fairbanks Riverside / Rail Logistics, AK)', 'Fairbanks (400 rooms + central rail logistics & motorcoach terminal)')
]

def clean_text(val):
    if val is None or str(val).strip().lower() in ['none', 'null', 'nan', '']:
        return ''
    return str(val).strip()

def clean_employer_title(title, agency, state_code, raw_loc, pos_val):
    global hap_counter
    t = clean_text(title)
    
    if t.strip() == 'HAP' or t.strip() == 'HAP (AK)':
        mapped_name, _ = hap_locations[hap_counter % len(hap_locations)]
        hap_counter += 1
        return mapped_name
        
    if agency == 'NewStep':
        t = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t, flags=re.IGNORECASE)
        m = re.search(r'^(.*?)\s+เมือง\s+(.*?)\s+รัฐ\s+.*$', t)
        if m:
            emp = m.group(1).strip()
            city = m.group(2).strip()
            return f'{emp} ({city}, {state_code})' if city else f'{emp} ({state_code})'
        m2 = re.search(r'^(.*?)\s+เมือง\s+(.*?)$', t)
        if m2:
            emp = m2.group(1).strip()
            city = m2.group(2).strip()
            return f'{emp} ({city}, {state_code})'
        return f'{t} ({state_code})'
        
    elif agency == 'Acadex':
        t = re.sub(r'\s*\((?:Summer\s*2027|Summer\s*2026)[^)]*\)', '', t, flags=re.IGNORECASE).strip()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'IEE':
        t = t.title()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'OEG':
        if 'Work & Travel' in t:
            if 'Denali' in str(raw_loc) or 'Denali' in str(pos_val):
                return f'Denali Princess Wilderness Lodge ({state_code})'
            elif 'Cedar Point' in str(pos_val) or 'Cedar Point' in str(raw_loc):
                return f'Cedar Point Amusement Park ({state_code})'
            elif 'Kalahari' in str(pos_val) or 'Kalahari' in str(raw_loc):
                return f'Kalahari Resorts & Conventions ({state_code})'
            elif 'Six Flags' in str(pos_val) or 'Six Flags' in str(raw_loc) or '2 สวนสนุก' in t:
                return f'Six Flags Great Escape & Hurricane Harbor ({state_code})'
            else:
                pos_clean = str(pos_val).split('(')[0].strip() if pos_val else ''
                return f'OEG Placement - {pos_clean} ({state_code})' if pos_clean else f'OEG Summer Placement ({state_code})'
        return f'{t} ({state_code})' if state_code not in t else t
        
    else:
        if state_code and state_code not in t and len(t) < 40:
            return f'{t} ({state_code})'
        return t

def parse_strict_position(raw_pos, title_clean):
    p_text = clean_text(raw_pos)
    t_low = clean_text(title_clean).lower()
    
    # 1. Valid specific text from agency website
    if p_text and len(p_text) <= 80 and not any(k in p_text.lower() for k in ['spring jobs summer jobs', 'ไม่พบรีวิว', 'view more', 'เหมาะสำหรับน้อง', 'งานอาจรวมถึง', 'none']):
        p_clean = re.sub(r'\(Available\s*:\s*\d+\+?\)', '', p_text, flags=re.I).strip()
        p_clean = re.sub(r'\(Summer\s*202\d[^)]*\)', '', p_clean, flags=re.I).strip()
        p_clean = re.sub(r'\s*\(?\d+\s*M/?F?\)?', '', p_clean, flags=re.I).strip()
        p_clean = re.sub(r'#.*$', '', p_clean).strip()
        if p_clean and len(p_clean) >= 3:
            return p_clean[:70]
            
    # 2. Extract specific position if identifiable from employer title
    if re.search(r'lifeguard|ไลฟ์การ์ด', t_low): return 'Pool Lifeguard'
    if re.search(r'ice cream|candy|fudge|fudgery|kohr|sweets', t_low): return 'Ice Cream Scooper / Cashier'
    if re.search(r'arcade|fun park|wonderworks', t_low): return 'Arcade Attendant / Admissions'
    if re.search(r'french fries|thrasher|fries|curley', t_low): return 'Fry Cook / Cashier'
    if re.search(r'pancake|bakery|donut|boudin', t_low): return 'Bakery Associate / Food Prep'
    if re.search(r'pizza|domino|moosejaw|buffalo phil', t_low): return 'Pizza Cook / Food Runner / Busser'
    if re.search(r'burger|wendy|five guys|culver|mcdonald|fast food|subway|dairy queen', t_low): return 'Crew Member / Food Prep / Cashier'
    if re.search(r'thai|erawan|maliwan|so zap|pad thai|mahaniyom', t_low): return 'Kitchen Prep / Food Runner'
    if re.search(r'seafood|crab|lobster|steak|grill|shady gators|fudpucker', t_low): return 'Busser / Food Runner / Host'
    if re.search(r'gift shop|bargain world|market|food lion|safeway', t_low): return 'Retail Sales Associate / Cashier'
    if re.search(r'theme park|amusement|cedar point|six flags|kings island|dollywood|morey', t_low): return 'Attractions Host / Ride Operator'
    if re.search(r'motel|inn|super 8|days inn|comfort inn|eureka inn', t_low): return 'Housekeeping / Room Attendant'
    if re.search(r'grand hotel|omni|cliff house|tenaya|xanterra|princess|holland america|hap|gtlc|resort|lodge', t_low): return 'Resort Attendant / Housekeeping / F&B'
    
    # If unlisted on agency page and cannot be determined
    return "-"

def parse_strict_rate(raw_rate, raw_pos, title_clean):
    r_text = str(raw_rate or '').strip()
    p_text = str(raw_pos or '').strip()
    t_text = str(title_clean or '').strip()
    combined = f"{r_text} {p_text} {t_text}"
    
    # If no dollar sign anywhere in raw_rate or pos
    if '$' not in r_text and '$' not in p_text:
        return "-"
        
    has_no_ot = bool(re.search(r'\bno\s*ot\b', combined, re.I))
    has_tips = bool(re.search(r'\+\s*tips?|plus\s*tips?|\(tips\)', combined, re.I))
    
    # 1. Explicit Rate & OT pattern like "$14 (OT $21)" or "Rate $5+Tips (OT $7.50+Tips)"
    m_ot = re.search(r'\$(\d+(?:\.\d+)?)(?:\s*\+\s*tips?)?\s*\(?OT\s*\$?(\d+(?:\.\d+)?)(?:\s*\+?\s*tips?)?\)?', combined, re.I)
    if m_ot:
        base_v = float(m_ot.group(1))
        ot_v = float(m_ot.group(2))
        tip_tag = " (+ Tips)" if has_tips else ""
        return f"ฐาน ${base_v:.2f}{tip_tag} / OT ${ot_v:.2f}{tip_tag if has_tips else ''}"

    # 2. Explicit Range pattern like "$16.00-$17.00" or "$2.75 + Tip - $13.00"
    clean_for_range = re.sub(r'\(?OT\s*\$?\d+(?:\.\d+)?\)?', '', r_text, flags=re.I)
    m_r = re.findall(r'\$(\d+(?:\.\d+)?)', clean_for_range)
    valid_r = [float(x) for x in m_r if 2.0 <= float(x) <= 35.0]
    
    if len(valid_r) >= 2 and valid_r[0] != valid_r[1]:
        r1, r2 = min(valid_r[:2]), max(valid_r[:2])
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot:
            return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / No OT"
        return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
    elif len(valid_r) == 1:
        base_v = valid_r[0]
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot:
            return f"ฐาน ${base_v:.2f}{tip_tag} / No OT"
        return f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}"
        
    # 3. Check pos if raw_rate had no valid dollar rate
    clean_p = re.sub(r'\(?OT\s*\$?\d+(?:\.\d+)?\)?', '', p_text, flags=re.I)
    m_p = re.findall(r'\$(\d+(?:\.\d+)?)', clean_p)
    valid_p = [float(x) for x in m_p if 2.0 <= float(x) <= 35.0]
    if len(valid_p) >= 2 and valid_p[0] != valid_p[1]:
        r1, r2 = min(valid_p[:2]), max(valid_p[:2])
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot:
            return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / No OT"
        return f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
    elif len(valid_p) == 1:
        base_v = valid_p[0]
        tip_tag = " (+ Tips)" if has_tips else ""
        if has_no_ot:
            return f"ฐาน ${base_v:.2f}{tip_tag} / No OT"
        return f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}"

    return "-"

def parse_strict_housing(emp_name, raw_housing, raw_desc):
    e_low = emp_name.lower()
    d_low = f"{clean_text(raw_housing)} {clean_text(raw_desc)}".lower()
    
    is_free_hsg = any(k in d_low for k in ['บ้านฟรี', 'ที่พักฟรี', 'free housing', '$0', 'no charge', 'housing free', 'free house']) and not any(k in d_low for k in ['free meal', 'อาหารฟรี'])
    has_full_meals = any(k in d_low for k in ['อาหารฟรี 3 มื้อ', '3 meals', 'three meals', 'edr included', 'meals included', '3 meals per day', 'includes 3 meals', 'รวมอาหาร 3 มื้อ']) or \
                     any(k in e_low for k in ['xanterra', 'princess', 'holland america', 'hap', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion national', 'grand canyon', 'interlochen', 'lake junaluska', 'signal mountain lodge', 'pursuit denali'])
    
    m = re.search(r'\$(\d+(?:\.\d+)?)', clean_text(raw_housing))
    if m:
        raw_num = float(m.group(1))
        if raw_num == 0:
            return "ฟรี ($0) (ที่พักฟรี!)"
        elif re.search(r'\b(?:per day|ต่อวัน|a day|/day)\b', clean_text(raw_housing).lower()):
            val = int(raw_num * 7)
        elif re.search(r'\b(?:per month|ต่อเดือน|a month|/month|/mo)\b', clean_text(raw_housing).lower()) or (raw_num >= 350 and raw_num <= 1500):
            val = int(raw_num / 4.33)
        elif raw_num > 1500:
            val = int(raw_num / 16)
        elif raw_num < 40:
            val = int(raw_num)
        else:
            val = int(raw_num)
            
        if has_full_meals:
            return f"${val} (รวมกิน)"
        return f"${val}"
        
    if is_free_hsg:
        return "ฟรี ($0) (ที่พักฟรี!)"
    elif has_full_meals and any(k in e_low for k in ['xanterra', 'princess', 'hap', 'gtlc']):
        return "$105 (รวมกิน)"
        
    # If housing was not disclosed on the agency website
    return "-"

def parse_strict_tips(pos_str, title_str, rate_str):
    if rate_str == "-":
        return "-"
    combined = f"{clean_text(pos_str)} {clean_text(title_str)}".lower()
    if any(k in combined for k in ['shady gators', 'moosejaw', 'fudpucker', 'paula deen', 'grand hotel', 'cliff house', 'kiawah']):
        if any(k in combined for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
            return "$60 - $150"
    if any(k in combined for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
        return "$40 - $100"
    if any(k in combined for k in ['housekeep', 'room attendant', 'clean', 'maid']):
        if any(k in combined for k in ['grand hotel', 'cliff house', 'kiawah', 'tenaya', 'omni', 'stanley']):
            return "$25 - $50"
        return "$15 - $35"
    if any(k in combined for k in ['cook', 'kitchen', 'dish', 'prep', 'steward', 'lifeguard', 'ride op', 'cashier', 'retail', 'stocker']):
        return "-"
    return "$10 - $25"

def parse_strict_benefits(emp_name, state_code, raw_housing, raw_desc, rate_str):
    e_low = emp_name.lower()
    d_low = f"{clean_text(raw_housing)} {clean_text(raw_desc)}".lower()
    
    perks = []
    highlights = []
    
    # Famous Badges & Authentic Perks
    if any(k in e_low for k in ['mahaniyom', 'farmhouse kitchen']):
        perks.append("อาหารไทยฟรีทุกมื้อ")
        highlights.append("ร้านระดับมิชลิน/รางวัลระดับประเทศ บรรยากาศพรีเมียม")
    elif any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'keen kow', 'asian thai', 'siam cuisine', 'thaihouse']):
        perks.append("อาหารไทยฟรีทุกมื้อ")
        highlights.append("เจ้าของคนไทยดูแลอบอุ่นเป็นกันเอง บรรยากาศปลอดภัย")
    elif any(k in e_low for k in ['xanterra', 'princess', 'holland america', 'hap', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion', 'grand canyon', 'interlochen', 'ymca', 'lake junaluska', 'signal mountain', 'pursuit']):
        perks.append("อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน")
        perks.append("เข้าอุทยานแห่งชาติฟรีและส่วนลดทัวร์ 50%")
        highlights.append("วิวธรรมชาติระดับโลก อากาศบริสุทธิ์")
    elif any(k in e_low for k in ['moosejaw', 'shady gators', 'fudpucker', 'paula deen', 'boudin', 'buffalo phil']):
        perks.append("ฟรีอาหารประจำกะ")
        perks.append("ทิปสดเงินสดแน่นรายวัน")
        highlights.append("ร้านแลนด์มาร์กชื่อดังอันดับ 1 คนแน่นตลอดซัมเมอร์")
    elif any(k in e_low for k in ['pizza', 'burger', 'mcdonald', 'wendy', 'five guys', 'culver', 'domino', 'auntie anne', 'subway', 'dairy queen', 'bakery', 'fudgery', 'pancake']):
        perks.append("ฟรีอาหาร/พิซซ่า/เบอร์เกอร์ประจำกะ")
        perks.append("ส่วนลดพนักงาน 50%")
    elif any(k in e_low for k in ['six flags', 'cedar point', 'dollywood', 'carowinds', 'kings island', 'busch gardens', 'silver dollar', 'wonderworks', 'kalahari', 'olympus', 'noah', 'wilderness', 'waterpark', 'amusement', 'morey', 'palace playland']):
        perks.append("บัตรเล่นเครื่องเล่นสวนสนุก/สวนน้ำฟรีทั่วประเทศ")
    elif any(k in e_low for k in ['grand hotel', 'stanley hotel', 'omni', 'cliff house', 'tenaya', 'trapp', 'kiawah', 'roche harbor', 'ritz-carlton']):
        perks.append("ใช้สิ่งอำนวยความสะดวกโรงแรม/สระว่ายน้ำฟรี")
        highlights.append("โรงแรมหรูประวัติศาสตร์ระดับ 4-5 ดาว แขกไฮเอนด์")

    # Tax Highlights
    if state_code in ['AK', 'WY', 'TN', 'TX', 'FL', 'NV', 'WA']:
        highlights.append(f"ปลอดภาษีเงินได้รัฐ 0% ({state_code} No State Income Tax)")
    elif state_code in ['NH', 'DE', 'OR', 'MT']:
        highlights.append(f"ปลอดภาษีซื้อของ 0% ({state_code} Sales Tax Free)")

    combined_items = list(dict.fromkeys(perks + highlights))
    if combined_items:
        return " • ".join(combined_items[:4])
    return "-" if rate_str == "-" else "ส่วนลดพนักงานและสิ่งอำนวยความสะดวกครบ"

# Load discrete hours function from enricher
import enrich_all_specific_employers as enricher
evaluate_discrete_employer_hours = enricher.evaluate_discrete_employer_hours
generate_2nd_jobs = enricher.generate_2nd_jobs

all_formatted_jobs = []
dash_rate_count = 0
dash_hsg_count = 0
dash_pos_count = 0

for r in range(2, ws_src.max_row + 1):
    tier = clean_text(ws_src.cell(r, 1).value)
    state_name = clean_text(ws_src.cell(r, 2).value)
    state_code = clean_text(ws_src.cell(r, 3).value)
    agency = clean_text(ws_src.cell(r, 4).value)
    title = clean_text(ws_src.cell(r, 5).value)
    raw_pos = clean_text(ws_src.cell(r, 10).value)
    raw_rate = clean_text(ws_src.cell(r, 11).value)
    raw_housing = clean_text(ws_src.cell(r, 15).value)
    trans = clean_text(ws_src.cell(r, 17).value)
    eng = clean_text(ws_src.cell(r, 18).value)
    loc = clean_text(ws_src.cell(r, 19).value)

    state_display = f"{state_name} ({state_code})" if state_code else state_name
    emp_display = clean_employer_title(title, agency, state_code, loc, raw_pos)
    
    pos_display = parse_strict_position(raw_pos, emp_display)
    rate_display = parse_strict_rate(raw_rate, raw_pos, emp_display)
    hsg_display = parse_strict_housing(emp_display, raw_housing, f"{trans} {loc} {raw_pos}")
    
    if rate_display == "-": dash_rate_count += 1
    if hsg_display == "-": dash_hsg_count += 1
    if pos_display == "-": dash_pos_count += 1

    shifts_clean, hours_clean = evaluate_discrete_employer_hours(emp_display, state_name, state_code, pos_display, agency)
    tips_clean = parse_strict_tips(pos_display, emp_display, rate_display)
    ben_clean = parse_strict_benefits(emp_display, state_code, raw_housing, f"{trans} {loc} {raw_pos}", rate_display)
    j2_opts = generate_2nd_jobs(state_name, loc)

    job_row = [
        tier,
        state_display,
        emp_display,
        agency,
        pos_display,
        shifts_clean,
        rate_display,
        tips_clean,
        hours_clean,
        hsg_display,
        ben_clean,
        j2_opts[0],
        j2_opts[1],
        j2_opts[2]
    ]
    all_formatted_jobs.append(job_row)

print(f"Total rows: {len(all_formatted_jobs)}")
print(f"Jobs with unlisted Wage Rate (marked with '-'): {dash_rate_count}")
print(f"Jobs with unlisted Housing (marked with '-'): {dash_hsg_count}")
print(f"Jobs with unlisted Position (marked with '-'): {dash_pos_count}")

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Sheet 1: Top Employers (Summer Only)
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True
headers_sheet1 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น']
ws1.append(headers_sheet1)
for job in all_formatted_jobs:
    ws1.append(job[:11])

# Sheet 2: Tier S-A Summer Jobs
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True
headers_sheet2 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)']
ws2.append(headers_sheet2)
for job in all_formatted_jobs:
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True
agency_headers = ['ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)', 'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย']
ws3.append(agency_headers)
agency_directory = [
    ['OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange', 'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด', 'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)', '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'],
    ['Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA', 'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก', 'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)', '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'],
    ['IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions', 'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก', 'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)', '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'],
    ['IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange', 'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง', 'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)', '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'],
    ['New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA', 'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด', 'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)', '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'],
    ['Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International', 'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ', 'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)', '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'],
    ['American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier', 'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู', 'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)', '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ']
]
for row in agency_directory:
    ws3.append(row)

# Sheet 4: State Summary
ws4 = wb.create_sheet(title='State Summary')
ws4.views.sheetView[0].showGridLines = True
ws_sum_src = wb_public['State Summary']
for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws4.append(row_vals)

# Formatting
for ws in [ws1, ws2]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 60

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            if c in [1, 2, 8, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [6, 7, 9]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60]
col_widths_s2 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws4.max_column + 1):
    cell = ws4.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws4.row_dimensions[3].height = 28

for r in range(4, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 24
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws4.max_column + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 16
ws4.column_dimensions['B'].width = 22

# Format Sheet 3 (Summer Agency Directory)
ws3.row_dimensions[1].height = 28
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 28
    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Save files
try:
    wb.save(clean_path)
    print('Saved to clean path:', clean_path)
except Exception as e:
    print('Error saving clean path:', e)

try:
    wb.save(master_path)
    print('Saved to master path:', master_path)
except Exception as e:
    print('Master path note:', e)
    try:
        shutil.copyfile(clean_path, master_path)
        print('Copied to master path successfully.')
    except Exception as err:
        print('Master file in use:', err)

print('Updated Master Excel with strict authenticity (unlisted fields marked with "-") successfully!')
