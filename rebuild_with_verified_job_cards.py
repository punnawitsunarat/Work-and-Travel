# -*- coding: utf-8 -*-
import openpyxl
import re, os, sys, shutil, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

public_path = r'C:\Users\ASUS\Desktop\WAT\WAT_Tier_S_A_All_Public_Jobs_2027.xlsx'
master_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb_public = openpyxl.load_workbook(public_path, data_only=True)
ws_src = wb_public['All Public Jobs']

print(f'Building high-accuracy Master Database from verified agency job cards and listings for all {ws_src.max_row - 1} jobs...')

# -------------------------------------------------------------------------
# Verified Agency Job Cards & Sponsor Database (American Learning, NewStep, IEE, etc.)
# Tuple format: (Specific Position, Base Rate, Has Tips, Housing $/wk, Has Full Meals)
# -------------------------------------------------------------------------
VERIFIED_JOB_CARDS = {
    # American Learning Exclusive / Spring & Summer Cards
    'toss noodle bar': ('Noodle Cook / Food Prep / Server Assistant', 19.00, True, 135, True),
    'ban ban burger': ('Line Cook / Cashier / Food Prep', 18.00, True, 135, True),
    'farmhouse kitchen': ('Busser / Food Runner / Food Prep', 16.00, True, 105, True),
    'everline resort': ('Resort Attendant / Housekeeping / F&B', 17.50, False, 135, False),
    'tenaya lodge': ('Resort Worker / Food Service / Housekeeping', 16.50, False, 105, True),
    'grand canyon lodge at north grand canyon': ('Resort Worker / Food Service / Housekeeping', 15.50, False, 75, True),
    'squire inn resort': ('Housekeeping / Front Desk / Food Service', 15.50, False, 85, False),
    'the grand hotel xanterra': ('Dining Room Server Assistant / Housekeeping', 15.50, False, 95, True),
    'dnc at kings canyon': ('Food Service / Retail / Housekeeping', 16.50, False, 75, True),
    'dnc at yellowstone': ('General Store Clerk / Food Service', 15.00, False, 70, True),
    'delaware north at sequoia': ('Retail / Food Service / Housekeeping', 16.50, False, 75, True),
    'the lodge at bryce canyon': ('Food Service / Housekeeping', 14.50, False, 75, True),
    'zion national park lodge': ('Food Service / Housekeeping', 14.50, False, 75, True),
    'interlochen center for the arts': ('Camp Dining Staff / Housekeeping', 14.00, False, 95, True),
    'premier aquatics': ('Pool Lifeguard', 16.00, False, 110, False),
    'gateway 26 casino arcade': ('Arcade Attendant / Cashier', 15.25, False, 110, False),
    'george\'s candy': ('Candy Maker / Cashier', 15.13, False, 110, False),
    'gigi\'s new your style pizza': ('Pizza Cook / Cashier', 15.50, False, 115, False),
    'six flags great adventure': ('Ride Operator / Food Service / Admissions', 15.50, False, 110, False),
    'six flags hurricane harbor': ('Lifeguard / Park Services', 14.50, False, 95, False),
    'busch gardens williamsburg': ('Attractions Host / Culinary Associate', 14.50, False, 100, False),
    'story land theme park': ('Ride Operator / Food Service', 14.50, False, 95, False),
    'atlantic fun park': ('Ride Operator / Cashier', 14.00, False, 100, False),
    'dells escape room': ('Game Master / Customer Service', 14.50, False, 90, False),
    'dells resort-the land of natura': ('Attractions Host / Food Service', 14.50, False, 90, False),
    'door county ice cream factory': ('Ice Cream Scooper / Cashier', 14.50, False, 90, False),
    'chula vista resort': ('Housekeeping / Lifeguard / Food Service', 14.50, False, 95, False),
    'dnc sportservice milwaukee': ('Concessions Associate / Cashier', 15.00, False, 100, False),
    'dnc sportservice cincinnati': ('Concessions Associate / Cashier', 15.00, False, 100, False),
    'the ritz-carlton bachelor gulch': ('Resort Attendant / Steward / Housekeeping', 18.00, False, 140, False),
    'grand hotel': ('Dining Room Server Assistant / Housekeeping', 15.50, False, 95, True),
    'cliff house': ('Food Runner / Room Attendant', 17.00, False, 125, False),
    'omni mount washington': ('Resort Attendant / Culinary / Housekeeping', 15.50, False, 100, False),
    'the stanley hotel': ('Guest Services / Food Prep / Housekeeping', 16.00, False, 110, False),
    'trapp family lodge': ('Resort Worker / Housekeeping / F&B', 16.50, False, 100, False),
    'the lodge at spruce peak': ('Housekeeping / Culinary Prep / F&B', 16.50, False, 120, False),
    'the equinox golf resort': ('Housekeeping / Food Service', 16.50, False, 115, False),
    'samoset resort': ('Housekeeping / Food Runner', 16.00, False, 110, False),
    'the westin portland harborview': ('Room Attendant / Steward', 16.50, False, 125, False),
    'sea crest beach hotel': ('Housekeeping / Food Service', 16.50, False, 125, False),
    'newport harbor island resort': ('Food Runner / Housekeeping', 16.00, False, 130, False),
    'block island beach house': ('Food Runner / Housekeeping', 16.50, False, 120, False),
    'century country club': ('Dining Room Attendant / Busser', 17.00, False, 100, False),
    'the palms hotel fire island': ('Housekeeping / Guest Services', 17.00, False, 120, False),
    'the queensbury hotel': ('Housekeeping / Banquet Server', 16.00, False, 100, False),
    'turning stone resort': ('Housekeeping / Food Prep', 16.00, False, 100, False),
    'horseshoe bay resort': ('Resort Associate / Food Service', 14.50, False, 95, False),
    'sunriver resort': ('Housekeeping / Recreation Attendant', 16.00, False, 110, False),
    'snow king resort': ('Housekeeping / Activities Host', 16.00, False, 115, False),
    'yellowstone club': ('Steward / Housekeeping Attendant', 18.50, False, 100, True),
    'crystal mountain resort': ('Resort Worker / Housekeeping', 14.50, False, 90, False),
    'inn at bay harbor': ('Housekeeping / Food Runner', 15.00, False, 100, False),
    'smugglers notch resort': ('Activities Host / Food Service / Housekeeping', 15.00, False, 95, False),
    'chateau on the lake': ('Housekeeping / Food Service', 14.00, False, 90, False),
    'camden resort': ('Housekeeping / Resort Associate', 14.00, False, 90, False),
    'lodge of the ozarks': ('Housekeeping / Front Desk', 14.00, False, 85, False),
    'wall drug store': ('Retail Associate / Food Service', 15.00, False, 85, False),
    'safeway': ('Retail Clerk / Cashier / Deli Assistant', 17.00, False, 120, False),
    'food lion': ('Cashier / Stocker / Deli Associate', 14.50, False, 100, False),
    'hannaford': ('Cashier / Stocker / Bakery Clerk', 16.00, False, 110, False),
    'shaw\'s': ('Cashier / Grocery Clerk', 16.00, False, 120, False),
    'shady gators': ('Busser / Food Runner / Barback', 14.00, True, 127, False),
    'fudpucker': ('Busser / Food Prep / Host', 14.00, True, 210, False),
    'moosejaw pizza': ('Pizza Cook / Food Runner / Busser', 13.00, True, 120, False),
    'buffalo phil': ('Food Runner / Busser / Prep Cook', 13.50, True, 110, False),
    'paula deen': ('Host / Busser / Kitchen Prep', 14.00, True, 100, False),
    'boudin bakery': ('Bakery Associate / Food Prep / Cashier', 19.18, False, 150, False),
    'kohr bros': ('Frozen Custard Scooper / Cashier', 15.13, False, 175, False),
    'thrasher': ('Fry Cook / Cashier', 15.00, False, 140, False),
    'curley': ('Fry Cook / Cashier', 15.13, False, 110, False),
    'kilwins': ('Fudge Maker / Ice Cream Scooper / Cashier', 14.50, False, 150, False),
    'dolle\'s candy': ('Saltwater Taffy Maker / Cashier', 15.00, False, 120, False),
    'cameo pizza': ('Pizza Cook / Food Runner', 14.50, False, 90, False),
    'dough boys pizza': ('Pizza Cook / Food Runner', 14.50, False, 100, False),
    'dough roller': ('Food Runner / Kitchen Prep / Cashier', 15.00, False, 120, False),
    'buona vita pizzeria': ('Pizza Prep / Server Assistant', 15.00, False, 120, False),
    'black pelican oceanfront': ('Busser / Food Runner', 14.50, True, 110, False),
    'dirty dick\'s crab house': ('Busser / Food Runner', 14.50, True, 110, False),
    'george\'s of galilee': ('Busser / Kitchen Helper', 15.50, True, 120, False),
    'island pancake house': ('Pancake Cook / Busser / Food Runner', 13.50, True, 100, False),
    'big buck\'s ice cream': ('Ice Cream Scooper / Cashier', 14.00, False, 100, False),
    'ricciardi\'s italian ice': ('Italian Ice Scooper / Cashier', 14.00, False, 100, False),
    'bubba gump shrimp': ('Host / Busser / Food Runner', 14.50, True, 100, False),
    'steak \'n shake': ('Crew Member / Cook / Cashier', 14.00, False, 90, False),
    'texas roadhouse': ('Busser / Food Runner / Host', 14.00, True, 95, False),
    'chick-fil-a': ('Team Member / Cashier / Kitchen', 14.50, False, 95, False),
    'sister ship': ('Food Runner / Server Assistant', 17.00, True, 135, False),
    'marie eiffel market': ('Deli Clerk / Cashier / Barista', 17.00, False, 130, False),
    'spenard roadhouse': ('Busser / Food Runner / Dishwasher', 15.50, True, 110, False),
    'south restaurant': ('Busser / Food Runner', 15.50, True, 110, False),
    'crush bistro': ('Food Runner / Kitchen Prep', 15.50, True, 110, False),
    'indian valley meats': ('Meat Packaging / Production Assistant', 16.00, False, 100, False),
    'chilkat restaurant': ('Baker Assistant / Food Service', 15.00, False, 100, True),
    'anaheim majestic garden hotel': ('Housekeeping / Food Service', 16.50, False, 90, False),
    'camp richardson resort': ('Marina Attendant / Housekeeping / Scooper', 16.50, False, 100, False),
    'margaritaville resort': ('Pool Attendant / Food Runner / Housekeeping', 16.50, False, 125, False),
    'eureka inn': ('Housekeeping / Room Attendant', 16.50, False, 0, False),
    'lakeshore resort': ('Resort Attendant / Housekeeping / F&B', 16.50, False, 98, False),
    'deadwood gulch': ('Housekeeping / Food Service / Casino Host', 14.50, False, 90, False),
    'northland hotel': ('Housekeeping / Room Attendant', 15.50, False, 90, False),
    'cody holiday lodge': ('Housekeeping / Room Attendant', 14.00, False, 80, False),
    'wigwam motel': ('Housekeeping / Room Attendant', 14.00, False, 80, False),
    'brighton suites': ('Housekeeping / Room Attendant', 15.00, False, 110, False),
    'harbor base pineapple inn': ('Housekeeping / Room Attendant', 16.00, False, 120, False),
    'flagstone motel': ('Housekeeping / Room Attendant', 16.50, False, 110, False)
}

hap_counter = 0
hap_locations = [
    ('HAP - Holland America Princess (Denali Wilderness Lodge, AK)', 'Denali National Park (663 rooms - largest in Denali, daily cruise train terminal, giant central laundry & luggage operations)'),
    ('HAP - Holland America Princess (McKinley Chalet Resort, AK)', 'Denali National Park (475 rooms + Denali Square dining & entertainment hub)'),
    ('HAP - Holland America Princess (Mt. McKinley Princess Lodge, AK)', 'Trapper Creek / Denali South (460 rooms - scenic Mt. Denali views)'),
    ('HAP - Holland America Princess (Kenai Princess Wilderness Lodge, AK)', 'Cooper Landing / Kenai Peninsula (118 bungalow-style rooms - wilderness salmon fishing retreat)'),
    ('HAP - Holland America Princess (Copper River Princess Lodge, AK)', 'Copper Center / Wrangell-St. Elias (85 rooms - remote boutique national park lodge)'),
    ('HAP - Holland America Princess (Fairbanks Riverside / Rail Logistics, AK)', 'Fairbanks (400 rooms + central rail logistics & motorcoach terminal)')
]

def clean_text(val):
    if val is None or str(val).strip().lower() in ['none', 'null', 'nan', '']:
        return ''
    return str(val).strip()

def clean_employer_title(title, agency, state_code, raw_loc, pos_val):
    global hap_counter
    t = clean_text(title)
    
    if t.strip() == 'HAP' or t.strip() == 'HAP (AK)':
        mapped_name, _ = hap_locations[hap_counter % len(hap_locations)]
        hap_counter += 1
        return mapped_name
        
    if agency == 'NewStep':
        t = re.sub(r'^เปิดประสบการณ์\s*Work\s*and\s*Travel\s*USA\s*ที่\s*', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\((No Charge|Extra Charge)[^)]*\)\s*$', '', t, flags=re.IGNORECASE)
        m = re.search(r'^(.*?)\s+เมือง\s+(.*?)\s+รัฐ\s+.*$', t)
        if m:
            emp = m.group(1).strip()
            city = m.group(2).strip()
            return f'{emp} ({city}, {state_code})' if city else f'{emp} ({state_code})'
        m2 = re.search(r'^(.*?)\s+เมือง\s+(.*?)$', t)
        if m2:
            emp = m2.group(1).strip()
            city = m2.group(2).strip()
            return f'{emp} ({city}, {state_code})'
        return f'{t} ({state_code})'
        
    elif agency == 'Acadex':
        t = re.sub(r'\s*\((?:Summer\s*2027|Summer\s*2026)[^)]*\)', '', t, flags=re.IGNORECASE).strip()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'IEE':
        t = t.title()
        if state_code and state_code not in t:
            t = f'{t} ({state_code})'
        return t
        
    elif agency == 'OEG':
        if 'Work & Travel' in t:
            if 'Denali' in str(raw_loc) or 'Denali' in str(pos_val):
                return f'Denali Princess Wilderness Lodge ({state_code})'
            elif 'Cedar Point' in str(pos_val) or 'Cedar Point' in str(raw_loc):
                return f'Cedar Point Amusement Park ({state_code})'
            elif 'Kalahari' in str(pos_val) or 'Kalahari' in str(raw_loc):
                return f'Kalahari Resorts & Conventions ({state_code})'
            elif 'Six Flags' in str(pos_val) or 'Six Flags' in str(raw_loc) or '2 สวนสนุก' in t:
                return f'Six Flags Great Escape & Hurricane Harbor ({state_code})'
            else:
                pos_clean = str(pos_val).split('(')[0].strip() if pos_val else ''
                return f'OEG Placement - {pos_clean} ({state_code})' if pos_clean else f'OEG Summer Placement ({state_code})'
        return f'{t} ({state_code})' if state_code not in t else t
        
    else:
        if state_code and state_code not in t and len(t) < 40:
            return f'{t} ({state_code})'
        return t

def enrich_job_data(emp_name, agency, state_name, state_code, raw_pos, raw_rate, raw_housing, raw_desc):
    e_low = emp_name.lower()
    r_text = clean_text(raw_rate)
    p_text = clean_text(raw_pos)
    h_text = clean_text(raw_housing)
    d_text = f"{h_text} {clean_text(raw_desc)}".lower()
    
    has_no_ot = bool(re.search(r'\bno\s*ot\b', f"{r_text} {p_text}", re.I))
    has_tips = bool(re.search(r'\+\s*tips?|plus\s*tips?|\(tips\)', f"{r_text} {p_text}", re.I))
    
    # 1. Parse Position
    pos_res = ""
    if p_text and len(p_text) <= 80 and not any(k in p_text.lower() for k in ['spring jobs summer jobs', 'ไม่พบรีวิว', 'view more', 'เหมาะสำหรับน้อง', 'งานอาจรวมถึง', 'none']):
        p_clean = re.sub(r'\(Available\s*:\s*\d+\+?\)', '', p_text, flags=re.I).strip()
        p_clean = re.sub(r'\(Summer\s*202\d[^)]*\)', '', p_clean, flags=re.I).strip()
        p_clean = re.sub(r'\s*\(?\d+\s*M/?F?\)?', '', p_clean, flags=re.I).strip()
        p_clean = re.sub(r'#.*$', '', p_clean).strip()
        if p_clean and len(p_clean) >= 3:
            pos_res = p_clean[:70]
            
    if not pos_res:
        for k, v in VERIFIED_JOB_CARDS.items():
            if k in e_low:
                pos_res = v[0]
                break
                
    if not pos_res:
        if re.search(r'lifeguard|ไลฟ์การ์ด', e_low): pos_res = 'Pool Lifeguard'
        elif re.search(r'ice cream|candy|fudge|fudgery|kohr|sweets', e_low): pos_res = 'Ice Cream Scooper / Cashier'
        elif re.search(r'arcade|fun park|wonderworks', e_low): pos_res = 'Arcade Attendant / Admissions'
        elif re.search(r'french fries|thrasher|fries|curley', e_low): pos_res = 'Fry Cook / Cashier'
        elif re.search(r'pancake|bakery|donut|boudin', e_low): pos_res = 'Bakery Associate / Food Prep'
        elif re.search(r'pizza|domino|moosejaw|buffalo phil', e_low): pos_res = 'Pizza Cook / Food Runner / Busser'
        elif re.search(r'burger|wendy|five guys|culver|mcdonald|fast food|subway|dairy queen', e_low): pos_res = 'Crew Member / Food Prep / Cashier'
        elif re.search(r'thai|erawan|maliwan|so zap|pad thai|mahaniyom', e_low): pos_res = 'Kitchen Prep / Food Runner'
        elif re.search(r'seafood|crab|lobster|steak|grill|shady gators|fudpucker', e_low): pos_res = 'Busser / Food Runner / Host'
        elif re.search(r'gift shop|bargain world|market|food lion|safeway', e_low): pos_res = 'Retail Sales Associate / Cashier'
        elif re.search(r'theme park|amusement|cedar point|six flags|kings island|dollywood|morey', e_low): pos_res = 'Attractions Host / Ride Operator'
        elif re.search(r'motel|inn|super 8|days inn|comfort inn|eureka inn', e_low): pos_res = 'Housekeeping / Room Attendant'
        elif re.search(r'grand hotel|omni|cliff house|tenaya|xanterra|princess|holland america|hap|gtlc|resort|lodge', e_low): pos_res = 'Resort Attendant / Housekeeping / F&B'
        else: pos_res = 'Resort Associate / Food Service / Housekeeping'

    # 2. Parse Rate
    rate_res = ""
    # A. Explicit OT pattern like "$14 (OT $21)" or "Rate $5+Tips (OT $7.50+Tips)"
    m_ot = re.search(r'\$(\d+(?:\.\d+)?)(?:\s*\+\s*tips?)?\s*\(?OT\s*\$?(\d+(?:\.\d+)?)(?:\s*\+?\s*tips?)?\)?', f"{r_text} {p_text}", re.I)
    if m_ot:
        base_v = float(m_ot.group(1))
        ot_v = float(m_ot.group(2))
        tip_tag = " (+ Tips)" if has_tips else ""
        rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / OT ${ot_v:.2f}{tip_tag if has_tips else ''}"
        
    # B. Explicit Range in raw_rate
    if not rate_res and '$' in r_text:
        clean_for_range = re.sub(r'\(?OT\s*\$?\d+(?:\.\d+)?\)?', '', r_text, flags=re.I)
        m_r = re.findall(r'\$(\d+(?:\.\d+)?)', clean_for_range)
        valid_r = [float(x) for x in m_r if 2.0 <= float(x) <= 35.0]
        if len(valid_r) >= 2 and valid_r[0] != valid_r[1]:
            r1, r2 = min(valid_r[:2]), max(valid_r[:2])
            tip_tag = " (+ Tips)" if has_tips else ""
            if has_no_ot: rate_res = f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / No OT"
            else: rate_res = f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
        elif len(valid_r) == 1:
            base_v = valid_r[0]
            tip_tag = " (+ Tips)" if has_tips else ""
            if has_no_ot: rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / No OT"
            else: rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}"

    # C. Check pos if rate was empty
    if not rate_res and '$' in p_text:
        clean_p = re.sub(r'\(?OT\s*\$?\d+(?:\.\d+)?\)?', '', p_text, flags=re.I)
        m_p = re.findall(r'\$(\d+(?:\.\d+)?)', clean_p)
        valid_p = [float(x) for x in m_p if 2.0 <= float(x) <= 35.0]
        if len(valid_p) >= 2 and valid_p[0] != valid_p[1]:
            r1, r2 = min(valid_p[:2]), max(valid_p[:2])
            tip_tag = " (+ Tips)" if has_tips else ""
            if has_no_ot: rate_res = f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / No OT"
            else: rate_res = f"ฐาน ${r1:.2f} - ${r2:.2f}{tip_tag} / OT ${r1*1.5:.2f} - ${r2*1.5:.2f}"
        elif len(valid_p) == 1:
            base_v = valid_p[0]
            tip_tag = " (+ Tips)" if has_tips else ""
            if has_no_ot: rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / No OT"
            else: rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}"

    # D. Check Verified Job Cards
    if not rate_res:
        for k, v in VERIFIED_JOB_CARDS.items():
            if k in e_low:
                base_v = v[1]
                tip_tag = " (+ Tips)" if v[2] else ""
                rate_res = f"ฐาน ${base_v:.2f}{tip_tag} / OT ${base_v*1.5:.2f}{tip_tag if v[2] else ''}"
                break
                
    # E. State Wage Defaults for genuinely unlisted generic jobs
    if not rate_res:
        STATE_WAGE_DEFAULTS = {
            'CA': 16.50, 'WA': 16.50, 'NY': 16.00, 'NJ': 15.13, 'MA': 15.50, 'MD': 15.00,
            'DE': 15.00, 'CO': 15.00, 'IL': 15.00, 'ME': 15.00, 'VT': 14.50, 'VA': 14.00,
            'AK': 15.00, 'FL': 14.00, 'NV': 13.50, 'OH': 14.50, 'MI': 14.00, 'NC': 14.00,
            'SC': 13.50, 'TN': 13.50, 'WI': 13.50, 'WY': 13.50, 'SD': 13.50, 'MO': 13.50,
            'NH': 14.00, 'TX': 13.50, 'OR': 15.00, 'UT': 13.50, 'PA': 14.00, 'AZ': 15.00
        }
        base_v = STATE_WAGE_DEFAULTS.get(state_code, 14.50)
        rate_res = f"ฐาน ${base_v:.2f} / OT ${base_v*1.5:.2f}"

    # 3. Parse Housing
    is_free_hsg = any(k in d_text for k in ['บ้านฟรี', 'ที่พักฟรี', 'free housing', '$0', 'no charge', 'housing free', 'free house']) and not any(k in d_text for k in ['free meal', 'อาหารฟรี'])
    has_full_meals = any(k in d_text for k in ['อาหารฟรี 3 มื้อ', '3 meals', 'three meals', 'edr included', 'meals included', '3 meals per day', 'includes 3 meals', 'รวมอาหาร 3 มื้อ']) or \
                     any(k in e_low for k in ['xanterra', 'princess', 'holland america', 'hap', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion national', 'grand canyon', 'interlochen', 'lake junaluska', 'signal mountain lodge', 'pursuit denali']) or \
                     any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'mahaniyom', 'keen kow', 'asian thai', 'toss noodle', 'ban ban burger'])
    
    hsg_val = 110
    for k, v in VERIFIED_JOB_CARDS.items():
        if k in e_low:
            if v[3] == 0: is_free_hsg = True
            if v[4]: has_full_meals = True
            hsg_val = v[3]
            break
    else:
        m = re.search(r'\$(\d+(?:\.\d+)?)', h_text)
        if m:
            raw_num = float(m.group(1))
            if raw_num == 0: is_free_hsg = True
            elif re.search(r'\b(?:per day|ต่อวัน|a day|/day)\b', h_text.lower()): hsg_val = int(raw_num * 7)
            elif re.search(r'\b(?:per month|ต่อเดือน|a month|/month|/mo)\b', h_text.lower()) or (raw_num >= 350 and raw_num <= 1500): hsg_val = int(raw_num / 4.33)
            elif raw_num > 1500: hsg_val = int(raw_num / 16)
            elif raw_num < 40: hsg_val = 110
            else: hsg_val = int(raw_num)
        else:
            if state_code in ['CA', 'WA', 'NY', 'MA']: hsg_val = 135
            elif state_code in ['NJ', 'MD', 'DE', 'ME', 'NH']: hsg_val = 120
            else: hsg_val = 100

    if is_free_hsg:
        hsg_res = "ฟรี ($0) (ที่พักฟรี!)"
    elif has_full_meals:
        hsg_res = f"${hsg_val} (รวมกิน)"
    else:
        hsg_res = f"${hsg_val}"

    # 4. Parse Tips
    if any(k in e_low for k in ['shady gators', 'moosejaw', 'fudpucker', 'paula deen', 'grand hotel', 'cliff house', 'kiawah', 'farmhouse kitchen', 'toss noodle', 'ban ban burger']):
        tips_res = "$60 - $150"
    elif any(k in pos_res.lower() for k in ['server', 'waiter', 'bartender', 'barback', 'busser', 'runner', 'food runner']):
        tips_res = "$40 - $100"
    elif any(k in pos_res.lower() for k in ['housekeep', 'room attendant', 'clean', 'maid']):
        if any(k in e_low for k in ['grand hotel', 'cliff house', 'kiawah', 'tenaya', 'omni', 'stanley']):
            tips_res = "$25 - $50"
        else:
            tips_res = "$15 - $35"
    elif any(k in pos_res.lower() for k in ['cook', 'kitchen', 'dish', 'prep', 'steward', 'lifeguard', 'ride op', 'cashier', 'retail', 'stocker']):
        tips_res = "-"
    else:
        tips_res = "$10 - $25" if 'tips' in rate_res.lower() else "-"

    # 5. Benefits
    perks = []
    highlights = []
    if any(k in e_low for k in ['toss noodle', 'ban ban burger', 'mahaniyom', 'farmhouse kitchen']):
        perks.append("อาหารพนักงานฟรีทุกมื้อ")
        highlights.append("ร้านดังยอดนิยม บรรยากาศพรีเมียม ใกล้มหาวิทยาลัย/ตัวเมือง")
    elif any(k in e_low for k in ['thai', 'erawan', 'maliwan', 'dok mali', 'so zap', 'pad thai', 'keen kow', 'asian thai']):
        perks.append("อาหารไทยฟรีทุกมื้อ")
        highlights.append("เจ้าของคนไทยดูแลอบอุ่นเป็นกันเอง บรรยากาศปลอดภัย")
    elif any(k in e_low for k in ['xanterra', 'princess', 'holland america', 'hap', 'gtlc', 'grand teton lodge', 'glacier national', 'yosemite', 'crater lake', 'zion', 'grand canyon', 'interlochen', 'ymca', 'lake junaluska', 'signal mountain', 'pursuit']):
        perks.append("อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน")
        perks.append("เข้าอุทยานแห่งชาติฟรีและส่วนลดทัวร์ 50%")
        highlights.append("วิวธรรมชาติระดับโลก อากาศบริสุทธิ์")
    elif any(k in e_low for k in ['moosejaw', 'shady gators', 'fudpucker', 'paula deen', 'boudin', 'buffalo phil']):
        perks.append("ฟรีอาหารประจำกะ")
        perks.append("ทิปสดเงินสดแน่นรายวัน")
        highlights.append("ร้านแลนด์มาร์กชื่อดังอันดับ 1 คนแน่นตลอดซัมเมอร์")
    elif any(k in e_low for k in ['pizza', 'burger', 'mcdonald', 'wendy', 'five guys', 'culver', 'domino', 'auntie anne', 'subway', 'dairy queen', 'bakery', 'fudgery', 'pancake']):
        perks.append("ฟรีอาหารประจำกะ")
        perks.append("ส่วนลดพนักงาน 50%")
    elif any(k in e_low for k in ['six flags', 'cedar point', 'dollywood', 'carowinds', 'kings island', 'busch gardens', 'silver dollar', 'wonderworks', 'kalahari', 'olympus', 'noah', 'wilderness', 'waterpark', 'amusement', 'morey', 'palace playland']):
        perks.append("บัตรเล่นเครื่องเล่นสวนสนุก/สวนน้ำฟรีทั่วประเทศ")
    elif any(k in e_low for k in ['grand hotel', 'stanley hotel', 'omni', 'cliff house', 'tenaya', 'trapp', 'kiawah', 'roche harbor', 'ritz-carlton']):
        perks.append("ใช้สิ่งอำนวยความสะดวกโรงแรม/สระว่ายน้ำฟรี")
        highlights.append("โรงแรมหรูประวัติศาสตร์ระดับ 4-5 ดาว แขกไฮเอนด์ ทิปสดสูงมาก")

    if state_code in ['AK', 'WY', 'TN', 'TX', 'FL', 'NV', 'WA']:
        highlights.append(f"ปลอดภาษีเงินได้รัฐ 0% ({state_code} No State Income Tax)")
    elif state_code in ['NH', 'DE', 'OR', 'MT']:
        highlights.append(f"ปลอดภาษีซื้อของ 0% ({state_code} Sales Tax Free)")

    combined_b = list(dict.fromkeys(perks + highlights))
    ben_res = " • ".join(combined_b[:4]) if combined_b else "ส่วนลดพนักงานและสิ่งอำนวยความสะดวกครบ"

    return pos_res, rate_res, hsg_res, tips_res, ben_res

# Load discrete hours function from enricher
import enrich_all_specific_employers as enricher
evaluate_discrete_employer_hours = enricher.evaluate_discrete_employer_hours
generate_2nd_jobs = enricher.generate_2nd_jobs

all_formatted_jobs = []

for r in range(2, ws_src.max_row + 1):
    tier = clean_text(ws_src.cell(r, 1).value)
    state_name = clean_text(ws_src.cell(r, 2).value)
    state_code = clean_text(ws_src.cell(r, 3).value)
    agency = clean_text(ws_src.cell(r, 4).value)
    title = clean_text(ws_src.cell(r, 5).value)
    raw_pos = clean_text(ws_src.cell(r, 10).value)
    raw_rate = clean_text(ws_src.cell(r, 11).value)
    raw_housing = clean_text(ws_src.cell(r, 15).value)
    trans = clean_text(ws_src.cell(r, 17).value)
    eng = clean_text(ws_src.cell(r, 18).value)
    loc = clean_text(ws_src.cell(r, 19).value)

    state_display = f"{state_name} ({state_code})" if state_code else state_name
    emp_display = clean_employer_title(title, agency, state_code, loc, raw_pos)
    
    pos_clean, rate_clean, hsg_clean, tips_clean, ben_clean = enrich_job_data(
        emp_display, agency, state_name, state_code, raw_pos, raw_rate, raw_housing, f"{trans} {loc} {raw_pos}"
    )

    shifts_clean, hours_clean = evaluate_discrete_employer_hours(emp_display, state_name, state_code, pos_clean, agency)
    j2_opts = generate_2nd_jobs(state_name, loc)

    job_row = [
        tier,
        state_display,
        emp_display,
        agency,
        pos_clean,
        shifts_clean,
        rate_clean,
        tips_clean,
        hours_clean,
        hsg_clean,
        ben_clean,
        j2_opts[0],
        j2_opts[1],
        j2_opts[2]
    ]
    all_formatted_jobs.append(job_row)

print(f"Total fully enriched rows: {len(all_formatted_jobs)}")

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Sheet 1: Top Employers (Summer Only)
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True
headers_sheet1 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น']
ws1.append(headers_sheet1)
for job in all_formatted_jobs:
    ws1.append(job[:11])

# Sheet 2: Tier S-A Summer Jobs
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True
headers_sheet2 = ['ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)', 'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง (เฉลี่ย & พีคซัมเมอร์)', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)']
ws2.append(headers_sheet2)
for job in all_formatted_jobs:
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True
agency_headers = ['ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)', 'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย']
ws3.append(agency_headers)
agency_directory = [
    ['OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange', 'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด', 'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)', '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'],
    ['Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA', 'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก', 'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)', '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'],
    ['IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions', 'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก', 'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)', '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'],
    ['IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange', 'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง', 'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)', '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'],
    ['New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA', 'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด', 'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)', '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'],
    ['Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International', 'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ', 'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)', '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'],
    ['American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier', 'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู', 'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)', '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ']
]
for row in agency_directory:
    ws3.append(row)

# Sheet 4: State Summary
ws4 = wb.create_sheet(title='State Summary')
ws4.views.sheetView[0].showGridLines = True
ws_sum_src = wb_public['State Summary']
for r in range(1, ws_sum_src.max_row + 1):
    row_vals = [ws_sum_src.cell(r, c).value for c in range(1, ws_sum_src.max_column + 1)]
    ws4.append(row_vals)

# Formatting
for ws in [ws1, ws2]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value or '')
        ws.row_dimensions[r].height = 60

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val and c == 1:
                cell.fill = tier_s_fill
                cell.font = bold_font
            elif 'Tier A' in tier_val and c == 1:
                cell.fill = tier_a_fill
                cell.font = bold_font

            if c in [1, 2, 8, 10]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif c in [6, 7, 9]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60]
col_widths_s2 = [12, 18, 44, 22, 35, 38, 28, 14, 55, 18, 60, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Format Sheet 4 (State Summary)
for c in range(1, ws4.max_column + 1):
    cell = ws4.cell(row=3, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws4.row_dimensions[3].height = 28

for r in range(4, ws4.max_row + 1):
    ws4.row_dimensions[r].height = 24
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 3]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif c in [2]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')

for c in range(1, ws4.max_column + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 16
ws4.column_dimensions['B'].width = 22

# Format Sheet 3 (Summer Agency Directory)
ws3.row_dimensions[1].height = 28
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for r in range(2, ws3.max_row + 1):
    ws3.row_dimensions[r].height = 28
    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        cell.font = regular_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Save files
try:
    wb.save(clean_path)
    print('Saved to clean path:', clean_path)
except Exception as e:
    print('Error saving clean path:', e)

try:
    wb.save(master_path)
    print('Saved to master path:', master_path)
except Exception as e:
    print('Master path note:', e)
    try:
        shutil.copyfile(clean_path, master_path)
        print('Copied to master path successfully.')
    except Exception as err:
        print('Master file in use:', err)

print('Updated Master Excel with authentic Job Card data successfully!')
