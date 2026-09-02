import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

excel_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027.xlsx'
excel_clean_path = r'C:\Users\ASUS\Desktop\WAT\Work_and_Travel_Master_Job_Database_2027_Clean.xlsx'

wb = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
regular_font = Font(name='Segoe UI', size=10)
bold_font = Font(name='Segoe UI', size=10, bold=True)
tier_s_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
tier_a_fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

master_summer_jobs = [
    # =========================================================================
    # 🏆 TIER S STATES (12 States)
    # =========================================================================

    # 1. ALASKA (AK)
    [
        'Tier S', 'Alaska (AK)',
        'Denali Princess Wilderness Lodge (Summer Season)',
        'Acadex, OEG, IEE, Higher, IEO, ALC',
        'Housekeeping / Room Attendant',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $16.00 / OT $24.00', '$15 - $30', '50 – 54 ชม./wk', '$105 (รวมกิน)',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน, ทิปหัวเตียง, รถรับส่งพนักงาน, ส่วนลดทัวร์/รถไฟ 50%',
        '🚌 Fleet Detailer (ในโรงแรม): ล้างรถทัวร์กะดึก 21:00-01:30 น. ได้เรต OT $24.00/ชม. ทันที (ไม่ต้องขอสปอนเซอร์)',
        '🍕 Prospectors Pizzeria: เดิน 10 นาที ล้างจาน/Barback 18:30-00:30 น. ($15/ชม. + ทิปสด $40-$70/คืน)',
        '🍽️ Fannie Q’s Saloon (ในโรงแรม): ช่วยจัดเลี้ยง/Runner 17:00-22:00 น. ได้เรต OT $24/ชม. + ทิป'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Denali Princess Wilderness Lodge (Summer Season)',
        'Acadex, OEG, IEE, Higher, IEO',
        'Kitchen Stewarding / Dishwasher (ล้างจาน)',
        'ปกติ: 16:00 – 00:00 (8 ชม.)\nพีคซัมเมอร์: 15:00 – 01:30 (10.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '-', '50 – 55 ชม./wk', '$105 (รวมกิน)',
        'อาหารบุฟเฟต์ 3 มื้อใน EDR + เชฟทำสเต๊ก/เบอร์เกอร์เลี้ยงในครัวฟรีทุกกะ',
        '🧹 เช้าช่วยแม่บ้าน (Housekeeping): 08:30-14:30 น. ขอกะเช้าเพิ่มในโรงแรม ได้เรต OT $24/ชม.',
        '🧳 Luggage Handler: ช่วยยกกระเป๋ารอบรถไฟเช้า 08:00-12:00 น. ได้ทิปสดจากทัวร์',
        '🛍️ Denali Boardwalk Gift Shops: แคชเชียร์/จัดของกะสาย 09:00-14:00 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'McKinley Chalet Resort (Holland America-Princess, Denali)',
        'OEG, Acadex, Higher, IEO, ALC',
        'Housekeeping / Culinary / F&B',
        'ปกติ: 07:30 – 15:30 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$15 - $30', '48 – 52 ชม./wk', '$105 (รวมกิน)',
        'รีสอร์ตหรูริมแม่น้ำ Nenana River ฝั่ง Denali, อาหาร 3 มื้อฟรีใน EDR, ส่วนลดทัวร์ล่องแก่ง',
        '🥩 Karstens Public House (ในรีสอร์ต): Busser/Barback 17:30-23:00 น. (ขอ OT ในรีสอร์ต $24/ชม. + ทิป)',
        '🍕 Grizzly Burger / Canyon Pizzeria: ผู้ช่วยครัว/ล้างจาน 18:00-23:30 น. ($15/ชม. + ทิป)',
        '🛒 Denali Square Retail Shops: แคชเชียร์รอบค่ำ 17:00-21:30 น. ($15.50/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Grande Denali Lodge (Denali National Park, AK)',
        'Acadex Thailand, IEO, Higher',
        'General Cleaning / Hostess / Bussing Tables / Crew Member',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $14.00 - $15.00 / OT $21.00 - $22.50', '-', '42 – 46 ชม./wk', '$140',
        'โบนัส Summer Bonus +$1.00/ชม. ท้ายทริป, โรงแรมหรู 4 ดาวบนยอดเขาชมวิวอุทยาน Denali, รถชัตเติลบัสรับส่ง',
        '🍷 Alpenglow Restaurant (ในโรงแรม): Busser 17:00-23:00 น. ($14/ชม. + ทิปสดเศรษฐี $50-$80/คืน)',
        '🥪 Subway Denali / The Black Bear: นั่งชัตเติลลงเขาไปทำกะเย็น 17:00-22:00 น. ($15/ชม.)',
        '🍽️ Denali Princess Kitchens: เดินลงเขา 15 นาที ล้างจานกะค่ำ 18:00-00:00 น. ($16/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Denali Bluffs Hotel (Denali National Park, AK)',
        'Acadex Thailand, IEO, Higher',
        'General Cleaning / Hostess / Bussing Tables / Crew Member',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 46 ชม./wk', '$140',
        'โบนัส Summer Bonus +$1.00/ชม. ท้ายทริป, โรงแรมสไตล์กระท่อมไม้บนเนินเขา Denali Canyon, รถชัตเติลบัส',
        '🥩 Mountaineer Grill (ในโรงแรม): Busser/Runner 17:00-22:30 น. (ขอ OT $22.50/ชม. + ทิป)',
        '🍕 Prospectors Pizzeria: เดิน 5 นาที ล้างจาน/บัสเซอร์กะค่ำ 18:30-00:30 น. ($15/ชม. + ทิป)',
        '☕ Moose-A.K.A. Espresso: แคชเชียร์/ผู้ช่วยบาริสต้า 06:30-10:30 น.'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Mountain High Pizza Pie (Talkeetna, Alaska)',
        'Acadex Thailand, Higher, OEG',
        'Food Preparation / Dishwashing / Crew Member (May & June Arrival)',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 22:30 (11 ชม.)',
        'ฐาน $13.00 / OT $19.50', '$20 - $50', '40 – 46 ชม./wk', '$100',
        'ร้านพิซซ่าเตาฟืนชื่อดังที่สุดในเมือง Talkeetna คนแน่นตลอดซัมเมอร์, เจ้าของร้านใจดีมาก, หอพักเคบินพนักงาน',
        '🍺 Denali Brewing Company (Talkeetna Pub): Barback กะค่ำ 18:30-00:30 น. ($14/ชม. + ทิปสด $60-$100/คืน)',
        '🏨 Talkeetna Alaskan Lodge: ขอกะเช้าช่วยแม่บ้าน 08:00-12:00 น. ($16/ชม.)',
        '🥐 Talkeetna Roadhouse: ช่วยเตรียมเบเกอรี่/ล้างจานช่วงเช้าตรู่ ($14.50/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Keen Kow Thai Food (Ninilchik, Alaska)',
        'OEG, Acadex, Higher',
        'Food Preparation / Dishwashing / Crew Member',
        'ปกติ: 10:30 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:00 – 21:00 (10 ชม.)',
        'ฐาน $14.00 / OT $21.00', '$20 - $50', '42 – 48 ชม./wk', '$90 (รวมกิน)',
        'ร้านอาหารไทยยอดฮิตในเมืองตากอากาศตกปลาแซลมอน Ninilchik (Kenai Peninsula), ทานอาหารไทยฟรีทุกมื้อ, เจ้าของคนไทยใจดี',
        '🐟 Deep Creek Halibut & Salmon Charters: ช่วยล้างทำความสะอาดเรือตกปลา 18:00-21:30 น. ($16/ชม. + ทิปปลาสด)',
        '🏕️ Ninilchik Cabins & RV Park: ผู้ช่วยดูแลความสะอาดแคมป์ 07:00-09:30 น. ($15/ชม.)',
        '🛒 Ninilchik General Store: แคชเชียร์รอบค่ำ 18:30-22:00 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        '49th State Brewing Company (Healy & Anchorage, AK)',
        'OEG, Acadex, IEO, ALC',
        'Dishwasher / Line Cook / Server Assistant (Busser)',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 23:00 (11.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$30 - $70', '46 – 52 ชม./wk', '$110',
        'โรงเบียร์คราฟต์และร้านอาหารชื่อดังที่สุดในอลาสกา สาขา Healy (หน้าทางเข้า Denali NP) มีรถเมล์ Into the Wild, ทิปสดแน่น',
        '🥩 49th State Outdoor Beer Garden: Busser/Runner โซนลานเบียร์กลางแจ้ง 17:30-23:30 น. (ทิปสด $60-$110/คืน)',
        '🍕 Otto’s Healy Pizzeria: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น. ($15/ชม.)',
        '🚌 Denali Park Transit Helper: ช่วยตรวจตั๋วรถบัสอุทยานรอบเช้า 07:00-10:00 น.'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Four Points by Sheraton Anchorage Downtown (Alaska)',
        'OEG, Acadex, Higher',
        'Housekeeping / Busser / Front Desk Support',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$15 - $35', '42 – 48 ชม./wk', '$125',
        'โรงแรมใจกลางเมืองแองเคอเรจ (เมืองใหญ่ที่สุดในอลาสกา), มีรถเมล์ People Mover เดินทางสะดวก หางานสองง่ายที่สุด',
        '🦀 Glacier Brewhouse (ร้านซีฟู้ดอันดับ 1 ในเมือง): Busser/Dishwasher 18:00-00:30 น. (ทิปสด $70-$130/คืน)',
        '🍜 Thai Kitchen Anchorage: ผู้ช่วยครัว/เสิร์ฟ 17:30-22:30 น. (ทิปสด + กินอาหารไทยฟรี)',
        '🛒 Anchorage 5th Avenue Mall: พนักงานร้านค้า/ปิดร้าน 18:00-21:30 น. ($16/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Seward Windsong Lodge & Kenai Fjords Tours (Pursuit)',
        'OEG, Acadex, ALC, IEO',
        'Housekeeping / Tour Operations / F&B',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '44 – 50 ชม./wk', '$120',
        'เมืองตากอากาศชายทะเล Seward ท่าเรือล่องเรือชมวาฬและธารน้ำแข็งซัมเมอร์, หอพักพนักงาน, ล่องเรือฟรี',
        '🦀 Ray’s Waterfront (ริมท่าเรือ Seward): Busser/Steward ร้านซีฟู้ดดัง 17:30-23:00 น. ($14/ชม. + ทิปสด $60-$100)',
        '🍺 Seward Brewing Company: Barback/Dishwasher กะค่ำ 18:00-00:00 น. ($15/ชม. + ทิป)',
        '🛥️ Kenai Fjords Boat Cleaning: ล้างทำความสะอาดเรือทัวร์รอบเย็น 18:00-21:30 น. ($16.50/ชม.)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Talkeetna Alaskan Lodge (Pursuit Collection)',
        'OEG, Acadex, Higher, IEO',
        'Housekeeping / Kitchen / Steward',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$15 - $30', '44 – 50 ชม./wk', '$115',
        'โรงแรมหรูชมวิวยอดเขา Denali ในเมืองคาวบอย Talkeetna, หอพักพนักงาน + รถรับส่ง',
        '🥩 Foraker Dining Room (ในโรงแรม): Busser ร้านอาหาร Fine Dining 17:30-23:00 น. (ทิปสด $50-$90/คืน)',
        '🍕 Mountain High Pizza Pie: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น. ($15/ชม. + ทิป)',
        '🍺 Denali Brewing Company (Talkeetna Pub): Barback กะค่ำ 18:30-00:30 น. ($14/ชม. + ทิปสด)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Alyeska Resort & Hotel Alyeska (Girdwood)',
        'OEG, Higher, Acadex, ALC',
        'Housekeeping / Culinary / Lift Ops',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '-', '42 – 48 ชม./wk', '$135',
        'รีสอร์ตหรูระดับโลก นั่งกระเช้าชมวิวภูเขาและธารน้ำแข็งซัมเมอร์, หอพักพนักงาน Glacier Valley',
        '🥩 Seven Glaciers Restaurant (บนยอดเขา): Busser 17:30-23:00 น. (ร้าน 4 ดาว AAA ทิปมหาศาล $70-$120)',
        '🍕 Coast Pizza / Girdwood Brewing: ผู้ช่วยครัว/Barback 18:00-23:00 น. ($15/ชม. + ทิป)',
        '🥖 The Bake Shop Girdwood: ช่วยเตรียมเบเกอรี่/ล้างจานช่วงเช้าตรู่หรือเย็น'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Mt. McKinley Princess Lodge (Trapper Creek)',
        'Higher, OEG, Acadex, IEO',
        'Housekeeping / Kitchen',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$10 - $20', '44 – 48 ชม./wk', '$105 (รวมกิน)',
        'อาหารบุฟเฟต์ 3 มื้อฟรีใน EDR, รถรับส่ง, ชมวิวยอดเขาเดนาลี บรรยากาศสงบ',
        '🥩 20,320 Alaskan Grill (ในโรงแรม): Busser/Steward กะค่ำ 17:30-22:30 น. (เรต OT $24/ชม. + ทิป)',
        '🍕 Northland Wood Fired Pizza (ในโรงแรม): ผู้ช่วยทำพิซซ่า/ครัว 16:30-21:30 น. (เรต OT $24/ชม.)',
        '🥐 Talkeetna Roadhouse: นั่งรถตู้พนักงานเข้าเมือง Talkeetna ช่วยล้างจาน/ทำขนมวันหยุด'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Copper River Princess Wilderness Lodge (Copper Center)',
        'Higher, Acadex, IEO',
        'Housekeeping / Culinary / Guest Services',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$10 - $20', '44 – 48 ชม./wk', '$105 (รวมกิน)',
        'รีสอร์ตริมแม่น้ำ Copper River และอุทยาน Wrangell-St. Elias NP, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Two Rivers Restaurant (ในโรงแรม): Busser/Runner 17:30-22:30 น. (ขอ OT $24/ชม. + ทิป)',
        '🍔 Whistle Stop Grill: ผู้ช่วยทำเบอร์เกอร์/ล้างจาน 16:30-21:30 น. (ขอ OT $24/ชม.)',
        '🐟 Klutina River Salmon Charters: ช่วยทำความสะอาดอุปกรณ์ตกปลาแซลมอน 17:00-20:30 น.'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Fairbanks Princess Riverside Lodge (Summer Season)',
        'IEE, Acadex, Higher, IEO',
        'Food Runner / Laundry / Steward',
        'ปกติ: 06:30 – 14:30 หรือ 15:00 – 23:00\nพีคซัมเมอร์: กะสลับตามรอบรถไฟ 8 ชม.',
        'ฐาน $15.50 / OT $23.25', '$20 - $40', '36 – 40 ชม./wk', '$105',
        'หอพักในเมืองแฟร์แบงก์ส + รถชัตเติลบัส, ใกล้ Walmart และร้านอาหารไทย หางาน 2 ง่าย',
        '🛒 Walmart Supercenter Fairbanks: จัดสต็อกสินค้า/แคชเชียร์ 16:30-22:30 น. ($16-$17/ชม. มีรถเมล์ผ่าน)',
        '🍜 Pad Thai Restaurant / Thai House: ผู้ช่วยครัว/เสิร์ฟ 17:00-22:30 น. (ทิปสด + กินอาหารไทยฟรี)',
        '🍺 The Pump House Restaurant: Busser/Dishwasher ร้านอาหารริมน้ำดัง 17:30-23:00 น. ($15/ชม. + ทิป)'
    ],
    [
        'Tier S', 'Alaska (AK)',
        'Skagway Westmark Inn & Chilkoot Tours (Holland America)',
        'OEG, Acadex, ALC',
        'Housekeeping / Tour Host / Retail',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '44 – 50 ชม./wk', '$110',
        'เมืองตากอากาศประวัติศาสตร์ยุคตื่นทอง ท่าเรือเรือสำราญแวะทุกวัน, เดินเท้าได้ทั่วเมือง',
        '🍺 Red Onion Saloon (บาร์ประวัติศาสตร์): Busser/Runner 17:30-23:30 น. (ทิปสด $60-$120/คืน)',
        '🦀 Skagway Fish Company: ผู้ช่วยครัว/ล้างจาน 18:00-23:00 น. ($15/ชม. + ทิป)',
        '🚂 White Pass & Yukon Route Depot: ช่วยทำความสะอาดขบวนรถไฟโบราณ 17:00-21:30 น. ($16/ชม.)'
    ],

    # 2. WYOMING (WY)
    [
        'Tier S', 'Wyoming (WY)',
        'Grand Teton Lodge Company - Jackson Lake Lodge (Summer)',
        'Acadex, Higher, OEG, IEE, IEO, ALC',
        'Housekeeping / Crew Member',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:30 (10 ชม. ควงกะ)',
        'ฐาน $17.75 / OT $26.62', '-', '54 – 60 ชม./wk', '$105 (รวมกิน)',
        'หอพัก + อาหารบุฟเฟต์ 3 มื้อ EDR ตักไม่อั้น 7 วัน ($105/wk), ส่วนลดซื้อของ 40%, เข้าอุทยาน Grand Teton & Yellowstone ฟรี',
        '🍽️ Mural Room & Pioneer Grill (ในโรงแรม): Busser 17:30-22:30 น. ได้เรต OT $26.62/ชม. + ทิป',
        '🍸 Blue Heron Lounge (ในโรงแรม): Barback ช่วยยกน้ำแข็ง/ล้างแก้ว 18:00-23:30 น. (เรต OT $26.62 + Tip-Out)',
        '🛶 Colter Bay Chuckwagon & Marina: นั่งชัตเติลฟรีไปช่วยมินิมาร์ท/เรือแคนู 17:00-21:30 น.'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Grand Teton Lodge Company - Colter Bay Village (Summer)',
        'Acadex, Higher, OEG, IEE, IEO',
        'Cabin Attendant / Grocery Clerk / Marina',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.75 / OT $26.62', '-', '52 – 58 ชม./wk', '$105 (รวมกิน)',
        'หอพัก + อาหารบุฟเฟต์ 3 มื้อ EDR ($105/wk), หมู่บ้านริมทะเลสาบ Jackson Lake สวยที่สุดในซัมเมอร์',
        '🥩 Chuckwagon Restaurant (ในหมู่บ้าน): Busser/Steward บุฟเฟต์มื้อเย็น 17:00-22:00 น. (เรต OT $26.62)',
        '🛒 Colter Bay General Store: แคชเชียร์/เติมสต็อกมินิมาร์ท 17:30-22:00 น. (เรต OT $26.62)',
        '🛶 Colter Bay Marina: ล้างเรือยนต์/เรือแคนู 16:30-20:30 น. (เรต OT $26.62)'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Grand Teton Lodge Company - Jackson Lake Lodge (Culinary)',
        'Acadex, Higher, OEG',
        'Kitchen Steward / Line Cook Helper',
        'ปกติ: 15:30 – 23:30 (8 ชม.)\nพีคซัมเมอร์: 14:30 – 00:30 (10 ชม.)',
        'ฐาน $17.75 / OT $26.62', '-', '52 – 58 ชม./wk', '$105 (รวมกิน)',
        'อาหาร 3 มื้อ EDR ฟรี + อาหารปรุงสดจากครัวเชฟ, มีโอกาสขอกะเช้าควงกะรับ OT มหาศาล',
        '🧹 กะเช้าช่วยแม่บ้าน (Housekeeping): 08:30-14:30 น. ขอทำเพิ่มรับเรต OT $26.62/ชม.',
        '🧳 Bell Staff Helper: ช่วยยกกระเป๋ารอบรถทัวร์เช้า 08:00-12:00 น. ได้ทิปสด',
        '🛒 Jackson Lake Lodge Gift Shop: จัดของรอบสาย 09:00-14:00 น. (เรต OT)'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Signal Mountain Lodge (Grand Teton National Park)',
        'OEG, Higher, Acadex, IEO',
        'Housekeeping / Marina / F&B',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$15 - $35', '48 – 54 ชม./wk', '$100 (รวมกิน)',
        'รีสอร์ตริมหาดทะเลสาบ Jackson Lake, อาหารพนักงานฟรีใน EDR, เช่าเรือพายเล่นฟรี',
        '🥩 Trapper Grill (ในรีสอร์ต): Busser/Dishwasher 17:30-23:00 น. (ขอ OT $24.75/ชม. + ทิปสด)',
        '🛶 Signal Mountain Marina: ดูแลเรือเช่า/ล้างเรือ 16:30-20:30 น. (ขอ OT ในรีสอร์ตได้)',
        '⛽ General Store & Gas Station: แคชเชียร์รอบค่ำ 17:00-21:30 น.'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Xanterra Yellowstone - Old Faithful Snow Lodge & Inn',
        'IEE, Acadex, Higher, IEO, ALC',
        'Housekeeping / Kitchen Steward / Guest Services',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '-', '48 – 54 ชม./wk', '$105 (รวมกิน)',
        'ทำงานหน้าแลนด์มาร์กระดับโลก น้ำพุร้อน Old Faithful Geyser, อาหาร 3 มื้อฟรีใน EDR ตักไม่อั้น 7 วัน',
        '🥩 Old Faithful Inn Dining Room: Busser/Runner 17:30-22:30 น. (ขอ OT ใน Xanterra $24.00 + ทิป)',
        '🍦 Bear Paw Ice Cream & Snack Shop: แคชเชียร์/ตักไอศกรีม 16:30-21:30 น. (เรต OT $24.00)',
        '🛒 Old Faithful Basin Store: จัดของมินิมาร์ทกะค่ำ 18:00-22:00 น.'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Xanterra Yellowstone - Canyon Lodge & Cabins',
        'IEE, Acadex, Higher, IEO',
        'Housekeeping / Fast Food Crew',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '-', '48 – 52 ชม./wk', '$105 (รวมกิน)',
        'หมู่บ้านที่พักที่ใหญ่ที่สุดใน Yellowstone อยู่ติด Grand Canyon of the Yellowstone, อาหาร 3 มื้อ EDR',
        '🍔 Canyon Eatery Food Court: พนักงานเสิร์ฟอาหาร/แคชเชียร์ 17:00-22:00 น. (เรต OT $24.00)',
        '🥩 Canyon Dining Room: Steward/Dishwasher กะดึก 18:00-23:30 น. (เรต OT $24.00)',
        '🛒 Canyon General Store (Delaware North): แคชเชียร์รอบค่ำ'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Xanterra Yellowstone - Mammoth Hot Springs & Lake Hotel',
        'IEE, Acadex, Higher, IEO',
        'Housekeeping / Dining Room Attendant',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $40', '46 – 52 ชม./wk', '$105 (รวมกิน)',
        'โซนประวัติศาสตร์ Mammoth และ Lake Hotel ริมทะเลสาบใหญ่ที่สุด, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Lake Yellowstone Hotel Dining Room: Busser 17:30-22:30 น. (ร้าน 4 ดาว ทิปสด $50-$90/คืน)',
        '🥪 Mammoth Terrace Grill: ผู้ช่วยทำแซนด์วิช/เบอร์เกอร์ 16:30-21:30 น. (เรต OT $24.00)',
        '🛒 General Stores Mammoth: เติมสต็อกสินค้ากะค่ำ 17:00-21:00 น.'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Delaware North Yellowstone General Stores (Summer)',
        'OEG, Higher, Acadex',
        'Retail Associate / Food Service / Cashier',
        'ปกติ: 08:30 – 17:00 หรือ 12:00 – 20:30\nพีคซัมเมอร์: กะหมุนเวียน 8-9 ชม.',
        'ฐาน $16.00 / OT $24.00', '-', '42 – 48 ชม./wk', '$100 (รวมกิน)',
        'ร้านค้าสะดวกซื้อและของที่ระลึก 12 จุดทั่วอุทยานเยลโลว์สโตน, อาหาร 3 มื้อ EDR, ส่วนลดสินค้า 30%',
        '🍔 Soda Fountain & Diner (ในสโตร์): ผู้ช่วยทำเบอร์เกอร์/ตักไอศกรีม 17:00-21:30 น. (ขอ OT ได้)',
        '🧹 Night Custodial: ล้างพื้น/ปิดร้านสโตร์ 20:30-23:30 น. (ได้เรต OT $24/ชม.)',
        '🍽️ Xanterra Dining Rooms: สมัครช่วยล้างจาน/Busser กะค่ำในโซนเดียวกัน'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Four Seasons Resort Jackson Hole & Snake River Lodge (Teton Village)',
        'Higher, Acadex, ALC',
        'Housekeeping / Public Area / F&B',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.50 / OT $26.25', '$30 - $70', '42 – 48 ชม./wk', '$140',
        'รีสอร์ตหรูระดับ 5 ดาว AAA 5-Diamond แขกมหาเศรษฐีซัมเมอร์, หอพักพนักงาน Teton Village',
        '🥩 Westbank Grill (ใน Four Seasons): Busser 17:30-23:30 น. (ทิปสดคืนละ $80-$150)',
        '🍺 Mangy Moose Saloon (Teton Village): Barback/Dishwasher 18:00-00:30 น. (ทิปสดแน่นมาก)',
        '🍕 Osteria / Handle Bar: Food Runner ร้านอิตาเลียน 17:00-22:30 น. ($15/ชม. + ทิป)'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'The Wort Hotel & Silver Dollar Bar (Jackson Town)',
        'Higher, Acadex',
        'Housekeeping / Busser / Banquet',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $17.00 / OT $25.50', '$30 - $60', '40 – 46 ชม./wk', '$130',
        'โรงแรมประวัติศาสตร์ใจกลาง Town Square เมืองแจ็กสัน, เดินเท้าได้ทั่วเมือง หางานสองง่ายที่สุดในไวโอมิง',
        '🍸 Silver Dollar Bar & Grill (ในโรงแรม): Barback 18:30-00:30 น. (ทิปสด $70-$120/คืน)',
        '🤠 Million Dollar Cowboy Bar: เดิน 2 นาที Barback/Busser บาร์คาวบอยอันดับ 1 ของโลก (ทิปมหาศาล)',
        '🥩 Gun Barrel Steak & Game House: Busser ร้านสเต๊กสัตว์ป่า 17:30-23:00 น. ($14/ชม. + ทิป $60-$100)'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Mangy Moose Restaurant & Saloon (Teton Village, WY)',
        'Higher, Acadex, ALC',
        'Busser / Food Runner / Dishwasher / Barback',
        'ปกติ: 15:30 – 23:30 (8 ชม.)\nพีคซัมเมอร์: 14:00 – 00:30 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '$40 - $100', '42 – 48 ชม./wk', '$130',
        'บาร์และร้านอาหารชื่อดังระดับตำนานในหมู่บ้าน Teton Village ทิปสดมหาศาลจากนักท่องเที่ยวซัมเมอร์',
        '🧹 กะเช้าช่วยโรงแรม Four Seasons / Snake River Lodge: 08:30-14:00 น. ($17.50/ชม.)',
        '🚡 Jackson Hole Tram Host Helper: ช่วยตรวจตั๋วกระเช้าลอยฟ้ารอบเช้า 08:00-11:30 น.',
        '☕ Teton Village Coffee & Bakery: บาริสต้า/แคชเชียร์ช่วงเช้าตรู่'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Under Canvas Grand Teton / Yellowstone (Summer Only)',
        'Higher Education, IEO, ALC',
        'Housekeeping / Guest Service',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$50 - $100', '42 – 46 ชม./wk', '$100',
        'แคมป์กระโจมซาฟารีหรูเปิดเฉพาะ พ.ค.-ก.ย., อาหารพนักงาน, ทิปเงินสดจากแขกไฮเอนด์',
        '🏕️ Embers Restaurant (ในแกลมปิ้ง): Food Runner กะค่ำ 17:30-22:00 น. (ได้ทิปสดคืนละ $40-$60)',
        '☕ Moose-Wilson Road Cafes: ผู้ช่วยบาริสต้า/เบเกอรี่ช่วงเช้าตรู่ 06:00-08:00 น. หรือเย็น',
        '🛒 Jackson Town Local Diners: นั่งรถเข้าเมืองแจ็กสัน ทำงานร้านอาหารวันหยุด'
    ],
    [
        'Tier S', 'Wyoming (WY)',
        'Togwotee Mountain Lodge (Moran / Dubois)',
        'Acadex, Higher, IEE',
        'Housekeeping / Culinary / Guest Activities',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$15 - $30', '44 – 50 ชม./wk', '$110 (รวมกิน)',
        'รีสอร์ตภูเขาบนช่องเขา Togwotee Pass ใกล้ Grand Teton NP, หอพักพร้อมอาหาร 3 มื้อ',
        '🥩 Red Fox Saloon (ในรีสอร์ต): Busser/Barback 17:30-22:30 น. (ขอ OT $24/ชม. + ทิป)',
        '🍳 Grizzly Grill (ในรีสอร์ต): ผู้ช่วยครัว/ล้างจาน 17:00-22:00 น. (เรต OT $24/ชม.)',
        '🐴 Horseback Riding Barn: ช่วยดูแลม้าและทำความสะอาดคอกม้าช่วงเย็น 16:30-20:00 น.'
    ],

    # 3. WISCONSIN (WI)
    [
        'Tier S', 'Wisconsin (WI)',
        'Kalahari Resorts & Conventions (Wisconsin Dells)',
        'OEG, Acadex, Higher, IEO, New Step, ALC',
        'Lifeguard / Housekeeping (Summer High Season)',
        'ปกติ: 09:30 – 18:00 หรือ 12:00 – 20:30\nพีคซัมเมอร์: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '38 – 44 ชม./wk', '$105',
        'หอพัก Kalahari Village, บัตรเล่นสวนน้ำฟรี, เมืองปั่นจักรยาน 100%, หางาน 2 ร้านอาหารง่าย',
        '🍕 Moosejaw Pizza & Dells Brewing: ปั่นจักรยาน 10 นาที Busser/Barback 18:30-23:30 น. ($13/ชม. + ทิปสด $60-$90/คืน)',
        '🍔 Buffalo Phil’s Grille: Food Runner/Busser เสิร์ฟอาหารด้วยรถไฟจิ๋ว 18:00-22:30 น. ($13/ชม. + ทิปสด $50-$80)',
        '🧀 MACS Mac & Cheese / Dairy Queen: แคชเชียร์/ครัวกะดึก 18:30-23:00 น. ($14.50-$15.50/ชม.)'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Moosejaw Pizza & Dells Brewing Co. (Wisconsin Dells)',
        'OEG, Acadex, Higher, IEO',
        'Food Runner / Busser / Pizza Prep / Host',
        'ปกติ: 11:30 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 11:00 – 23:30 (12 ชม.)',
        'ฐาน $13.50 / OT $20.25', '$40 - $90', '42 – 48 ชม./wk', '$100',
        'ร้านพิซซ่าและโรงเบียร์ยอดนิยมอันดับ 1 ใน Wisconsin Dells ทิปสดดีมาก, อาหารพนักงานฟรี',
        '🧹 กะเช้าช่วยโรงแรม Kalahari / Wilderness: 08:30-14:00 น. ($15/ชม.)',
        '🍦 Dairy Queen Wisconsin Dells: ตักไอศกรีมรอบดึก 21:00-00:00 น. ($14.50/ชม.)',
        '🛒 Outlets at the Dells: พนักงานร้านค้า/จัดสต็อกวันหยุด'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Paul Bunyan’s Cook Shanty & Bakery (Wisconsin Dells)',
        'OEG, Acadex, Higher',
        'Bakery Assistant / Busser / Dishwasher / Food Server',
        'ปกติ: 06:30 – 14:30 (8 ชม.)\nพีคซัมเมอร์: 06:00 – 15:30 (9.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$30 - $70', '42 – 48 ชม./wk', '$95',
        'ร้านอาหารสไตล์แคมป์ตัดไม้ชื่อดังระดับประเทศ เสิร์ฟโดนัทชูการ์และอาหารเช้าไม่อั้น, ทานอาหารเช้าฟรีทุกวัน',
        '🎢 Mt. Olympus Waterpark: กะบ่าย-ค่ำ คุมเครื่องเล่น/สวนสนุก 16:00-22:00 น. ($15/ชม.)',
        '🍕 Pizza Pub Dells: ล้างจาน/ผู้ช่วยทำพิซซ่า 18:00-23:30 น. ($14/ชม. + ทิป)',
        '🍺 Showboat Saloon: Barback ผับดนตรีสดรอบดึก'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Wilderness Resort & Glacier Canyon Lodge (WI Dells)',
        'OEG, IEE, Acadex, IEO',
        'Waterpark Attendant / Housekeeping',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:00 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '40 – 46 ชม./wk', '$105',
        'อาณาจักรรีสอร์ตสวนน้ำที่ใหญ่ที่สุดในอเมริกา (600 เอเคอร์), หอพักพนักงานในรีสอร์ต',
        '🥩 Field’s at the Wilderness (สเต๊กหรูในรีสอร์ต): Busser 17:30-23:00 น. ($12/ชม. + ทิปสด $70-$120/คืน)',
        '🍕 Sarento’s Restaurant (ในรีสอร์ต): Food Runner/ล้างจาน 17:00-22:30 น. (ขอ OT ในรีสอร์ต $22.50)',
        '🍔 B-LUX Grill & Bar: Barback/Busser ร้านเบอร์เกอร์คราฟต์เบียร์ 18:00-00:00 น. (ทิปสดดีมาก)'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Mt. Olympus Water & Theme Park (WI Dells)',
        'New Step, Acadex, Higher, IEO',
        'Ride Operator / Housekeeping / Lifeguard',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 21:00 (11.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$95',
        'สวนสนุกธีมกรีกโบราณ โรลเลอร์โคสเตอร์ไม้, หอพักราคาถูก, ทำเลติดถนนใหญ่ Parkway',
        '🥞 Paul Bunyan’s Cook Shanty: ล้างจาน/ทำความสะอาดรอบค่ำ 17:30-22:00 น. ($14.50/ชม. + กินฟรี)',
        '🌭 Hot Dog Avenue: แคชเชียร์/ครัวปิด 18:00-22:30 น. ($14.50/ชม.)',
        '🍺 Showboat Saloon (Downtown): Barback ผับดนตรีสด 19:00-01:30 น. (ทิปสดแน่น)'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Noah’s Ark Waterpark (Wisconsin Dells - Summer Only)',
        'OEG, IEE, New Step, IEO',
        'Ride Operator / Park Services',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 18:30 (9.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '-', '40 – 46 ชม./wk', '$100',
        'สวนน้ำกลางแจ้งใหญ่ที่สุดในอเมริกา เปิดเฉพาะซัมเมอร์, เล่นสวนน้ำฟรี, หอพักพนักงานใกล้ที่ทำงาน',
        '🍻 Monk’s Bar & Grill (Downtown Dells): ปั่นจักรยานไปทำ Barback/Busser 18:30-00:00 น. (ทิปสดแน่นมาก)',
        '🍕 Pizza Pub Dells: ช่วยส่งอาหาร/ทำพิซซ่า/ล้างจานรอบดึก 18:00-00:00 น. ($14/ชม. + ทิป)',
        '🍦 Dairy Queen / Cold Stone Dells: ตักไอศกรีมกะค่ำ 18:00-22:30 น. ($14.50/ชม.)'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Chula Vista Resort & Waterpark (WI Dells)',
        'IEE, Acadex, Higher, IEO',
        'Housekeeping / Lifeguard / Food Service',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 18:30 (9.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '-', '38 – 44 ชม./wk', '$95',
        'รีสอร์ตริมแม่น้ำ Wisconsin River, สวนน้ำในร่ม/กลางแจ้ง, หอพักในรีสอร์ต',
        '🥩 Kaminski’s Chop House (ในรีสอร์ต): Busser ร้านสเต๊กหรู 17:30-23:00 น. ($12/ชม. + ทิปสดเศรษฐี $70-$120/คืน)',
        '🌮 Mexicali Rose: Busser/Runner ร้านอาหารเม็กซิกันริมน้ำ 18:00-23:00 น. ($13/ชม. + ทิปสด)',
        '🍕 Kilbourn City Grill (ในรีสอร์ต): ล้างจาน/ผู้ช่วยครัว 17:00-22:00 น. (ขอ OT ในรีสอร์ตได้)'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'Grand Geneva Resort & Spa (Lake Geneva, WI)',
        'Acadex, OEG, IEO, ALC',
        'Housekeeping / Banquet / Culinary',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '42 – 48 ชม./wk', '$110',
        'รีสอร์ตหรูระดับ 4 เพชร AAA เมืองตากอากาศทะเลสาบคนรวยชิคาโกช่วงซัมเมอร์, หอพักพนักงาน',
        '🥩 Geneva ChopHouse (ในรีสอร์ต): Busser 17:30-23:00 น. ($13/ชม. + ทิปสด $60-$100/คืน)',
        '🍕 Popeye’s on Lake Geneva: Busser/Dishwasher ร้านอาหารริมทะเลสาบดัง 18:00-23:30 น. (ทิปดี)',
        '🍺 The Bottle Shop / Local Pubs: Barback ร้านคราฟต์เบียร์ 18:30-00:00 น.'
    ],
    [
        'Tier S', 'Wisconsin (WI)',
        'University of Wisconsin-Madison (Housing Facilities & Dining)',
        'OEG, Acadex, Higher',
        'Facilities Crew / Food Service / Dining Associate',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '40 – 45 ชม./wk', '$100',
        'มหาวิทยาลัยชั้นนำใจกลางเมือง Madison, หอพักนักศึกษาพร้อมสิ่งอำนวยความสะดวกครบครัน, รถเมล์ฟรีทั่วเมือง',
        '🍕 Ian’s Pizza on State (สาขาดัง Madison): ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-00:30 น. ($15/ชม. + ทิป)',
        '🍺 State Street Brats: Barback/Busser ร้านเบอร์เกอร์และเบียร์ชื่อดัง 18:30-01:00 น. (ทิปสดดี)',
        '🍜 Asian Sweet Bakery & Local Eateries: แคชเชียร์/ผู้ช่วยครัว 17:30-22:00 น. ($15/ชม.)'
    ],

    # 4. TENNESSEE (TN)
    [
        'Tier S', 'Tennessee (TN)',
        'Dollywood Theme Park & Splash Country (Pigeon Forge)',
        'New Step, Acadex, IEO, OEG',
        'Ride Operator / Food Service / Retail',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 21:30 (12 ชม. สวนสนุกเปิดดึกมีพลุ)',
        'ฐาน $15.00 / OT $22.50', '-', '44 – 50 ชม./wk', '$110',
        'ปลอดภาษีเงินได้รัฐ 0%, นั่งรถราง Pigeon Forge Trolley $1/วัน, บัตรเข้าสวนสนุก Dollywood ฟรี',
        '🍗 Paula Deen’s Family Kitchen (The Island): นั่งรถรางไปทำ Busser 18:30-23:00 น. ($13/ชม. + ทิปสด $60-$90)',
        '🍕 Mellow Mushroom Pizza: Barback/Busser ร้านพิซซ่าคราฟต์เบียร์ 18:30-00:00 น. ($12/ชม. + ทิปสด $50-$80)',
        '🥞 Puckett’s Grocery & Restaurant: Food Runner 18:00-23:00 น. ($13/ชม. + ทิป)'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'The Island in Pigeon Forge (Arcades & Attractions)',
        'New Step, Acadex, IEO',
        'Attractions Host / Retail / F&B',
        'ปกติ: 10:00 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 10:00 – 23:00 (กะดึก)',
        'ฐาน $15.00 / OT $22.50', '$20 - $40', '42 – 48 ชม./wk', '$110',
        'ศูนย์รวมความบันเทิง ชิงช้าสวรรค์ยักษ์และร้านอาหารกลางเมือง, รถรางผ่านหน้าร้าน',
        '🦜 Margaritaville Restaurant (ใน The Island): Busser/Barback 18:30-00:00 น. ($13/ชม. + ทิปสด $60-$100)',
        '🥞 Timberwood Grill (ใน The Island): ล้างจาน/ผู้ช่วยครัว 18:00-23:30 น. ($14.50/ชม.)',
        '🍭 The Island Candy Kitchen: แคชเชียร์/ทำขนม 18:00-22:30 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'Paula Deen’s Family Kitchen (The Island in Pigeon Forge, TN)',
        'New Step, Acadex, Higher',
        'Food Runner / Busser / Line Cook / Host',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 23:00 (12 ชม.)',
        'ฐาน $14.00 / OT $21.00', '$40 - $90', '44 – 50 ชม./wk', '$110',
        'ร้านอาหารครอบครัวชื่อดังที่สุดใน The Island ปลอดภาษีรัฐ 0%, ทิปสดเงินสดแน่นมากทุกวัน',
        '🎡 Dollywood Theme Park: ช่วยคุมเครื่องเล่นรอบบ่าย-ค่ำ 16:00-22:00 น. ($15/ชม.)',
        '🍕 Mellow Mushroom Pizza: Barback รอบดึก 21:00-00:30 น. ($12/ชม. + ทิป)',
        '🍦 The Fudgery (The Island): แคชเชียร์/ทำฟัดจ์รอบค่ำ'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'Wilderness at the Smokies (Sevierville)',
        'OEG, Acadex, Higher, IEO, ALC',
        'Lifeguard / Housekeeping',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$105',
        'สวนน้ำในร่มและกลางแจ้งใหญ่ที่สุดในเทนเนสซี, ปลอดภาษีรัฐ 0%, หอพักพนักงาน Sevierville',
        '🍏 Applewood Farmhouse Restaurant: Busser ร้านอาหารดังคนแน่น 18:00-22:30 น. (ทิปสด $60-$100/คืน)',
        '🥩 Hidden Trail Restaurant (ในรีสอร์ต): ล้างจาน/Busser กะค่ำ 18:00-22:30 น. (ขอ OT ในรีสอร์ต $22.50/ชม.)',
        '🛍️ Tanger Outlets Sevierville: จัดสต็อกสินค้า/ปิดร้าน 18:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'Westgate Smoky Mountain Resort & Waterpark (Gatlinburg)',
        'OEG, Acadex, Higher',
        'Food & Beverage / Lifeguard / Activities Host',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$15 - $35', '42 – 48 ชม./wk', '$115',
        'รีสอร์ตสวนน้ำและสปาขนาดใหญ่บนไหล่เขา Gatlinburg ปลอดภาษีรัฐ 0%, หอพักพนักงาน Westgate',
        '🥩 Southern Comfort Restaurant (ในรีสอร์ต): Busser/Food Runner 17:30-23:00 น. (ขอ OT $24/ชม. + ทิป)',
        '🍺 Roaring Fork Snack Bar (ริมสระ): แคชเชียร์/ผู้ช่วยครัว 16:30-21:30 น. (ขอ OT $24/ชม.)',
        '🥞 Log Cabin Pancake House: ผู้ช่วยครัว/ล้างจานรอบบ่าย-เย็น 17:00-22:00 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'Anakeesta Mountaintop Theme Park (Gatlinburg)',
        'Higher, IEE, IEO',
        'Guest Host / F&B / Retail',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 20:00 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '$15 - $30', '42 – 48 ชม./wk', '$110',
        'สวนสนุกธรรมชาติบนยอดเขา Gatlinburg นั่งกระเช้า Chondola, ปลอดภาษีรัฐ 0%',
        '🍺 Smoky Mountain Brewery (Gatlinburg): Barback/Busser 18:30-00:30 น. ($12/ชม. + ทิปสด $70-$120/คืน)',
        '🥩 Alamo Steakhouse: Busser/Dishwasher ร้านสเต๊กคาวบอย 17:30-23:00 น. ($13/ชม. + ทิป)',
        '🥞 Pancake Pantry (ร้านแพนเค้กอันดับ 1): ช่วยเตรียมวัตถุดิบ/ล้างจานรอบบ่าย-เย็น'
    ],
    [
        'Tier S', 'Tennessee (TN)',
        'Ober Mountain Adventure Park & Ski Area (Gatlinburg)',
        'IEE, New Step, IEO',
        'Guest Services / Food & Beverage / Ride Operator',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '40 – 46 ชม./wk', '$105',
        'สวนสนุกภูเขาและกระเช้าลอยฟ้า Aerial Tramway ชมวิว Great Smoky Mountains, ปลอดภาษีรัฐ 0%',
        '🥩 The Peddler Steakhouse (Gatlinburg): Busser ร้านสเต๊กริมลำธาร 17:30-23:00 น. (ทิปสด $60-$100)',
        '🍕 Slice Pizza Bakery: ทำพิซซ่า/ล้างจานกะค่ำ 18:00-23:30 น. ($14.50/ชม.)',
        '☕ Gatlinburg Coffee & Donut Shop: ผู้ช่วยหน้าร้าน'
    ],

    # 5. MONTANA (MT)
    [
        'Tier S', 'Montana (MT)',
        'Glacier National Park Lodges (Pursuit - West Glacier)',
        'Higher, Acadex, OEG, IEO, ALC',
        'Housekeeping / Kitchen / Tour Host',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '48 – 54 ชม./wk', '$105 (รวมกิน)',
        'อุทยานกลาเชียร์ฝั่งตะวันตก (Apgar & West Glacier), อาหาร 3 มื้อฟรีใน EDR, ปลอดภาษีการค้า Sales Tax 0%',
        '🥩 Belton Chalet Grill Dining: Busser/Steward ร้านอาหารประวัติศาสตร์ 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🛶 Glacier Raft Company: ช่วยล้างแพยาง/อุปกรณ์ล่องแก่ง 17:00-20:30 น. ($16/ชม.)',
        '🥐 West Glacier Bakery & Cafe: บาริสต้า/แคชเชียร์ช่วงเช้าตรู่ 06:00-08:00 น.'
    ],
    [
        'Tier S', 'Montana (MT)',
        'Many Glacier Hotel & Swiftcurrent Motor Inn (Xanterra - East Glacier)',
        'IEE, Higher, Acadex, IEO',
        'Housekeeping / Kitchen Steward / Dining Room',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '50 – 56 ชม./wk', '$105 (รวมกิน)',
        'สวิตเซอร์แลนด์แห่งอเมริกา โรงแรมไม้สไตล์สวิสริมทะเลสาบ Swiftcurrent Lake, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Ptarmigan Dining Room (ใน Many Glacier): Busser 17:30-22:30 น. (ขอ OT ใน Xanterra $24/ชม. + ทิป)',
        '🛶 Many Glacier Boat Tours: ช่วยล้างทำความสะอาดเรือไม้โบราณ 17:00-20:30 น. (เรต OT $24/ชม.)',
        '☕ Heidi’s Snack Shoppe: ผู้ช่วยร้านขนม/แคชเชียร์ 16:30-21:30 น.'
    ],
    [
        'Tier S', 'Montana (MT)',
        'Big Sky Resort & Mountain Village (Big Sky, MT)',
        'Acadex, Higher, ALC, IEO',
        'Housekeeping / Lift Ops / Culinary',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $17.00 / OT $25.50', '-', '42 – 48 ชม./wk', '$135',
        'สกีรีสอร์ตและเมืองตากอากาศฤดูร้อนระดับโลก แขกมหาเศรษฐี, หอพักพนักงาน Big Sky, กิจกรรมเอาท์ดอร์ฟรี',
        '🥩 Riverhouse BBQ (ร้านบาร์บีคิวอันดับ 1): Busser/Dishwasher 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍺 Lone Peak Brewery & Taphouse: Barback 18:00-00:00 น. ($15/ชม. + ทิป)',
        '🛒 Hungry Moose Market & Deli: จัดของ/แคชเชียร์ 17:00-21:30 น. ($16.50/ชม.)'
    ],
    [
        'Tier S', 'Montana (MT)',
        'Under Canvas Glacier (Coram, MT - Summer Season)',
        'Higher Education, IEO',
        'Housekeeping / Camp Host / Dining',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$50 - $100', '42 – 46 ชม./wk', '$100',
        'แกลมปิ้งหรูหรา 5 ดาวริมอุทยาน Glacier NP, ทิปสดแน่นจากนักท่องเที่ยวไฮเอนด์, ปลอดภาษีซื้อของ 0%',
        '🥩 Glacier Grill (Coram): Busser/Steward 17:30-22:30 น. ($14/ชม. + ทิปสด)',
        '🌲 Glacier Outdoor Center: ช่วยดูแลอุปกรณ์แค้มปิ้ง/เดินป่า 16:30-20:30 น.',
        '☕ Stonefly Lounge: Barback กะค่ำ 18:30-23:30 น.'
    ],

    # 6. MAINE (ME)
    [
        'Tier S', 'Maine (ME)',
        'Bar Harbor Grand Hotel & Bluenose Inn (Acadia NP)',
        'Acadex, Higher, OEG, IEO',
        'Housekeeping / Front Desk / Laundry',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '44 – 50 ชม./wk', '$120',
        'เมืองตากอากาศชายทะเลหน้าอุทยาน Acadia NP นักท่องเที่ยวแน่นตลอดซัมเมอร์, รถเมล์ Island Explorer ฟรี',
        '🦞 Stewman’s Lobster Pound (ริมทะเล): Busser/Steward ร้านกุ้งมังกรดัง 17:30-23:30 น. (ทิปสด $70-$120/คืน)',
        '🍔 Geddy’s Down Under: Barback/Runner ร้านเบอร์เกอร์และซีฟู้ดใจกลางเมือง 18:00-00:00 น. (ทิปสดแน่น)',
        '🍦 Ben & Bill’s Chocolate Emporium: ตักไอศกรีมล็อบสเตอร์/แคชเชียร์ 18:00-22:30 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Maine (ME)',
        'Stewman’s Lobster Pound (Bar Harbor Waterfront, ME)',
        'Acadex, Higher, OEG',
        'Busser / Food Runner / Dishwasher / Host',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 11:00 – 23:30 (12 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$50 - $120', '44 – 50 ชม./wk', '$115',
        'ร้านล็อบสเตอร์ริมทะเลชื่อดังที่สุดใน Bar Harbor หน้าอุทยาน Acadia NP ทิปสดมหาศาล, อาหารทะเลทานฟรี',
        '🧹 กะเช้าช่วยแม่บ้าน Bar Harbor Grand / Bluenose Inn: 08:00-12:30 น. ($16.50/ชม.)',
        '🍦 Ben & Bill’s Chocolate: ตักไอศกรีมรอบดึก 21:30-00:00 น. ($15/ชม.)',
        '🚢 Bar Harbor Whale Watch: ช่วยทำความสะอาดเรือดูวาฬรอบเย็น'
    ],
    [
        'Tier S', 'Maine (ME)',
        'Cliff House Maine (Cape Neddick / Ogunquit)',
        'Acadex, Higher, ALC',
        'Housekeeping / Culinary / Steward',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $17.50 / OT $26.25', '$30 - $60', '42 – 48 ชม./wk', '$135',
        'รีสอร์ตหรูระดับ 5 ดาวบนหน้าผาริมมหาสมุทรแอตแลนติก, หอพักพนักงาน, ส่วนลดอาหาร',
        '🦞 The Tiller Restaurant (ในรีสอร์ต): Busser ร้าน Fine Dining ริมผา 17:30-23:30 น. (ทิปสด $80-$140/คืน)',
        '🍔 Nubb’s Lobster Shack (ในรีสอร์ต): ผู้ช่วยครัว/Runner 16:30-22:30 น. (ขอ OT $26.25/ชม.)',
        '🏖️ Ogunquit Beachfront Cafes: ขี่จักรยานไปทำงานร้านอาหารกะค่ำ'
    ],
    [
        'Tier S', 'Maine (ME)',
        'The Nonantum Resort (Kennebunkport, ME)',
        'Higher, Acadex, IEO',
        'Housekeeping / Banquet / Busser',
        'ปกติ: 08:00 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$30 - $70', '42 – 48 ชม./wk', '$120',
        'รีสอร์ตประวัติศาสตร์ริมแม่น้ำ Kennebunk River เมืองพักตากอากาศประธานาธิบดีสหรัฐฯ, ทิปสดแน่น',
        '🦞 Ocean Restaurant (ในรีสอร์ต): Busser ร้านอาหารริมน้ำ 17:30-23:00 น. (ทิปสด $70-$120/คืน)',
        '🥪 The Clam Shack Kennebunkport: ผู้ช่วยครัว/ทอดซีฟู้ด 17:00-22:00 น. ($15/ชม. + ทิป)',
        '🚲 Kennebunkport Bicycle Rentals: ช่วยดูแลจักรยานเช่ารอบเช้า 07:30-10:30 น.'
    ],

    # 7. UTAH (UT)
    [
        'Tier S', 'Utah (UT)',
        'Zion National Park Lodge (Xanterra - Springdale)',
        'IEE, Higher, Acadex, IEO, ALC',
        'Housekeeping / Kitchen Steward / Retail',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '-', '46 – 52 ชม./wk', '$105 (รวมกิน)',
        'โรงแรมเดียวที่ตั้งอยู่ข้างในหุบเขา Zion Canyon อุทยานแห่งชาติไซออน, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Red Rock Grill (ใน Zion Lodge): Busser/Steward 17:30-22:30 น. (ขอ OT ใน Xanterra $24/ชม. + ทิป)',
        '🍺 Bit & Spur Saloon (Springdale): นั่งชัตเติลฟรีออกไปทำ Barback 18:30-00:30 น. (ทิปสด $60-$100/คืน)',
        '🍕 Zion Pizza & Noodle Co.: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:00 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Utah (UT)',
        'Ruby’s Inn & Bryce Canyon Grand Hotel (Best Western)',
        'Acadex, Higher, IEO, ALC',
        'Housekeeping / Fast Food / Grocery',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '48 – 54 ชม./wk', '$90 (รวมกิน)',
        'อาณาจักรรีสอร์ตหน้าอุทยาน Bryce Canyon NP, หอพักพร้อมอาหารบุฟเฟต์ 3 มื้อ $90/wk ประหยัดมาก',
        '🥩 Cowboy’s Buffet & Steak Room: Busser/Runner 17:00-22:00 น. (ขอ OT $23.25/ชม. + ทิป)',
        '🤠 Ebenezer’s Barn & Grill: ช่วยจัดเลี้ยงอาหารคาวบอยพร้อมโชว์ดนตรีสด 17:30-22:30 น. (ทิปสดดี)',
        '🛒 Ruby’s General Store: แคชเชียร์/เติมสต็อกกะค่ำ 18:00-22:30 น. (เรต OT $23.25)'
    ],
    [
        'Tier S', 'Utah (UT)',
        'Park City Mountain Resort (Vail Resorts - Summer Season)',
        'OEG, Higher, Acadex, ALC',
        'Mountain Host / Lift Operator / F&B',
        'ปกติ: 08:30 – 17:00 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 18:00 (9.5 ชม.)',
        'ฐาน $17.00 / OT $25.50', '$20 - $40', '40 – 46 ชม./wk', '$135',
        'สกีรีสอร์ตอันดับ 1 ของยูทาห์ช่วงซัมเมอร์ กิจกรรม Zip Line & Alpine Slide, รถบัสฟรีทั่วเมือง Park City',
        '🥩 High West Distillery & Saloon: Busser/Barback ร้านวิสกี้ชื่อดัง 17:30-23:30 น. (ทิปสด $80-$140/คืน)',
        '🍕 Red Banjo Pizza (Main Street): ล้างจาน/ทำพิซซ่า 18:00-00:00 น. ($15/ชม. + ทิป)',
        '☕ Park City Coffee Roaster: บาริสต้าช่วงเช้าตรู่ 06:30-09:30 น.'
    ],
    [
        'Tier S', 'Utah (UT)',
        'Red Cliffs Lodge & Winery (Moab / Arches National Park)',
        'Higher, Acadex, IEO',
        'Housekeeping / Winery Host / Culinary',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$30 - $60', '44 – 50 ชม./wk', '$115',
        'รีสอร์ตหรูริมแม่น้ำ Colorado River หน้าอุทยาน Arches & Canyonlands, อาหารพนักงาน, ทิปสดดี',
        '🥩 The Cowboy Grill (ในรีสอร์ต): Busser/Steward ร้านสเต๊กริมแม่น้ำ 17:30-23:00 น. (ขอ OT $24.75 + ทิป)',
        '🛶 Moab Adventure Center: ช่วยล้างเรือยาง/อุปกรณ์ล่องแก่ง 17:00-20:30 น. ($16/ชม.)',
        '🍕 Moab Brewery: Barback/Busser ร้านคราฟต์เบียร์ใจกลางเมือง 18:30-00:00 น. (ทิปสดแน่น)'
    ],

    # 8. MARYLAND (MD)
    [
        'Tier S', 'Maryland (MD)',
        'Premier Aquatics / High Sierra Pools (Lifeguard MD/DC/VA)',
        'American Learning (ALC), Acadex, OEG',
        'Pool Lifeguard (Summer Guaranteed Hours)',
        'ปกติ: 10:30 – 20:00 (9 ชม.)\nพีคซัมเมอร์: 10:00 – 20:30 (10 ชม. + ควงกะวันหยุด)',
        'ฐาน $16.00 / OT $24.00', '-', '52 – 60 ชม./wk', '$140',
        'การันตีชั่วโมงทำงานและ OT สูงสุดในสายไลฟ์การ์ด, มีฝึกอบรมและสอบใบรับรอง Red Cross ให้, สระว่ายน้ำคอนโดหรู',
        '🍕 Domino’s / Pizza Hut Delivery Helper: ช่วยส่งพิซซ่า/ล้างจานรอบดึก 20:30-00:30 น. ($15/ชม. + ทิป)',
        '🛒 Giant Food / Safeway Supermarket: จัดสต็อกสินค้ากะดึก 21:00-01:00 น. ($16/ชม.)',
        '🏊 Swim Instructor: สอนว่ายน้ำเด็กช่วงเช้า 08:30-10:30 น. (ได้เรตพิเศษ $25-$35/ชม.)'
    ],
    [
        'Tier S', 'Maryland (MD)',
        'Ocean City Boardwalk Hotels & Oceanfront Resorts',
        'New Step, Acadex, OEG, IEO',
        'Housekeeping / Front Desk / Laundry',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$15 - $30', '40 – 46 ชม./wk', '$125',
        'เมืองตากอากาศชายหาดอันดับ 1 ในแมริแลนด์ บอร์ดวอล์กคึกคักตลอด 24 ชม., รถเมล์ Coastal Highway $3/วัน',
        '🦀 Seacrets Jamaica USA (ผับริมหาดใหญ่ที่สุดในอเมริกา): Busser/Barback 18:00-01:30 น. (ทิปสด $80-$150/คืน)',
        '🍟 Thrasher’s French Fries (Boardwalk): พนักงานทอดเฟรนช์ฟรายส์/แคชเชียร์ 17:30-23:00 น. ($15/ชม.)',
        '🍕 The Dough Roller Pizza: ล้างจาน/บัสเซอร์กะค่ำ 18:00-00:00 น. ($14.50/ชม. + ทิป)'
    ],
    [
        'Tier S', 'Maryland (MD)',
        'Seacrets Jamaica USA (Ocean City, MD)',
        'New Step, Acadex, Higher',
        'Barback / Busser / Food Runner / Dishwasher',
        'ปกติ: 15:30 – 00:00 (8 ชม.)\nพีคซัมเมอร์: 14:00 – 02:00 (11.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$60 - $150', '44 – 52 ชม./wk', '$125',
        'บาร์และร้านอาหารริมหาดสไตล์จาไมกาที่ใหญ่ที่สุดในสหรัฐฯ ทิปสดแน่นที่สุดใน Ocean City, ดนตรีสดทั้งคืน',
        '🧹 กะเช้าช่วยแม่บ้านโรงแรมริมหาด (Clarion / Carousel): 08:30-14:00 น. ($15.50/ชม.)',
        '🍟 Boardwalk Fries / Dumser’s Dairyland: แคชเชียร์ช่วงบ่าย 12:00-16:00 น.',
        '🛒 Ocean City Beach Rentals: ดูแลร่ม/เก้าอี้ชายหาดช่วงเช้า'
    ],
    [
        'Tier S', 'Maryland (MD)',
        'Jolly Roger Amusement Park & Splash Mountain (Ocean City, MD)',
        'New Step, IEE, Acadex, IEO',
        'Ride Operator / Lifeguard / Games & Retail',
        'ปกติ: 12:00 – 20:30 (8 ชม.)\nพีคซัมเมอร์: 10:00 – 23:00 (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.00 / OT $22.50', '-', '44 – 50 ชม./wk', '$120',
        'สวนสนุกและสวนน้ำที่ใหญ่ที่สุดใน Ocean City, เล่นเครื่องเล่นฟรี, หอพักพนักงานใกล้ 30th Street',
        '🦀 Higgins Crab House: Busser/Runner ร้านปูย่างดัง 18:00-23:30 น. (ทิปสด $60-$100/คืน)',
        '🍕 Piezano’s Pizza: ผู้ช่วยทำพิซซ่า/ล้างจานรอบดึก 21:00-01:30 น. ($14.50/ชม.)',
        '🍿 Fisher’s Popcorn (Boardwalk): แคชเชียร์/บรรจุป๊อปคอร์นคาราเมล'
    ],

    # 9. OHIO (OH)
    [
        'Tier S', 'Ohio (OH)',
        'Cedar Point Amusement Park & Hotel Breakers (Sandusky)',
        'OEG, IEE, Higher, IEO, ALC',
        'Ride Operator / Housekeeping / F&B',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:30 (12.5 ชม. ควงกะ)',
        'ฐาน $15.00 / OT $22.50', '-', '48 – 55 ชม./wk', '$85',
        'เมืองหลวงรถไฟเหาะของโลกริมทะเลสาบ Lake Erie, หอพักพนักงานราคาถูกมาก $85/wk, อาหารโรงอาหารลด 50%, เล่นเครื่องเล่นฟรี',
        '🥩 Famous Dave’s BBQ (ริมท่าเรือ Cedar Point Marina): Busser 18:00-23:30 น. ($13/ชม. + ทิปสด $50-$90/คืน)',
        '🍕 Hotel Breakers Pizzeria (ในสวนสนุก): ผู้ช่วยครัว/ล้างจาน 18:00-00:00 น. (ขอ OT ใน Cedar Fair ได้ $22.50)',
        '🍻 Thirsty Pony / Castaway Bay: นั่งรถบัสพนักงานเข้าเมืองทำร้านอาหารกะดึก'
    ],
    [
        'Tier S', 'Ohio (OH)',
        'Kings Island Theme Park & Soak City (Mason / Cincinnati)',
        'OEG, Higher, Acadex, IEO',
        'Ride Operator / Food Service / Park Services',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:00 (12 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '44 – 50 ชม./wk', '$90',
        'สวนสนุกและสวนน้ำชั้นนำเครือ Cedar Fair, หอพักพนักงาน Kings Island, บัตรเข้าสวนสนุกฟรี',
        '🥩 Great Wolf Lodge Mason (ติด Kings Island): Busser/Steward 17:30-23:00 น. ($14/ชม. + ทิป)',
        '🍦 Graeter’s Ice Cream Mason: ตักไอศกรีม/แคชเชียร์ 18:00-22:30 น. ($14.50/ชม.)',
        '🍔 Skyline Chili Mason: ผู้ช่วยครัว/ล้างจาน 18:00-23:00 น. ($14.50/ชม.)'
    ],
    [
        'Tier S', 'Ohio (OH)',
        'Kalahari Resorts & Conventions (Sandusky, OH)',
        'OEG, Acadex, Higher, ALC',
        'Lifeguard / Housekeeping / Culinary',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 20:00 (10.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$95',
        'รีสอร์ตสวนน้ำในร่มขนาดใหญ่ที่สุดในโอไฮโอ, หอพักพนักงาน, เล่นสวนน้ำฟรี',
        '🥩 B-Lux Grill & Bar (ในรีสอร์ต): Busser/Barback 17:30-23:30 น. (ขอ OT $22.50/ชม. + ทิปสด)',
        '🎢 Cedar Point (ฝั่งตรงข้าม): ขอกะคุมเครื่องเล่นรอบบ่าย-ค่ำ 16:00-22:30 น.',
        '🍕 Chet & Matt’s Pizza Sandusky: ล้างจาน/ทำพิซซ่ารอบค่ำ ($14/ชม.)'
    ],

    # 10. COLORADO (CO)
    [
        'Tier S', 'Colorado (CO)',
        'YMCA of the Rockies (Estes Park Center - RMNP)',
        'Higher Education, Acadex, IEO',
        'Housekeeping / Food Service / Camp Activities',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '42 – 48 ชม./wk', '$95 (รวมกิน)',
        'ศูนย์รีสอร์ตธรรมชาติหน้าอุทยาน Rocky Mountain NP, หอพักพร้อมอาหาร 3 มื้อบุฟเฟต์ ($95/wk), กิจกรรมปีนเขา/พายเรือฟรี',
        '🥩 The Stanley Hotel (โรงแรมประวัติศาสตร์ The Shining): Busser/Runner 17:30-23:00 น. (ทิปสด $60-$110/คืน)',
        '🍺 Estes Park Brewery: Barback/Dishwasher ร้านเบียร์และพิซซ่า 18:00-23:30 น. ($14/ชม. + ทิป)',
        '🍔 Penelope’s Old Time Burgers: แคชเชียร์/ครัวกะค่ำ 17:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier S', 'Colorado (CO)',
        'The Stanley Hotel & Cascades Restaurant (Estes Park, CO)',
        'Higher, Acadex, ALC',
        'Housekeeping / Busser / Food Runner / Front Desk',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$30 - $70', '42 – 48 ชม./wk', '$125',
        'โรงแรมประวัติศาสตร์ชื่อดังระดับโลก (แรงบันดาลใจหนัง The Shining) วิวเทือกเขา Rocky Mountain, ทิปสดแน่น',
        '🥩 Cascades Restaurant (ในโรงแรม): Busser 17:30-23:30 น. (ทิปสดเศรษฐี $70-$130/คืน)',
        '🍺 Estes Park Brewery: Barback 18:00-00:00 น. ($14.50/ชม. + ทิป)',
        '🥐 Cinnamon’s Bakery: ช่วยเตรียมอบขนมช่วงเช้าตรู่ 06:00-09:00 น.'
    ],

    # 11. SOUTH CAROLINA (SC)
    [
        'Tier S', 'South Carolina (SC)',
        'Myrtle Beach Oceanfront Resorts & Hotels (Summer High)',
        'New Step, Acadex, Higher, IEO, ALC',
        'Housekeeping / Pool Attendant / Front Desk',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$15 - $30', '40 – 46 ชม./wk', '$115',
        'เมืองตากอากาศชายหาดยอดนิยมอันดับ 1 ใน South Carolina, บอร์ดวอล์กและสวนสนุกริมหาด, หางาน 2 ง่ายที่สุด',
        '🦀 Sea Captain’s House (ร้านซีฟู้ดริมทะเลอันดับ 1): Busser/Steward 17:30-23:00 น. (ทิปสด $70-$120/คืน)',
        '🎡 Broadway at the Beach: พนักงานร้านอาหาร/บาร์/ของที่ระลึก 18:00-00:00 น. ($14/ชม. + ทิป)',
        '🍔 Peaches Corner (ริม Boardwalk): แคชเชียร์/ทำเบอร์เกอร์กะดึก 18:30-01:00 น. ($14/ชม.)'
    ],
    [
        'Tier S', 'South Carolina (SC)',
        'Sea Captain’s House Restaurant (Myrtle Beach Oceanfront, SC)',
        'New Step, Acadex, Higher',
        'Busser / Food Runner / Dishwasher / Host',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 23:00 (12 ชม.)',
        'ฐาน $13.50 / OT $20.25', '$50 - $110', '42 – 48 ชม./wk', '$110',
        'ร้านอาหารซีฟู้ดระดับตำนานริมหาด Myrtle Beach คนแน่นตลอดทั้งวัน ทิปสดเงินสดสูงมาก, ทานอาหารฟรี',
        '🧹 กะเช้าช่วยแม่บ้านรีสอร์ตริมหาด (Breakers / Caribbean): 08:00-13:00 น. ($14.50/ชม.)',
        '🎡 Broadway at the Beach: ช่วยร้านขายของที่ระลึกรอบดึก 21:30-00:30 น.',
        '🍦 Mad Myrtle’s Ice Cream: ตักไอศกรีมริมบอร์ดวอล์ก'
    ],

    # 12. NORTH CAROLINA (NC)
    [
        'Tier S', 'North Carolina (NC)',
        'Outer Banks Coastal Resorts & Hotels (Nags Head / Kill Devil Hills)',
        'Acadex, Higher, OEG, IEO',
        'Housekeeping / Beach Attendant / F&B',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$20 - $40', '42 – 48 ชม./wk', '$120',
        'เกาะตากอากาศ Outer Banks (OBX) ชายหาดสวยงามและบ้านพักตากอากาศหรู, ปั่นจักรยานสะดวก',
        '🦀 Owens’ Restaurant (ร้านซีฟู้ดเก่าแก่ OBX): Busser 17:30-23:00 น. (ทิปสด $60-$110/คืน)',
        '🍺 Outer Banks Brewing Station: Barback/Dishwasher โรงเบียร์กังหันลม 18:00-00:00 น. (ทิปสดแน่น)',
        '🍩 Duck Donuts (สาขาต้นตำรับ Duck): ทำโดนัท/แคชเชียร์ 06:30-10:30 น. ($14.50/ชม.)'
    ],
    [
        'Tier S', 'North Carolina (NC)',
        'Biltmore Estate & Village (Asheville, NC)',
        'Higher, Acadex, ALC',
        'Hospitality Associate / Food Service / Retail',
        'ปกติ: 08:30 – 17:00 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '40 – 46 ชม./wk', '$125',
        'คฤหาสน์ส่วนบุคคลที่ใหญ่ที่สุดในอเมริกา เมืองแห่งศิลปะและเบียร์คราฟต์ Asheville ทิวเขา Blue Ridge',
        '🥩 Stable Cafe & Bistro (ในคฤหาสน์): Busser/Runner 17:00-22:00 น. (เรต OT $24.75 + ทิป)',
        '🍺 Wicked Weed Brewing (Downtown Asheville): Barback ผับคราฟต์เบียร์ระดับโลก 18:30-00:30 น. (ทิปสดแน่น)',
        '🥐 French Broad Chocolates: แคชเชียร์/ผู้ช่วยทำขนมรอบค่ำ 18:00-22:30 น.'
    ],

    # =========================================================================
    # 🏅 TIER A STATES (19 States)
    # =========================================================================

    # 13. TEXAS (TX)
    [
        'Tier A', 'Texas (TX)',
        'Schlitterbahn Waterpark (New Braunfels / Galveston, TX)',
        'OEG, Acadex, Higher, IEO',
        'Lifeguard / Ride Attendant / Park Services',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 19:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$100',
        'ปลอดภาษีเงินได้รัฐ 0%, สวนน้ำแม่น้ำธรรมชาติอันดับ 1 ของโลก, หอพักพนักงาน, เล่นสวนน้ำฟรี',
        '🥩 Krause’s Cafe & Biergarten (New Braunfels): Busser 17:30-23:00 น. ($13/ชม. + ทิปสด $60-$90/คืน)',
        '🍕 Gristmill River Restaurant (Gruene Historic District): Busser/Steward ริมแม่น้ำกัวดาลูป (ทิปสดแน่น)',
        '🍔 Buc-ee’s New Braunfels (ปั๊มใหญ่ที่สุดในโลก): แคชเชียร์/จัดของกะดึก ($16-$18/ชม.)'
    ],
    [
        'Tier A', 'Texas (TX)',
        'Kalahari Resorts & Conventions (Round Rock / Austin, TX)',
        'OEG, Acadex, Higher, ALC',
        'Lifeguard / Housekeeping / Culinary',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '40 – 46 ชม./wk', '$110',
        'สวนน้ำในร่มขนาดใหญ่ที่สุดในอเมริกา ปลอดภาษีรัฐ 0%, ชานเมือง Austin เมืองเทคโนโลยีและดนตรีสด',
        '🥩 Double Cut Steak House (ในรีสอร์ต): Busser 17:30-23:00 น. ($13/ชม. + ทิปสด $70-$120/คืน)',
        '🍕 Cinco Niños Mexican (ในรีสอร์ต): Food Runner/ล้างจาน 17:00-22:30 น. (ขอ OT $23.25/ชม.)',
        '🛒 Round Rock Premium Outlets: จัดสต็อกสินค้า/ปิดร้าน 18:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Texas (TX)',
        'Six Flags Fiesta Texas (San Antonio, TX)',
        'New Step, IEE, Acadex',
        'Ride Operator / Games / Food Service',
        'ปกติ: 10:00 – 18:30 (8 ชม.)\nพีคซัมเมอร์: 09:30 – 22:00 (12 ชม.)',
        'ฐาน $14.50 / OT $21.75', '-', '42 – 48 ชม./wk', '$105',
        'สวนสนุกในเหมืองหินปูนธรรมชาติ ปลอดภาษีรัฐ 0%, บัตรเล่นเครื่องเล่น Six Flags ฟรีทั่วประเทศ',
        '🌮 San Antonio River Walk Restaurants: นั่งรถเมล์ไปทำ Busser/Runner 18:30-00:30 น. (ทิปสด $60-$100)',
        '🥩 The Rustic San Antonio: Barback/Busser ร้านอาหารและลานดนตรีสด 18:00-00:00 น.',
        '🛒 The Shops at La Cantera: พนักงานร้านค้า/ปิดร้าน 18:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Texas (TX)',
        'Kemah Boardwalk Resort & Theme Park (Kemah / Houston, TX)',
        'Higher, Acadex, IEO',
        'Attractions Host / Retail / Food Service',
        'ปกติ: 11:00 – 19:00 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 22:30 (11.5 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$20 - $40', '40 – 46 ชม./wk', '$105',
        'ศูนย์รวมความบันเทิงริมอ่าว Galveston Bay ปลอดภาษีรัฐ 0%, ร้านอาหารซีฟู้ดและเครื่องเล่นริมทะเล',
        '🦀 Aquarium Restaurant Kemah: Busser/Runner 17:30-23:00 น. ($12/ชม. + ทิปสด $60-$100/คืน)',
        '🥩 Saltgrass Steak House: ล้างจาน/ผู้ช่วยครัว 17:00-23:00 น. ($14.50/ชม.)',
        '🍦 Boardwalk Sweet Shoppe: แคชเชียร์/ทำขนม 18:00-22:30 น.'
    ],

    # 14. FLORIDA (FL)
    [
        'Tier A', 'Florida (FL)',
        'Universal Orlando Resort & Volcano Bay (Orlando, FL)',
        'IEO Abroad, Acadex, Higher',
        'Attractions Host / Lifeguard / Food Service',
        'ปกติ: 08:30 – 17:00 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 22:30 (กะดึก/ควงกะ)',
        'ฐาน $15.50 / OT $23.25', '-', '38 – 44 ชม./wk', '$135',
        'ปลอดภาษีเงินได้รัฐ 0%, ทำงานในสวนสนุกระดับโลก ยูนิเวอร์แซล สตูดิโอส์ & สวนน้ำภูเขาไฟ Volcano Bay, บัตรเข้าสวนสนุกฟรี',
        '🍔 Universal CityWalk Dining: Food Runner/Busser ร้านอาหารริมทางเดิน 18:00-01:00 น. (ทิปสด $60-$100/คืน)',
        '🍕 Bubba Gump Shrimp Co. / Hard Rock Cafe: Busser 18:30-00:30 น. (ทิปสดแน่น)',
        '🛍️ Orlando International Premium Outlets: จัดสต็อกสินค้า/แคชเชียร์ 18:00-22:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Florida (FL)',
        'Busch Gardens & Adventure Island (Tampa Bay, FL)',
        'New Step, Acadex, IEO',
        'Ride Operator / Park Services / Lifeguard',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 21:30 (12 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '40 – 46 ชม./wk', '$120',
        'สวนสนุกธีมแอฟริกาและสวนน้ำ Adventure Island ปลอดภาษีรัฐ 0%, บัตรเข้าสวนสนุกและสวนน้ำฟรี',
        '🥩 Columbia Restaurant (Ybor City): Busser ร้านอาหารสเปนเก่าแก่ที่สุดในฟลอริดา 18:00-23:30 น. (ทิปสด $70-$120)',
        '🍺 Tampa Bay Brewing Co.: Barback/Runner 18:30-00:00 น. ($14/ชม. + ทิป)',
        '🍔 Portillo’s Tampa: แคชเชียร์/ครัวกะดึก 18:00-23:00 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Florida (FL)',
        'Gaylord Palms Resort & Convention Center (Kissimmee, FL)',
        'OEG, Higher, Acadex',
        'Housekeeping / Culinary / Banquet / Lifeguard',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '40 – 46 ชม./wk', '$125',
        'รีสอร์ตในเครือ Marriott ขนาดใหญ่ อะเทรียมกระจกปรับอากาศ 4 เอเคอร์ ปลอดภาษีรัฐ 0%',
        '🥩 MOOR Seafood Restaurant (ในรีสอร์ต): Busser 17:30-23:00 น. ($13/ชม. + ทิปสด $60-$100/คืน)',
        '🍔 Old Town Kissimmee Eateries: Food Runner ร้านอาหารย่านเมืองเก่า 18:00-00:00 น. ($14/ชม. + ทิป)',
        '🛒 Disney Springs (ใกล้รีสอร์ต): พนักงานร้านค้า/บริการ 18:30-23:30 น. ($15.50/ชม.)'
    ],

    # 15. NEW JERSEY (NJ)
    [
        'Tier A', 'New Jersey (NJ)',
        'Morey’s Piers & Beachfront Waterparks (Wildwood, NJ)',
        'IEE, New Step, IEO, Acadex, Higher',
        'Ride Operator / Lifeguard / Food Service',
        'ปกติ: 11:00 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 10:30 – 23:30 (12.5 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.50 / OT $23.25', '-', '46 – 52 ชม./wk', '$115',
        'สวนสนุกและสวนน้ำบนท่าเรือริมหาดยาว 6 บล็อก เมืองตากอากาศ Wildwood ปั่นจักรยานเลียบหาดได้ 100%',
        '🍕 Jumbo Slice Pizza (ริมบอร์ดวอล์ก): พนักงานทำพิซซ่า/ล้างจานรอบดึก 19:00-01:30 น. ($15/ชม. + ทิป)',
        '🦀 The Wharf Restaurant: Busser/Runner ร้านซีฟู้ดริมน้ำ 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍦 Kohr Bros Frozen Custard: ตักไอศกรีมกะค่ำ 18:00-23:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'New Jersey (NJ)',
        'Six Flags Great Adventure & Safari (Jackson, NJ)',
        'New Step, IEE, Acadex',
        'Ride Operator / Safari Guide / Food Service',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:00 (12 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '42 – 48 ชม./wk', '$110',
        'สวนสนุก Six Flags ที่ใหญ่ที่สุดอันดับ 2 ของโลก รถไฟเหาะ Kingda Ka, หอพักพนักงาน, บัตรเข้าสวนสนุกฟรี',
        '🍕 Jackson Diner & Pizzeria: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น. ($15/ชม.)',
        '🛍️ Jackson Premium Outlets: จัดสต็อกสินค้า/แคชเชียร์ 18:00-21:30 น. ($15.50/ชม.)',
        '🍦 Rita’s Italian Ice: พนักงานหน้าร้านกะค่ำ 17:30-22:00 น.'
    ],

    # 16. SOUTH DAKOTA (SD)
    [
        'Tier A', 'South Dakota (SD)',
        'Xanterra Mount Rushmore National Memorial (Keystone)',
        'IEE, Higher, Acadex, IEO',
        'Retail Associate / Food Service / Custodial',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '44 – 50 ชม./wk', '$95 (รวมกิน)',
        'อนุสรณ์สถานแห่งชาติภูผา 4 ประธานาธิบดี ปลอดภาษีรัฐ 0%, หอพักพร้อมอาหาร 3 มื้อ EDR ($95/wk)',
        '🥩 Carver’s Cafe (ใน Mount Rushmore): Busser/Steward 17:00-22:00 น. (ขอ OT $23.25/ชม.)',
        '🍦 Thomas Jefferson Ice Cream Shop: ตักไอศกรีมสูตรดั้งเดิม 16:30-21:30 น. (เรต OT $23.25)',
        '🤠 Keystone Boardwalk Eateries: เดินเข้าเมืองเคย์สโตนทำร้านสเต๊กคาวบอย 18:00-23:00 น. (ทิปสด)'
    ],
    [
        'Tier A', 'South Dakota (SD)',
        'Custer State Park Resorts & Game Lodge (Custer, SD)',
        'Higher, Acadex, IEO',
        'Housekeeping / Culinary / Activities Host',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$20 - $40', '44 – 50 ชม./wk', '$95 (รวมกิน)',
        'อุทยานฝูงควายไบซันป่า Black Hills ปลอดภาษีรัฐ 0%, หอพักพร้อมอาหาร 3 มื้อ EDR ($95/wk)',
        '🥩 State Game Lodge Restaurant: Busser 17:30-22:30 น. (ร้านอาหารทำเนียบขาวฤดูร้อน ทิปสด $60-$100)',
        '🦬 Buffalo Safari Jeep Tours: ช่วยดูแลและล้างรถจี๊ปทัวร์ 17:00-20:30 น. ($16/ชม.)',
        '🛶 Sylvan Lake Marina: ดูแลเรือพายเช่าริมทะเลสาบ 16:30-20:00 น.'
    ],
    [
        'Tier A', 'South Dakota (SD)',
        'Wall Drug Store & Western Art Gallery (Wall, SD)',
        'Higher, Acadex',
        'Retail Associate / Food Service / Donut Maker',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:30 (10 ชม.)',
        'ฐาน $15.00 / OT $22.50', '$10 - $25', '44 – 50 ชม./wk', '$85 (รวมกิน)',
        'จุดเช็คอินระดับตำนานของอเมริกาหน้าอุทยาน Badlands NP ปลอดภาษีรัฐ 0%, หอพักและอาหารพนักงานราคาถูกมาก',
        '🍩 Wall Drug Famous Donut Shop: ช่วยปั้นโดนัท/แคชเชียร์ 06:00-10:30 น. (เรต OT)',
        '🥩 Red Rock Restaurant (Wall): Busser/Steward ร้านสเต๊กคาวบอย 17:30-22:30 น. (ทิปสด)',
        '🛒 Badlands Trading Post: แคชเชียร์รอบค่ำ 17:00-21:30 น.'
    ],

    # 17. VIRGINIA (VA)
    [
        'Tier A', 'Virginia (VA)',
        'Busch Gardens Williamsburg & Water Country USA (VA)',
        'New Step, Acadex, Higher, IEO',
        'Ride Operator / Lifeguard / Culinary / Retail',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:00 (12 ชม. สวนสนุกเปิดดึก)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$115',
        'สวนสนุกที่สวยที่สุดในโลกธีมยุโรปโบราณ และสวนน้ำ Water Country USA, หอพักพนักงาน, รถรับส่งฟรี',
        '🥩 Das Festhaus (เบียร์ฮอลล์เยอรมันในสวนสนุก): Busser/Runner 17:30-22:30 น. (ขอ OT $22.50 + ทิป)',
        '🍗 Pierce’s Pitt Bar-B-Que (Williamsburg): ผู้ช่วยครัว/ล้างจาน 18:00-23:00 น. ($14.50/ชม. + กินฟรี)',
        '🛍️ Williamsburg Premium Outlets: พนักงานร้านค้า/จัดสต็อก 18:00-21:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Virginia (VA)',
        'Virginia Beach Oceanfront Resorts & Boardwalk Hotels (VA)',
        'New Step, Acadex, Higher',
        'Housekeeping / Front Desk / Beach Services',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '$15 - $30', '40 – 46 ชม./wk', '$120',
        'เมืองตากอากาศชายหาดมหาสมุทรแอตแลนติก บอร์ดวอล์กยาว 3 ไมล์, รถบัส The Wave Trolley $2/วัน',
        '🦀 Waterman’s Surfside Grille: Busser/Barback ร้านซีฟู้ดริมหาดดัง 18:00-00:00 น. (ทิปสด $70-$120/คืน)',
        '🍔 Catch 31 Fishhouse & Bar: Food Runner ริมบอร์ดวอล์ก 17:30-23:30 น. (ทิปสดแน่น)',
        '🍦 Dairy Queen / Kohr Bros (Boardwalk): พนักงานหน้าร้านกะค่ำ'
    ],

    # 18. MASSACHUSETTS (MA)
    [
        'Tier A', 'Massachusetts (MA)',
        'Cape Cod Oceanfront Resorts & Inns (Hyannis / Yarmouth, MA)',
        'Acadex, Higher, OEG, IEO',
        'Housekeeping / Pool Attendant / F&B',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $50', '40 – 46 ชม./wk', '$135',
        'แหลมเคปคอด เมืองตากอากาศชั้นสูงของชาวบอสตันและนิวยอร์ก ทะเลสวย ทิปสดแน่น',
        '🦞 The Lobster Pot (Provincetown/Hyannis): Busser/Steward ร้านล็อบสเตอร์ดัง 17:30-23:30 น. (ทิปสด $70-$130)',
        '🍔 Tugboats on Hyannis Marina: Barback/Runner ร้านอาหารริมท่าเรือยอชต์ 18:00-00:00 น. (ทิปดีมาก)',
        '🍦 Cape Cod Creamery: ตักไอศกรีม/แคชเชียร์ 18:00-22:30 น. ($15.50/ชม.)'
    ],
    [
        'Tier A', 'Massachusetts (MA)',
        'Six Flags New England (Agawam / Springfield, MA)',
        'New Step, Acadex, IEO',
        'Ride Operator / Food Service / Games',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 21:30 (12 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '42 – 48 ชม./wk', '$110',
        'สวนสนุก Six Flags เก่าแก่ที่สุดในอเมริกา รถไฟเหาะ Superman the Ride, บัตรเล่นสวนสนุกฟรี',
        '🍕 Agawam Pizzeria & Diner: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:00 น. ($15/ชม.)',
        '🍔 Five Guys / Wendy’s Agawam: แคชเชียร์/ครัวปิด 18:30-23:30 น. ($15.50/ชม.)',
        '🎳 Riverside Bowling & Arcade: พนักงานรอบค่ำ'
    ],

    # 19. NEW HAMPSHIRE (NH)
    [
        'Tier A', 'New Hampshire (NH)',
        'Omni Mount Washington Resort (Bretton Woods, NH)',
        'Higher, Acadex, ALC',
        'Housekeeping / Culinary / Banquet / Steward',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '44 – 50 ชม./wk', '$110 (รวมกิน)',
        'ปลอดภาษีเงินได้รัฐ 0% และ ปลอดภาษีซื้อของ 0% (Double Zero Tax), รีสอร์ตปราสาทประวัติศาสตร์ 4 เพชร AAA, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Main Dining Room (ใน Omni Resort): Busser ร้านหรู 17:30-23:00 น. (ทิปสด $60-$110/คืน)',
        '🍺 The Cave (บาร์ประวัติศาสตร์ยุคห้ามขายเหล้าในโรงแรม): Barback 19:00-01:00 น. (ทิปสดดีมาก)',
        '🚂 Mount Washington Cog Railway: ช่วยทำความสะอาดสถานีรถไฟไต่เขา 16:30-20:30 น. ($16/ชม.)'
    ],
    [
        'Tier A', 'New Hampshire (NH)',
        'Mount Washington Cog Railway & Base Lodge (NH)',
        'Higher, Acadex',
        'Railway Host / Retail / Food Service',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$15 - $30', '42 – 48 ชม./wk', '$105 (รวมกิน)',
        'รถไฟไต่เขาพลังไอน้ำสายแรกของโลกขึ้นสู่ยอดเขา Mount Washington ปลอดภาษี 0%, หอพักพร้อมอาหาร',
        '🥩 Omni Mount Washington Resort: เดินทาง 10 นาที ช่วยงานกะค่ำในครัว/จัดเลี้ยง ($16.50/ชม.)',
        '🍔 Base Station Grill: ผู้ช่วยครัว/ทำเบอร์เกอร์ 16:00-20:30 น. (ขอ OT ได้)',
        '🛒 Cog Gift Shop: แคชเชียร์รอบปิดร้าน'
    ],

    # 20. MISSOURI (MO)
    [
        'Tier A', 'Missouri (MO)',
        'Silver Dollar City Theme Park & Showboat Branson Belle (MO)',
        'IEE, Acadex, Higher, IEO',
        'Ride Operator / Food Service / Retail / Show Host',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 22:00 (12 ชม. เทศกาลซัมเมอร์)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$95',
        'สวนสนุกธีมศตวรรษที่ 18 ยุคคาวบอยแร่เงินในเทือกเขา Ozarks และเรือสำราญ Showboat Branson Belle, บัตรเข้าสวนสนุกฟรี',
        '🥩 Branson Landing Waterfront Restaurants: Busser 18:00-23:30 น. ($13/ชม. + ทิปสด $60-$90/คืน)',
        '🍗 Lambert’s Cafe (Home of Throwed Rolls): ผู้ช่วยครัว/เสิร์ฟ 17:30-22:30 น. (คนแน่นมาก ทิปสดดี)',
        '🚢 Showboat Branson Belle: ช่วยจัดเลี้ยงบนเรือล่องทะเลสาบกะค่ำ (เรต OT $22.50)'
    ],
    [
        'Tier A', 'Missouri (MO)',
        'Six Flags St. Louis & Hurricane Harbor (Eureka, MO)',
        'New Step, IEE, Acadex',
        'Ride Operator / Lifeguard / Park Services',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 21:30 (12 ชม.)',
        'ฐาน $14.50 / OT $21.75', '-', '42 – 48 ชม./wk', '$95',
        'สวนสนุกและสวนน้ำยอดนิยมในรัฐมิสซูรี, หอพักพนักงาน, บัตรเล่นเครื่องเล่นฟรี',
        '🍕 Eureka Pizzeria & Pub: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:00 น. ($14.50/ชม.)',
        '🍔 Culver’s Eureka: แคชเชียร์/ครัวกะค่ำ 18:30-23:00 น. ($15/ชม.)',
        '🍦 Dairy Queen Eureka: ตักไอศกรีมรอบดึก'
    ],

    # 21. PENNSYLVANIA (PA)
    [
        'Tier A', 'Pennsylvania (PA)',
        'Hersheypark & Hershey Entertainment Resorts (Hershey, PA)',
        'Higher, Acadex, OEG, IEO, ALC',
        'Ride Operator / Food & Beverage / Housekeeping / Games',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:30 (12.5 ชม. ควงกะ)',
        'ฐาน $15.25 / OT $22.87', '-', '44 – 50 ชม./wk', '$105',
        'เมืองแห่งช็อกโกแลตเฮอร์ชีย์ สวนสนุก Hersheypark & The Hotel Hershey 4 เพชร AAA, หอพักพนักงาน, ช็อกโกแลตและบัตรสวนสนุกฟรี',
        '🥩 The Chocolatier Restaurant & Bar (ในสวนสนุก): Busser/Food Runner 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍫 Hershey’s Chocolate World: แคชเชียร์/จัดของ 18:00-22:30 น. ($15/ชม.)',
        '🍕 Houlihan’s Hershey: Barback/Busser ร้านอาหารใจกลางเมือง 18:30-00:00 น. (ทิปสดแน่น)'
    ],
    [
        'Tier A', 'Pennsylvania (PA)',
        'Kalahari Resorts Pocono Mountains (Pocono Manor, PA)',
        'OEG, Acadex, Higher, ALC',
        'Lifeguard / Housekeeping / Culinary',
        'ปกติ: 09:00 – 17:30 (8 ชม.)\nพีคซัมเมอร์: 08:30 – 19:30 (10.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '42 – 48 ชม./wk', '$110',
        'รีสอร์ตสวนน้ำในร่มขนาดใหญ่ในเทือกเขาโพโคโน แหล่งตากอากาศคนรวยนิวยอร์กและฟิลาเดลเฟีย',
        '🥩 Double Cut Steakhouse (ในรีสอร์ต): Busser 17:30-23:30 น. (ทิปสด $70-$120/คืน)',
        '🍕 Sortino’s Italian Kitchen (ในรีสอร์ต): ล้างจาน/ผู้ช่วยครัว 17:00-22:30 น. (ขอ OT $23.25/ชม.)',
        '🛍️ The Crossings Premium Outlets: จัดสต็อกสินค้า/ปิดร้าน 18:00-21:30 น. ($15.50/ชม.)'
    ],
    [
        'Tier A', 'Pennsylvania (PA)',
        'Dorney Park & Wildwater Kingdom (Allentown, PA)',
        'OEG, IEE, Acadex, IEO',
        'Ride Operator / Lifeguard / Park Services',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 22:00 (12 ชม.)',
        'ฐาน $15.00 / OT $22.50', '-', '42 – 48 ชม./wk', '$95',
        'สวนสนุกและสวนน้ำชั้นนำเครือ Cedar Fair ในเพนซิลเวเนีย, หอพักพนักงาน, บัตรเข้าสวนสนุกฟรี',
        '🥩 Lehigh Valley Mall Eateries: Busser/Runner 18:00-23:00 น. ($14/ชม. + ทิป)',
        '🍕 Allentown Pizza & Grill: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:30-23:30 น. ($14.50/ชม.)',
        '🍦 Rita’s Water Ice Allentown: แคชเชียร์กะค่ำ'
    ],

    # 22. MICHIGAN (MI)
    [
        'Tier A', 'Michigan (MI)',
        'Grand Hotel Mackinac Island (Mackinac Island, MI)',
        'Acadex, Higher, ALC, IEO',
        'Housekeeping / Kitchen Steward / Dining Room / Baggage',
        'ปกติ: 07:30 – 15:30 หรือ 16:00 – 00:00\nพีคซัมเมอร์: 8-10 ชม. ควงกะ',
        'ฐาน $16.00 / OT $24.00', '$30 - $70', '48 – 54 ชม./wk', '$105 (รวมกิน)',
        'โรงแรมประวัติศาสตร์หรูหราระดับตำนานบนเกาะ Mackinac (ห้ามใช้รถยนต์ ใช้จักรยาน/รถม้า 100%), อาหาร 3 มื้อฟรีใน EDR, ทิปสดแน่น',
        '🥩 Main Dining Room (ใน Grand Hotel): Busser อาหารค่ำ 5 คอร์ส 17:30-23:30 น. (ทิปสดเศรษฐี $80-$150/คืน)',
        '🐴 Mackinac Island Carriage Tours: ช่วยดูแลรถม้า/ทำความสะอาด 17:00-21:00 น. ($16/ชม.)',
        '🍫 Ryba’s Fudge Shops: ช่วยทำขนมฟัดจ์/ขายของหน้าร้าน 18:00-22:30 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Michigan (MI)',
        'Mission Point Resort (Mackinac Island, MI - Summer)',
        'Acadex, Higher, ALC',
        'Housekeeping / Culinary / Activities Host',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '44 – 50 ชม./wk', '$110 (รวมกิน)',
        'รีสอร์ตริมทะเลสาบ Lake Huron บนเกาะ Mackinac Island บรรยากาศโรแมนติก, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Chianti Restaurant (ในรีสอร์ต): Busser Fine Dining 17:30-23:00 น. (ทิปสด $70-$120/คืน)',
        '🍕 Round Island Bar & Grill (ในรีสอร์ต): Food Runner/ล้างจาน 17:00-23:00 น. (ขอ OT ในรีสอร์ต $24)',
        '🚲 Mackinac Bike Rentals: ช่วยจัดและซ่อมบำรุงจักรยานเช่าช่วงเช้า 07:30-10:30 น.'
    ],

    # 23. NEW YORK (NY)
    [
        'Tier A', 'New York (NY)',
        'The Sagamore Resort (Lake George / Bolton Landing, NY)',
        'Higher, Acadex, IEO, ALC',
        'Housekeeping / Culinary / Marina / Banquet',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$30 - $70', '44 – 50 ชม./wk', '$125',
        'รีสอร์ตหรูระดับ 4 เพชร AAA บนเกาะส่วนตัวในทะเลสาบ Lake George เทือกเขา Adirondacks แขกมหาเศรษฐีซัมเมอร์',
        '🥩 La Bella Vita Restaurant (ในรีสอร์ต): Busser 17:30-23:30 น. (ทิปสด $80-$150/คืน)',
        '🚢 The Morgan Boat (เรือสำราญรีสอร์ต): Food Runner จัดเลี้ยงบนเรือล่องทะเลสาบ 17:00-22:30 น. (ทิปสดดี)',
        '🍕 Bolton Landing Pizzeria: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:30 น. ($15.50/ชม.)'
    ],
    [
        'Tier A', 'New York (NY)',
        'Six Flags Great Escape & Hurricane Harbor (Queensbury, NY)',
        'New Step, IEE, Acadex, IEO',
        'Ride Operator / Lifeguard / Food Service',
        'ปกติ: 09:30 – 18:00 (8 ชม.)\nพีคซัมเมอร์: 09:00 – 21:30 (12 ชม.)',
        'ฐาน $15.50 / OT $23.25', '-', '42 – 48 ชม./wk', '$105',
        'สวนสนุกและสวนน้ำใกล้ทะเลสาบ Lake George, หอพักพนักงาน Great Escape Lodge, บัตรเข้าสวนสนุกฟรี',
        '🥩 The Log Jam Restaurant (Lake George): Busser/Steward 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍔 Lake George Boardwalk Diners: Food Runner ร้านอาหารริมหาด 18:00-00:00 น. (ทิปดี)',
        '🍦 Martha’s Dandee Creme: พนักงานตักไอศกรีมชื่อดังคิวยาว 18:00-23:00 น. ($15.50/ชม.)'
    ],

    # 24. WASHINGTON (WA)
    [
        'Tier A', 'Washington (WA)',
        'Mount Rainier NP Lodges - Paradise Inn & National Park Inn',
        'Higher, Acadex, IEE, IEO',
        'Housekeeping / Kitchen Steward / Dining Room / Retail',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '44 – 50 ชม./wk', '$105 (รวมกิน)',
        'ปลอดภาษีเงินได้รัฐ 0%, โรงแรมประวัติศาสตร์บนทุ่งดอกไม้ป่าเชิงภูเขาไฟ Mount Rainier, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Paradise Inn Dining Room: Busser/Runner 17:30-22:30 น. (ขอ OT $24.75/ชม. + ทิป)',
        '☕ Tatoosh Cafe (ใน Paradise Inn): บาริสต้า/แคชเชียร์ 16:30-21:30 น. (เรต OT $24.75)',
        '🛒 National Park General Store: จัดสต็อกสินค้า/ปิดร้าน 17:00-21:30 น.'
    ],
    [
        'Tier A', 'Washington (WA)',
        'Olympic National Park Lodges - Lake Quinault & Lake Crescent',
        'Higher, Acadex, ALC',
        'Housekeeping / Culinary / Front Desk',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '42 – 48 ชม./wk', '$110 (รวมกิน)',
        'ปลอดภาษีรัฐ 0%, ลอดจ์ริมทะเลสาบในป่าดงดิบชื้น temperate rainforest อุทยาน Olympic NP, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Roosevelt Dining Room (ใน Lake Quinault Lodge): Busser 17:30-22:30 น. (ทิปสด $50-$90/คืน)',
        '🛶 Lake Crescent Boat Rentals: ช่วยดูแลเรือพาย/แคนู 16:30-20:00 น. (เรต OT)',
        '🌲 Olympic Park Guided Tours: ผู้ช่วยไกด์ทัวร์ธรรมชาติ'
    ],

    # 25. ARIZONA (AZ)
    [
        'Tier A', 'Arizona (AZ)',
        'Grand Canyon National Park South Rim Lodges (Xanterra, AZ)',
        'IEE, Acadex, Higher, IEO, ALC',
        'Housekeeping / Kitchen Steward / Food Runner / Retail',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$20 - $50', '48 – 54 ชม./wk', '$105 (รวมกิน)',
        'โรงแรมริมขอบเหวแกรนด์แคนยอนระดับโลก (El Tovar Hotel, Bright Angel Lodge), อาหาร 3 มื้อฟรีใน EDR',
        '🥩 El Tovar Dining Room: Busser ร้าน Fine Dining ริมหน้าผา 17:30-23:00 น. (ทิปสด $70-$130/คืน)',
        '☕ Bright Angel Fountain & Coffee: ตักไอศกรีม/บาริสต้า 16:30-21:30 น. (เรต OT $23.25)',
        '🛒 Desert View Trading Post: แคชเชียร์/เติมสต็อกกะค่ำ 17:00-21:30 น.'
    ],
    [
        'Tier A', 'Arizona (AZ)',
        'Grand Canyon North Rim Lodge (Forever Resorts - Summer Only)',
        'Higher, Acadex, IEO',
        'Housekeeping / Culinary / Guest Services',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$20 - $40', '46 – 52 ชม./wk', '$100 (รวมกิน)',
        'เปิดเฉพาะซัมเมอร์ (พ.ค.-ต.ค.) ฝั่ง North Rim สงบเงียบและอากาศเย็นสบาย, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Grand Canyon North Rim Dining Room: Busser 17:30-22:30 น. (ขอ OT $23.25 + ทิป)',
        '🥪 Deli in the Pines: ผู้ช่วยทำแซนด์วิช/เบอร์เกอร์ 16:30-21:00 น. (เรต OT)',
        '🐴 Canyon Trail Rides: ช่วยดูแลม้าและทำความสะอาดคอกม้าช่วงเย็น'
    ],

    # 26. NEVADA (NV)
    [
        'Tier A', 'Nevada (NV)',
        'Hyatt Regency Lake Tahoe Resort, Spa & Casino (Incline Village, NV)',
        'Higher, Acadex, ALC',
        'Housekeeping / Culinary / Casino Support / Beach Attendant',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$30 - $70', '42 – 48 ชม./wk', '$135',
        'ปลอดภาษีเงินได้รัฐ 0%, รีสอร์ตคาสิโน 4 เพชร AAA ริมหาดทะเลสาบทาโฮ แขกมหาเศรษฐีซัมเมอร์',
        '🥩 Lone Eagle Grille (ร้านอาหารริมหาดส่วนตัวของไฮแอท): Busser 17:30-23:30 น. (ทิปสด $80-$150/คืน)',
        '🍔 Sierra Cafe & Lakeside Bar: Food Runner ริมทะเลสาบ 18:00-00:00 น. (ทิปสดแน่น)',
        '🛶 Action Water Sports Lake Tahoe: ช่วยดูแลเจ็ทสกี/เรือใบ 16:30-20:30 น. ($16/ชม.)'
    ],

    # 27. CALIFORNIA (CA)
    [
        'Tier A', 'California (CA)',
        'Yosemite National Park Lodges (Aramark / Curry Village / Valley Lodge)',
        'OEG, Higher, Acadex, IEE, IEO, ALC',
        'Housekeeping / Food Service / Retail / Bike Rental',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '-', '44 – 50 ชม./wk', '$105 (รวมกิน)',
        'อุทยานแห่งชาติโยเซมิตี น้ำตกและหน้าผาหินแกรนิตระดับโลก, อาหาร 3 มื้อฟรีใน EDR, รถชัตเติลบัสฟรีทั่วหุบเขา',
        '🥩 The Ahwahnee Dining Room: Busser โรงแรม 4 ดาวระดับตำนาน 17:30-23:00 น. (ทิปสดเศรษฐี $80-$140/คืน)',
        '🍕 Curry Village Pizza Deck: ผู้ช่วยทำพิซซ่า/ล้างจาน 18:00-23:00 น. (ขอ OT $24.75/ชม.)',
        '🍦 Yosemite Valley Lodge Food Court: แคชเชียร์/เสิร์ฟอาหาร 17:00-22:00 น. (เรต OT)'
    ],
    [
        'Tier A', 'California (CA)',
        'Tenaya Lodge at Yosemite (Fish Camp, CA - Summer)',
        'American Learning (ALC), Acadex, Higher',
        'Housekeeping / Culinary / Guest Activities',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $50', '42 – 48 ชม./wk', '$120',
        'รีสอร์ตหรูระดับ 4 เพชร AAA หน้าทางเข้าอุทยาน Yosemite NP ฝั่งทิศใต้ (Mariposa Grove ต้นสนยักษ์ Sequoia)',
        '🥩 Jackalope’s Bar & Grill (ในรีสอร์ต): Busser/Barback 17:30-23:30 น. (ทิปสด $60-$110/คืน)',
        '🍕 Timberloft Pizzeria (ในรีสอร์ต): ล้างจาน/ทำพิซซ่า 17:00-22:30 น. (ขอ OT $24.75/ชม.)',
        '🌲 Yosemite Mountain Sugar Pine Railroad: ช่วยดูแลสถานีรถไฟโบราณ 16:30-20:30 น.'
    ],

    # 28. DELAWARE (DE)
    [
        'Tier A', 'Delaware (DE)',
        'Rehoboth Beach Boardwalk Hotels & Oceanfront Resorts',
        'Higher, Acadex, New Step, IEO',
        'Housekeeping / Front Desk / Retail',
        'ปกติ: 08:30 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $15.00 / OT $22.50', '$20 - $40', '40 – 46 ชม./wk', '$120',
        'ปลอดภาษีซื้อของ 0% Sales Tax (Tax-Free Capital), เมืองตากอากาศชายหาดคนรวยวอชิงตัน ดี.ซี. ซื้อของช้อปปิ้งถูกที่สุด',
        '🍕 Grotto Pizza (สาขา The Grand Slam Rehoboth): ผู้ช่วยทำพิซซ่า/ล้างจาน 18:30-01:00 น. ($14.50/ชม. + ทิป)',
        '🦀 Crab House / The Starboard: Busser/Barback ร้านปูและบาร์ชื่อดัง 18:00-00:30 น. (ทิปสด $70-$120/คืน)',
        '🎡 Funland Rehoboth Beach: คุมเครื่องเล่น/เกมงานวัดริมบอร์ดวอล์ก 18:00-23:00 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Delaware (DE)',
        'Grotto Pizza (Rehoboth Beach Boardwalk & Dewey Beach, DE)',
        'Higher, Acadex, New Step',
        'Pizza Prep / Line Cook / Busser / Cashier',
        'ปกติ: 11:30 – 19:30 (8 ชม.)\nพีคซัมเมอร์: 11:00 – 00:00 (13 ชม.)',
        'ฐาน $14.50 / OT $21.75', '$40 - $90', '42 – 48 ชม./wk', '$115',
        'ร้านพิซซ่าซิกเนเจอร์อันดับ 1 ในรัฐเดลาแวร์ ปลอดภาษีซื้อของ 0%, ทิปสดแน่นมากตลอดซัมเมอร์',
        '🧹 กะเช้าช่วยแม่บ้านโรงแรมริมหาด (Boardwalk Plaza / Avenue Inn): 08:30-13:30 น. ($15/ชม.)',
        '🎡 Funland Amusements: ช่วยคุมซุ้มเกมริมบอร์ดวอล์ก 19:00-23:00 น.',
        '🍦 Kohr Bros Rehoboth: พนักงานตักไอศกรีม'
    ],

    # 29. RHODE ISLAND (RI)
    [
        'Tier A', 'Rhode Island (RI)',
        'Block Island Historic Hotels & Inns (The National Hotel, RI)',
        'Higher, Acadex, IEO',
        'Housekeeping / Food Runner / Busser / Front Desk',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:30 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$30 - $70', '42 – 48 ชม./wk', '$120',
        'เกาะตากอากาศประวัติศาสตร์นอกชายฝั่งเกาะโรดไอแลนด์ (เดินทางด้วยเรือเฟอร์รี่เท่านั้น), แขกเรือยอชต์ ทิปสดแน่น',
        '🥩 The National Hotel Tap & Grill: Busser/Runner หน้าระเบียงชมวิวทะเล 17:30-23:30 น. (ทิปสด $80-$140/คืน)',
        '🦞 The Oar Restaurant: Barback ร้านอาหารและบาร์ริมท่าเรือ 18:00-00:30 น. (ทิปสดดีมาก)',
        '🚲 Block Island Bike & Moped Rentals: ช่วยดูแลจักรยานเช่ารอบเช้า 07:30-10:00 น.'
    ],

    # 30. OREGON (OR)
    [
        'Tier A', 'Oregon (OR)',
        'Crater Lake National Park Lodges (Crater Lake, OR)',
        'Higher, Acadex, IEE, IEO',
        'Housekeeping / Culinary / Retail / Front Desk',
        'ปกติ: 07:30 – 16:00 (8 ชม.)\nพีคซัมเมอร์: 07:00 – 17:00 (9.5 ชม.)',
        'ฐาน $15.50 / OT $23.25', '$15 - $30', '44 – 50 ชม./wk', '$105 (รวมกิน)',
        'ปลอดภาษีซื้อของ 0% Sales Tax, ทะเลสาบปล่องภูเขาไฟที่ลึกที่สุดและน้ำสีน้ำเงินใสที่สุดในอเมริกา, อาหาร 3 มื้อฟรีใน EDR',
        '🥩 Crater Lake Lodge Dining Room: Busser 17:30-22:30 น. (ขอ OT $23.25/ชม. + ทิป)',
        '🥪 Annie Creek Restaurant (ริมแคมป์): ผู้ช่วยครัว/ล้างจาน 16:30-21:30 น. (เรต OT)',
        '🚢 Crater Lake Volcano Boat Cruises: ช่วยดูแลท่าเรือและล้างเรือทัวร์ 16:00-19:30 น.'
    ],
    [
        'Tier A', 'Oregon (OR)',
        'Timberline Lodge & Ski Area (Mount Hood, OR)',
        'Higher, Acadex, ALC',
        'Housekeeping / Culinary / Lift Operations',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $40', '42 – 48 ชม./wk', '$115 (รวมกิน)',
        'ลอดจ์ประวัติศาสตร์แห่งชาติบนยอดภูเขาไฟ Mount Hood สกีซัมเมอร์แห่งเดียวในอเมริกา ปลอดภาษี 0%, อาหาร 3 มื้อฟรี',
        '🥩 Cascade Dining Room (ใน Timberline Lodge): Busser 17:30-23:00 น. (ทิปสด $60-$100/คืน)',
        '🍺 Blue Ox Bar & Wy’East Cafe: ผู้ช่วยครัว/Barback 16:30-22:00 น. (ขอ OT $24/ชม.)',
        '🚡 Magic Mile Chairlift: ช่วยดูแลผู้โดยสารกระเช้าชมวิวรอบสาย'
    ],

    # 31. VERMONT (VT)
    [
        'Tier A', 'Vermont (VT)',
        'Stowe Mountain Resort & Spruce Peak (Vail Resorts, VT)',
        'OEG, Higher, Acadex, ALC',
        'Housekeeping / Mountain Host / Culinary / Retail',
        'ปกติ: 08:30 – 17:00 (8 ชม.)\nพีคซัมเมอร์: 08:00 – 17:30 (9.5 ชม.)',
        'ฐาน $16.50 / OT $24.75', '$20 - $40', '40 – 46 ชม./wk', '$130',
        'สกีรีสอร์ตและหมู่บ้าน Spruce Peak สุดหรูในเทือกเขา Green Mountains ทิวทัศน์ซัมเมอร์สวยงาม กิจกรรมซิปไลน์',
        '🥩 Solstice Restaurant & Alpine Lounge: Busser 17:30-23:00 น. (ทิปสด $70-$120/คืน)',
        '🍺 The Matterhorn Bar & Grill (Stowe): Barback/Dishwasher บาร์คึกคัก 18:30-00:30 น. (ทิปสดแน่น)',
        '🥐 The Skinny Pancake Stowe: ทำเครป/แคชเชียร์ 17:00-22:00 น. ($15/ชม.)'
    ],
    [
        'Tier A', 'Vermont (VT)',
        'Trapp Family Lodge & von Trapp Brewing (Stowe, VT)',
        'Higher, Acadex, IEO',
        'Housekeeping / Brewery Dining / Guest Services',
        'ปกติ: 08:00 – 16:30 (8 ชม.)\nพีคซัมเมอร์: 07:30 – 17:00 (9.5 ชม.)',
        'ฐาน $16.00 / OT $24.00', '$20 - $50', '42 – 48 ชม./wk', '$110 (รวมกิน)',
        'รีสอร์ตสไตล์เทือกเขาแอลป์ออสเตรียของครอบครัว The Sound of Music และโรงเบียร์คราฟต์ชื่อดัง, อาหาร 3 มื้อฟรีใน EDR',
        '🍺 von Trapp Brewery & Bierhall: Busser/Runner ลานเบียร์ออสเตรีย 17:30-23:30 น. (ทิปสด $70-$120/คืน)',
        '🥩 Main Dining Room (ในลอดจ์): Food Runner อาหารยุโรป 17:00-22:30 น. (ขอ OT $24/ชม. + ทิป)',
        '🥐 Trapp Family Bakery: ช่วยทำขนมเบเกอรี่ออสเตรียช่วงเช้าตรู่ 06:30-10:00 น.'
    ]
]

# Sheet 1: Top Employers (Summer Only) - 11 Clean Columns Overview
ws1 = wb.active
ws1.title = 'Top Employers (Summer Only)'
ws1.views.sheetView[0].showGridLines = True

headers_sheet1 = [
    'ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)',
    'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น'
]
ws1.append(headers_sheet1)

for r_idx, job in enumerate(master_summer_jobs, start=2):
    row_data = job[:11]
    ws1.append(row_data)

# Sheet 2: Tier S-A Summer Jobs - 14 Columns with 2nd Job Options
ws2 = wb.create_sheet(title='Tier S-A Summer Jobs')
ws2.views.sheetView[0].showGridLines = True

headers_sheet2 = [
    'ระดับ Tier', 'รัฐ (State)', 'ชื่องาน / สถานที่ทำงาน', 'Agency ในไทย', 'ตำแหน่งงาน (ซัมเมอร์)',
    'ช่วงเวลาทำงาน (พีค/ปกติ)', 'ค่าแรงฐาน / OT', 'ทิป ($/wk)', 'ชั่วโมงทำงานจริง', 'ค่าที่พัก ($/wk)', 'สวัสดิการ & จุดเด่น',
    'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 1)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 2)', 'ตัวเลือกงานที่ 2 (กะค่ำ / ร้านที่ 3)'
]
ws2.append(headers_sheet2)

for r_idx, job in enumerate(master_summer_jobs, start=2):
    ws2.append(job)

# Sheet 3: Summer Agency Directory
ws3 = wb.create_sheet(title='Summer Agency Directory')
ws3.views.sheetView[0].showGridLines = True

agency_headers = [
    'ชื่อเอเจนซี่ (Agency)', 'ค่าโครงการโดยประมาณ', 'Sponsor หลักในสหรัฐฯ', 'จุดเด่นสำหรับเด็ก Summer (พ.ค.-ก.ย.)',
    'รัฐเด่น / งานยอดฮิตช่วง Summer', 'เบอร์ติดต่อ', 'LINE Official', 'Facebook / Website', 'ที่ตั้งออฟฟิศในไทย'
]
ws3.append(agency_headers)

agency_directory = [
    [
        'OEG (Overseas Ed Group)', '65,000 – 78,000 บาท', 'CIEE / Spirit Cultural Exchange',
        'เอเจนซี่ที่ใหญ่และเก่าแก่ที่สุดในไทย โควตางานอุทยานแห่งชาติและสวนสนุกยักษ์ใหญ่เยอะที่สุด',
        'Alaska (Denali), Wyoming (Yellowstone/Teton), Ohio (Cedar Point), Wisconsin (Kalahari)',
        '02-263-3666', '@oeg_workandtravel', 'facebook.com/OEGWorkAndTravel', 'อาคาร ซีพี ทาวเวอร์ (สีลม) ชั้น 11 กรุงเทพฯ'
    ],
    [
        'Acadex Thailand', '62,000 – 75,000 บาท', 'Intrax / CCUSA / CHI / IENA',
        'ระบบสัมภาษณ์งานออนไลน์เสถียร มีงานรัฐ Tier S หลากหลาย ดูแลเอกสารวีซ่าละเอียดมาก',
        'Alaska (Denali/Grande Denali/Talkeetna), Wyoming (Grand Teton/Yellowstone), Wisconsin (Dells), Michigan (Grand Hotel)',
        '02-129-3547', '@acadex', 'facebook.com/acadexthailand', 'อาคาร ทู แปซิฟิค เพลส ชั้น 18 (BTS นานา) กรุงเทพฯ'
    ],
    [
        'IEE Thailand', '59,000 – 72,000 บาท', 'CCUSA / InterExchange / GeoVisions',
        'ค่าโครงการสมเหตุสมผล โดดเด่นเรื่องงานอุทยาน Xanterra และงานสวนน้ำ/สวนสนุก',
        'Wyoming (Xanterra Yellowstone), South Dakota (Mount Rushmore), Wisconsin (Chula Vista/Noah’s Ark), New Jersey (Morey’s)',
        '02-612-9511', '@ieethailand', 'facebook.com/ieethailand', 'อาคาร พญาไทพลาซ่า ชั้น 15 (BTS พญาไท) กรุงเทพฯ'
    ],
    [
        'IEO Abroad', '60,000 – 74,000 บาท', 'CIEE / Intrax / InterExchange',
        'ให้คำปรึกษาเป็นกันเอง เหมาะสำหรับผู้ที่ต้องการเลือกงานแบบคัดกรองเมืองและเรตค่าแรง',
        'Alaska (Denali Princess), Wisconsin (Kalahari), Tennessee (Pigeon Forge), Florida (Universal Orlando)',
        '02-650-3532', '@ieoabroad', 'facebook.com/ieoabroad', 'อาคาร อรกานต์ ชั้น 16 (ชิดลม) กรุงเทพฯ'
    ],
    [
        'New Step Thailand', '58,000 – 70,000 บาท', 'Intrax / CHI / AWA',
        'ชำนาญพื้นที่ฝั่งตะวันออกและสวนสนุก/สวนน้ำขนาดใหญ่ มีระบบแบ่งจ่ายค่าโครงการเป็นงวด',
        'Tennessee (Dollywood/The Island), Maryland (Ocean City), New Jersey (Wildwood Morey’s), Virginia (Busch Gardens)',
        '02-246-0430', '@newstep', 'facebook.com/newstepthailand', 'อาคาร ฟอร์จูนทาวน์ ชั้น 16 (พระราม 9) กรุงเทพฯ'
    ],
    [
        'Higher Education', '63,000 – 76,000 บาท', 'IENA / Spirit / Janus International',
        'โดดเด่นมากเรื่องงานรีสอร์ตพรีเมียม 4-5 ดาว, แกลมปิ้ง Under Canvas และอุทยานแห่งชาติ',
        'Wyoming (Grand Teton GTLC/Jackson Hole), Montana (Glacier NP/Big Sky), Alaska (Pursuit/Princess), New Hampshire (Omni)',
        '02-530-9111', '@higher', 'facebook.com/higherthailand', 'อาคาร เมเจอร์ ทาวเวอร์ ทองหล่อ ชั้น 10 กรุงเทพฯ'
    ],
    [
        'American Learning (ALC)', '60,000 – 73,000 บาท', 'Intrax / Spirit / CCUSA / Premier',
        'เจ้าใหญ่ด้านตำแหน่ง Pool Lifeguard รายได้สูงสุดในแมริแลนด์/เวอร์จิเนีย และงานรีสอร์ตหรู',
        'Maryland/VA/DC (Premier Aquatics Lifeguard), Michigan (Grand Hotel), California (Tenaya Yosemite), Pennsylvania (Kalahari)',
        '02-642-4520', '@americanlearning', 'facebook.com/americanlearning', 'อาคาร ซี.พี.ทาวเวอร์ 2 (ฟอร์จูน พระราม 9) ชั้น 19 กรุงเทพฯ'
    ]
]

for row in agency_directory:
    ws3.append(row)

# Formatting all sheets
for ws in [ws1, ws2, ws3]:
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        tier_val = str(ws.cell(row=r, column=1).value)
        ws.row_dimensions[r].height = 42 if ws != ws3 else 28

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border

            if 'Tier S' in tier_val:
                if c == 1:
                    cell.fill = tier_s_fill
                    cell.font = bold_font
            elif 'Tier A' in tier_val:
                if c == 1:
                    cell.fill = tier_a_fill
                    cell.font = bold_font

            if ws in [ws1, ws2]:
                if c in [1, 2, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif c in [6, 7]:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths_s1 = [12, 18, 38, 28, 30, 32, 24, 14, 18, 16, 45]
col_widths_s2 = [12, 18, 38, 28, 30, 32, 24, 14, 18, 16, 45, 42, 42, 42]
col_widths_s3 = [25, 20, 28, 42, 38, 16, 20, 30, 35]

for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for i, w in enumerate(col_widths_s3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

try:
    wb.save(excel_path)
    print("Successfully saved unified Master Excel to:", excel_path)
except Exception as e:
    print("Locked path:", excel_path)

try:
    wb.save(excel_clean_path)
    print("Successfully saved unified Master Excel to:", excel_clean_path)
except Exception as e:
    print("Error saving clean path:", e)
